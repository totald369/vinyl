#!/usr/bin/env python3
"""
부산 사상구 종량제·불연성(및 비고 건설마대) 판매소 xlsx
시트: 종량제봉투판매소 | 불연성마대 및 건설폐기물 마대 판매소

  python3 scripts/import_busan_sasang_from_xlsx.py \\
    --xlsx ~/Downloads/'종량제봉투판매소현황(2026.4.16.기준홈페이지용).xlsx'
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
OUT_TRASH = FRONTEND / "public" / "data" / "stores.busan-sasang-trash.json"
OUT_SPECIAL = FRONTEND / "public" / "data" / "stores.busan-sasang-special.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-busan-sasang.json"

USER_AGENT = "Mozilla/5.0 (compatible; VinylMapImport/1.0; +https://github.com/totald369/vinyl)"
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=12) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=12) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def collapse(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def full_busan_sasang(raw: str) -> str:
    s = collapse(raw).replace("\u200b", "")
    if not s:
        return s
    if s.startswith("부산광역시"):
        return s
    if s.startswith("부산시"):
        return collapse("부산광역시 " + s[3:].lstrip())
    if s.startswith("사상구"):
        return collapse("부산광역시 " + s)
    if s.startswith("부산 "):
        return collapse("부산광역시 " + s[3:].lstrip())
    return collapse(f"부산광역시 사상구 {s}")


def fmt_phone(raw: str | None) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    s = collapse(str(raw))
    if re.match(r"^051-\d{3}-\d{4}$", s) or re.match(r"^051-\d{4}-\d{4}$", s):
        return s
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if digits.startswith("051") and len(digits) == 11:
        return f"051-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 10 and digits.startswith("051"):
        return f"051-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 7:
        return f"051-{digits[:3]}-{digits[3:]}"
    return s


def geocode_query_variants(full_addr: str) -> list[str]:
    def _tail_paren_strip(s: str) -> str:
        return collapse(re.sub(r"\s*[\(（][^)）]*[\)）]\s*$", "", s))

    def add(q: str, out: list[str]) -> None:
        cq = collapse(q)
        if cq and cq not in out:
            out.append(cq)

    seen: list[str] = []
    collapsed = collapse(full_addr)
    comma_head = collapsed.split(",", 1)[0].strip()
    stripped_tail = _tail_paren_strip(collapsed)
    for q in (collapsed, comma_head, stripped_tail, _tail_paren_strip(comma_head)):
        add(q, seen)
    base = comma_head or stripped_tail or collapsed
    spaced_alley = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced_alley != base:
        add(spaced_alley, seen)
        base = spaced_alley
    ro_gil = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", base)
    if ro_gil != base:
        add(ro_gil, seen)
        base = ro_gil
    ro_only = re.sub(
        r"^(부산광역시\s+사상구\s+[가-힣\d\s]+?(?:로|대로))\s+\d.+",
        r"\1",
        base,
    )
    if ro_only != base:
        add(ro_only, seen)
    return seen


def resolve_coords(
    road_address: str, place_name: str, cache: dict, key: str
) -> tuple[float, float] | None:
    for qv in geocode_query_variants(road_address):
        c = kakao_geocode(qv, cache, key)
        if c:
            return c
    c = kakao_geocode(f"{road_address} {place_name}", cache, key)
    if c:
        return c
    for qv in geocode_query_variants(road_address):
        c = kakao_keyword_geocode(qv, cache, key)
        if c:
            return c
    name_bits = {place_name}
    if "(" in place_name:
        name_bits.add(re.sub(r"\([^)]*\)", "", place_name).strip())
    for base in name_bits:
        for q in (
            f"부산 사상구 {base}",
            f"사상구 {base}",
            f"부산 {base}",
            base,
        ):
            if len(q.strip()) < 2:
                continue
            c = kakao_keyword_geocode(q, cache, key)
            if c:
                return c
    return None


def trash_id(mgmt_raw: object, seq: int) -> str:
    if mgmt_raw is not None and str(mgmt_raw).strip():
        s = collapse(str(mgmt_raw)).replace("/", "-").replace("\\", "-")
        s = re.sub(r'\s+', "-", s)
        return f"busan-sasang-trash-{s}"
    return f"busan-sasang-trash-seq-{int(seq)}"


def parse_trash_sheet(path: Path) -> list[tuple[int, object, str, str]]:
    """(연번, 관리번호, 상호, 주소 원문)"""
    wb = load_workbook(path, read_only=True, data_only=True)
    sh = wb["종량제봉투판매소"]
    out = []
    for r in range(3, sh.max_row + 1):
        seq = sh.cell(r, 1).value
        mgmt = sh.cell(r, 2).value
        name = sh.cell(r, 3).value
        addr = sh.cell(r, 4).value
        nm = collapse(str(name or ""))
        ad = collapse(str(addr or ""))
        if not nm or not ad:
            continue
        si = (
            int(float(str(seq))) if seq is not None and str(seq).strip() else r
        )
        out.append((si, mgmt, nm, ad))
    wb.close()
    return out


def parse_special_sheet(path: Path) -> list[tuple[int, str, str, str, str | None]]:
    """연번, 상호, 주소, 전화, 비고"""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "불연성마대 및 건설폐기물 마대 판매소" in wb.sheetnames:
        sh_name = "불연성마대 및 건설폐기물 마대 판매소"
    else:
        cand = [
            x
            for x in wb.sheetnames
            if "불연성" in x and ("건설" in x or "마대" in x)
        ]
        sh_name = cand[0] if cand else wb.sheetnames[1]
    sh = wb[sh_name]
    out = []
    for r in range(2, sh.max_row + 1):
        seq_raw = sh.cell(r, 1).value
        name = collapse(str(sh.cell(r, 2).value or ""))
        addr = collapse(str(sh.cell(r, 3).value or ""))
        phone_raw = sh.cell(r, 4).value
        note = sh.cell(r, 5).value
        if not name or not addr:
            continue
        try:
            si = int(float(str(seq_raw))) if seq_raw is not None else r
        except ValueError:
            si = r
        out.append((si, name, addr, str(phone_raw or ""), note))
    wb.close()
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xls",
        type=Path,
        dest="xlsx",
        default=Path.home()
        / "Downloads"
        / "종량제봉투판매소현황(2026.4.16.기준홈페이지용).xlsx",
        help=".xlsx 경로",
    )
    ap.add_argument(
        "--ref-date",
        default="2026-04-16",
        help="기준일 dataReferenceDate",
    )
    args = ap.parse_args()
    xp = args.xlsx.expanduser()
    if not xp.exists():
        print(f"파일 없음: {xp}", file=sys.stderr)
        sys.exit(1)

    ref_date = args.ref_date.strip()

    trash_rows = parse_trash_sheet(xp)
    spec_rows = parse_special_sheet(xp)
    print(
        f"파싱: 종량제 {len(trash_rows)}건, 불연성/건설 마대 {len(spec_rows)}건 · ref={ref_date}",
        file=sys.stderr,
    )

    cache = load_cache()
    if not KAKAO_REST_KEY:
        print("KAKAO_REST_API_KEY 없음.", file=sys.stderr)

    trash_stores: list[dict] = []
    failed_t: list[str] = []

    for i, (seq, mgmt, name, addr_raw) in enumerate(trash_rows):
        road = full_busan_sasang(addr_raw)
        oid = trash_id(mgmt, seq)
        lat = lng = None
        if KAKAO_REST_KEY:
            c = resolve_coords(road, name, cache, KAKAO_REST_KEY)
            if c:
                lat, lng = c
        if lat is None:
            failed_t.append(f"{oid} {name} | {road}")
        row: dict = {
            "id": oid,
            "name": name,
            "roadAddress": road,
            "address": road,
            "businessStatus": "영업",
            "hasTrashBag": True,
            "hasSpecialBag": False,
            "hasLargeWasteSticker": False,
            "adminVerified": False,
            "dataReferenceDate": ref_date,
        }
        if lat is not None and lng is not None:
            row["lat"] = lat
            row["lng"] = lng
        trash_stores.append(row)
        if (i + 1) % 80 == 0:
            save_cache(cache)
            print(f"  [trash] {i+1}/{len(trash_rows)}", file=sys.stderr)

    spec_stores: list[dict] = []
    failed_s: list[str] = []

    for i, (seq, name, addr_raw, phone_raw, note) in enumerate(spec_rows):
        road = full_busan_sasang(addr_raw)
        oid = f"busan-sasang-special-{seq}"
        note_s = collapse(str(note or ""))
        건설 = "건설" in note_s
        lat = lng = None
        if KAKAO_REST_KEY:
            c = resolve_coords(road, name, cache, KAKAO_REST_KEY)
            if c:
                lat, lng = c
        if lat is None:
            failed_s.append(f"{oid} {name} | {road}")
        row: dict = {
            "id": oid,
            "name": name,
            "roadAddress": road,
            "address": road,
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": 건설,
            "adminVerified": False,
            "dataReferenceDate": ref_date,
        }
        fn = fmt_phone(phone_raw)
        if fn and fn not in ("051-0-", "051-0000-0000", "0"):
            row["phone"] = fn
        if lat is not None and lng is not None:
            row["lat"] = lat
            row["lng"] = lng
        spec_stores.append(row)
        if (i + 1) % 15 == 0:
            save_cache(cache)
            print(f"  [special] {i+1}/{len(spec_rows)}", file=sys.stderr)

    save_cache(cache)

    print(
        f"종량제 좌표 {sum(1 for s in trash_stores if 'lat' in s)}/{len(trash_stores)} "
        f"실패 {len(failed_t)}",
        file=sys.stderr,
    )
    for ln in failed_t[:20]:
        print(f"  {ln}", file=sys.stderr)
    print(
        f"불연성·건설좌표 {sum(1 for s in spec_stores if 'lat' in s)}/{len(spec_stores)} "
        f"실패 {len(failed_s)}",
        file=sys.stderr,
    )
    for ln in failed_s:
        print(f"  {ln}", file=sys.stderr)

    OUT_TRASH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_TRASH, "w", encoding="utf-8") as f:
        json.dump(trash_stores, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(OUT_SPECIAL, "w", encoding="utf-8") as f:
        json.dump(spec_stores, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"저장: {OUT_TRASH}", file=sys.stderr)
    print(f"저장: {OUT_SPECIAL}", file=sys.stderr)


if __name__ == "__main__":
    main()
