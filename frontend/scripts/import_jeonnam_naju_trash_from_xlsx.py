#!/usr/bin/env python3
"""
전라남도 나주시 종량제봉투·불연성마대 판매소 xlsx → stores.jeonnam-naju-trash.json

시트 `Sheet1` (헤더 2행, 데이터 3행~):
  연번(A) | 읍면동(B) | 상호(C) | 주소(D) | 불연성마대(E)

매핑:
  - 목록 등재 -> hasTrashBag: true
  - E열 'O'/'○' -> hasSpecialBag

사용:
  cd frontend
  python3 scripts/import_jeonnam_naju_trash_from_xlsx.py \\
    --input ~/Downloads/정보공개자료\\(나주시_16707870\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonnam-naju-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonnam-naju-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "정보공개자료(나주시_16707870).xlsx"
SHEET_NAME = "Sheet1"
DATA_START_ROW = 3
REF_DATE = "2026-06-09"
CACHE_VERSION = "v1-naju"

COL_AREA = 2
COL_NAME = 3
COL_ADDR = 4
COL_SPECIAL = 5

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")
NAJU_AREAS = (
    "남평읍",
    "세지면",
    "왕곡면",
    "공산면",
    "동강면",
    "다시면",
    "문평면",
    "노안면",
    "금천면",
    "산포면",
    "다도면",
    "봉황면",
    "반남면",
    "금남동",
    "빛가람동",
    "송월동",
    "성북동",
    "영강동",
    "영산동",
    "이창동",
    "교동",
    "안창동",
)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    t = collapse(str(val or "")).upper()
    return t in ("O", "○", "Y", "YES", "예", "여")


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def ref_date_from_path(p: Path) -> str:
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def in_naju_bbox(lat: float, lng: float) -> bool:
    return 34.90 <= lat <= 35.10 and 126.54 <= lng <= 126.86


def naju_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if any(x in t for x in ("서울", "대구", "부산", "인천", "대전", "울산", "화순군")):
        return False
    if "나주시" in t:
        return True
    return ("전남" in t or "전라남도" in t) and "나주" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^전남\s+", "전라남도 ", a)
    if not a:
        return ""
    if a.startswith("전라남도"):
        return a
    if a.startswith("나주시"):
        return f"전라남도 {a}"
    return f"전라남도 나주시 {a}"


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def normalize_naju_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if not a:
        return ""
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"([가-힣]+(?:리|동))(\d+)", r"\1 \2", a)
    a = re.sub(r"((?:대로|로|길))(\d+(?:-\d+)?)", r"\1 \2", a)
    if not a.startswith("전") and not a.startswith("나주"):
        a = f"나주시 {a}"
    return format_display_addr(a)


def geocode_target(full: str) -> str:
    a = collapse(full)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+(?:,\s*\d+)?\s*호.*$", "", a)
    a = re.sub(r"\s+상가동.*$", "", a)
    return a


def road_format_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(addr)
    compact = addr.replace(" ", "")
    push(compact)
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    push(re.sub(r"((?:로|길|대로))(\d+)", r"\1 \2", compact))
    return out


def naju_tail(full: str) -> str:
    for prefix in (
        "전라남도 나주시 ",
        "전라남도 ",
        "전남 나주시 ",
        "나주시 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def lot_query_variants(tail: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = LOT_RE.match(tail)
    if m:
        prefix, main, sub = m.group("prefix"), m.group("main"), m.group("sub")
        subs: list[str | None] = [sub, "1", "2", "3", "4", "5"] if sub else [None]
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def geocode_query_variants(road_full: str, name: str, area: str) -> list[str]:
    target = geocode_target(road_full)
    tail = naju_tail(target)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", target))
    if has_road:
        for base in road_format_variants(target):
            t = naju_tail(base)
            push(f"전남 나주시 {t}")
            push(f"전라남도 나주시 {t}")
    elif is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push(f"전남 나주시 {tv}")
            push(f"전라남도 나주시 {tv}")
    else:
        for base in road_format_variants(target):
            t = naju_tail(base)
            push(f"전남 나주시 {t}")
            push(f"전라남도 나주시 {t}")

    m = re.match(r"^(.+?(?:동|읍|면|리))", tail.replace(" ", ""))
    if m:
        push(f"전남 나주시 {collapse(m.group(1))}")
    if area:
        push(f"전남 나주시 {area}")

    push(f"{name} 나주")
    push(f"{name} 나주시")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class NajuRow:
    name: str
    addr_raw: str
    area: str
    has_special: bool


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_naju_bbox(lat, lng):
        return None
    if not naju_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_naju_bbox(lat, lng):
        return None
    if not naju_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(addr_raw: str, area: str, key: str, display_road: str) -> GeoHit | None:
    tail = naju_tail(geocode_target(addr_raw))
    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    candidate = collapse(m.group(1)) if m else area
    if not candidate:
        for a in NAJU_AREAS:
            if a in tail:
                candidate = a
                break
    if not candidate:
        return None
    for q in (f"전남 나주시 {candidate}", f"전라남도 나주시 {candidate}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, _ = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, area: str, key: str, display_road: str) -> GeoHit | None:
    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", display_road))
    for q in geocode_query_variants(display_road, name, area):
        if has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
        if not has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return area_fallback(addr_raw, area, key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[NajuRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    merged: dict[str, NajuRow] = {}

    def key_for(name: str, addr_raw: str) -> str:
        norm = normalize_naju_addr(addr_raw)
        road = geocode_target(norm)
        return f"{name}|{road}"

    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, COL_NAME).value)
        if not name or name in ("상호",):
            continue
        addr = cell_str(ws.cell(r, COL_ADDR).value)
        if not addr:
            continue
        if "나주" not in addr.replace(" ", ""):
            continue
        area = cell_str(ws.cell(r, COL_AREA).value)
        has_special = flag_y(ws.cell(r, COL_SPECIAL).value)
        k = key_for(name, addr)
        if k not in merged:
            merged[k] = NajuRow(name, addr, area, has_special)
        elif has_special:
            merged[k].has_special = True

    wb.close()
    return list(merged.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    seen: set[str] = set()

    for row in rows:
        full_norm = normalize_naju_addr(row.addr_raw)
        display_road = geocode_target(full_norm)
        if not display_road or "나주" not in display_road.replace(" ", ""):
            continue

        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, row.area, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = cr or display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 20 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"jeonnam-naju-trash-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road or display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref_date,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    trash_cnt = sum(1 for r in out if r["hasTrashBag"])
    special_cnt = sum(1 for r in out if r["hasSpecialBag"])
    both_cnt = sum(1 for r in out if r["hasTrashBag"] and r["hasSpecialBag"])
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo={geo_n}, miss={misses}, src={len(rows)}, "
        f"종량제 {trash_cnt}, 불연성마대 {special_cnt}, 겸업 {both_cnt})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["나주시"])


if __name__ == "__main__":
    main()
