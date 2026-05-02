#!/usr/bin/env python3
"""
대구광역시 북구(garbagebag_vendor.xlsx, 시트=동별) + 달서구(CSV 지정판매소) 종량제봉투 판매처.

  python3 scripts/import_daegu_buk_dalseo_trash.py \\
    --xlsx ~/Downloads/garbagebag_vendor.xlsx \\
    --csv \"$HOME/Downloads/대구광역시 달서구_종량제봉투_지정판매소정보_20240911.csv\"

필요: .env.local KAKAO_REST_KEY
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.daegu-buk-dalseo-trash.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-daegu-buk-dalseo-trash.json"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06

PRI_XLSX = 10
PRI_CSV = 20


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)


def collapse(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("\u200b", "").replace("\xa0", " ")).strip()


def to_float(val):
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        rr = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(rr, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def full_daegu_bukgu(addr_raw: str) -> str:
    s = collapse(str(addr_raw or ""))
    if not s:
        return s
    if "대구광역시" in s:
        return collapse(s)
    if s.startswith("달서구"):
        return collapse(f"대구광역시 {s}")
    if s.startswith("북구"):
        return collapse(f"대구광역시 {s}")
    return collapse(f"대구광역시 북구 {s}")


def full_daegu_dalseo(addr_raw: str) -> str:
    s = collapse(str(addr_raw or ""))
    if not s:
        return s
    s = re.sub(r"달서구([가-힣])", r"달서구 \1", s)
    if "대구광역시" in s:
        return collapse(s)
    return collapse(f"대구광역시 달서구 {s}")


def massage_for_geocode(road: str) -> str:
    s = collapse(road)
    s = re.sub(r"(로)(\d+)(번길|길)", r"\1 \2\3", s)
    s = re.sub(r"(로)(\d+길)", r"\1 \2", s)
    return collapse(s)


def fmt_phone(raw: object) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    s = collapse(str(raw))
    if re.match(r"^053-\d{3,4}-\d{4}$", s):
        return s
    if re.match(r"^054-\d{3,4}-\d{4}$", s):
        return s
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if digits.startswith("053") and len(digits) == 11:
        return f"053-{digits[3:7]}-{digits[7:]}"
    if digits.startswith("054") and len(digits) == 11:
        return f"054-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10 and digits.startswith(("053", "054")):
        prefix = digits[:3]
        return f"{prefix}-{digits[3:7]}-{digits[7:]}"
    if len(digits) in (12, 13) and "-" not in s:
        return ""
    return s[:40] if digits else ""


def geocode_query_variants(full_addr: str) -> list[str]:
    def tail_paren_strip(x: str) -> str:
        return collapse(re.sub(r"\s*[\(（][^)）]*[\)）]\s*$", "", x))

    def add(q: str, out: list[str]) -> None:
        cq = collapse(q)
        if cq and cq not in out:
            out.append(cq)

    seen: list[str] = []
    collapsed = collapse(full_addr)
    comma_head = collapsed.split(",", 1)[0].strip()
    stripped = tail_paren_strip(collapsed)
    for q in (collapsed, comma_head, stripped, tail_paren_strip(comma_head)):
        add(q, seen)

    base = comma_head or stripped or collapsed
    spaced = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced != base:
        add(spaced, seen)
        base = spaced
    ro_join = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", base)
    if ro_join != base:
        add(ro_join, seen)
    return seen


def resolve_coords(full_addr: str, name: str, gu_hint: str, cache: dict, key: str) -> tuple[float, float] | None:
    for qv in geocode_query_variants(full_addr):
        c = kakao_geocode(qv, cache, key)
        if c:
            return c
    for qv in geocode_query_variants(full_addr):
        c = kakao_keyword_geocode(qv, cache, key)
        if c:
            return c
    c = kakao_keyword_geocode(f"{full_addr} {name}", cache, key)
    if c:
        return c
    for q in (
        f"대구 {gu_hint} {name}",
        f"대구광역시 {gu_hint} {name}",
        name,
    ):
        if len(collapse(q)) < 3:
            continue
        c = kakao_keyword_geocode(q, cache, key)
        if c:
            return c
    return None


def merge_key(name: str, addr: str) -> str:
    def norm(x: str) -> str:
        return re.sub(r"\s+", " ", (x or "").lower())

    return f"{norm(name)}|{norm(addr)}"


def parse_buk_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        h1 = collapse(str(ws.cell(1, 1).value or ""))
        h3 = collapse(str(ws.cell(1, 3).value or ""))
        if "번호" not in h1 or "판매소" not in h3:
            continue
        for r in range(2, ws.max_row + 1):
            nm = collapse(str(ws.cell(r, 3).value or ""))
            addr_raw = collapse(str(ws.cell(r, 4).value or ""))
            if not nm or not addr_raw:
                continue
            road = massage_for_geocode(full_daegu_bukgu(addr_raw))
            out.append(
                {
                    "_pri": PRI_XLSX,
                    "name": nm,
                    "roadAddress": road,
                    "daeguGu": "북구",
                    "sourceVendor": "daegu_bukgu_vendor_xlsx",
                    "phone": "",
                }
            )
    wb.close()
    return out


def parse_dalseo_csv(path: Path) -> list[dict]:
    out: list[dict] = []
    try:
        raw_text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        raw_text = path.read_text(encoding="cp949")

    rdr = csv.reader(StringIO(raw_text))
    for row in rdr:
        if not row:
            continue
        if row[0].strip().startswith("\ufeff"):
            row[0] = row[0].lstrip("\ufeff")
        if row[0] == "사업장이름" or ("사업장" in row[0] and "이름" in row[0]):
            continue
        if len(row) < 2:
            continue
        nm = collapse(row[0])
        addr_raw = collapse(row[1])
        tel_raw = collapse(row[2]) if len(row) > 2 else ""
        if not nm or not addr_raw:
            continue
        road = massage_for_geocode(full_daegu_dalseo(addr_raw))
        ph = fmt_phone(tel_raw)
        out.append(
            {
                "_pri": PRI_CSV,
                "name": nm,
                "roadAddress": road,
                "daeguGu": "달서구",
                "sourceVendor": "daegu_dalseo_csv",
                "phone": ph,
            }
        )
    return out


def merged_records(parts: list[dict]) -> list[dict]:
    bucket: dict[str, dict] = {}
    for rec in sorted(parts, key=lambda x: (-x["_pri"], x["roadAddress"])):
        k = merge_key(rec["name"], rec["roadAddress"])
        if k not in bucket:
            bucket[k] = rec
    return list(bucket.values())


def ref_date_default(csv_path: Path | None, override: str) -> str:
    if override.strip():
        return override.strip()
    if csv_path and csv_path.exists():
        m = re.search(r"(\d{8})", csv_path.name)
        if m:
            d = m.group(1)
            return f"{d[:4]}-{d[4:6]}-{d[6:]}"
    return "2026-03-31"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path.home() / "Downloads" / "garbagebag_vendor.xlsx",
    )
    ap.add_argument(
        "--csv",
        type=Path,
        dest="csv_path",
        default=Path.home()
        / "Downloads"
        / "대구광역시 달서구_종량제봉투_지정판매소정보_20240911.csv",
    )
    ap.add_argument("--ref-date", default="", help="YYYY-MM-DD")
    args = ap.parse_args()
    xp = args.xlsx.expanduser()
    cp = args.csv_path.expanduser()
    ref = ref_date_default(cp, args.ref_date)

    rows: list[dict] = []
    if xp.exists():
        bk = parse_buk_xlsx(xp)
        print(f"북구 xlsx {len(bk)}건 {xp.name}", file=sys.stderr)
        rows.extend(bk)
    else:
        print(f"경고: 북구 xlsx 없음 {xp}", file=sys.stderr)

    if cp.exists():
        ds = parse_dalseo_csv(cp)
        print(f"달서구 CSV {len(ds)}건 {cp.name}", file=sys.stderr)
        rows.extend(ds)
    else:
        print(f"경고: 달서구 CSV 없음 {cp}", file=sys.stderr)

    merged = merged_records(rows)
    print(f"병합 후 {len(merged)}건 · dataReferenceDate={ref}", file=sys.stderr)

    cache = load_cache()
    if not KAKAO_REST_KEY:
        print("KAKAO_REST_KEY 없음.", file=sys.stderr)

    merged.sort(key=lambda x: (x["daeguGu"], x["roadAddress"], x["name"]))
    stores: list[dict] = []
    failed = 0

    for i, rec in enumerate(merged):
        road = rec["roadAddress"]
        nm = rec["name"]
        gu = rec["daeguGu"]
        oid = f"daegu-trash-bd-{i + 1:05d}"
        lat = lng = None
        if KAKAO_REST_KEY:
            c = resolve_coords(road, nm, gu, cache, KAKAO_REST_KEY)
            if c:
                lat, lng = c

        obj: dict = {
            "id": oid,
            "name": nm,
            "roadAddress": road,
            "address": road,
            "businessStatus": "영업",
            "hasTrashBag": True,
            "hasSpecialBag": False,
            "hasLargeWasteSticker": False,
            "adminVerified": False,
            "dataReferenceDate": ref,
            "sourceVendor": rec["sourceVendor"],
            "daeguGu": gu,
        }
        if rec.get("phone"):
            obj["phone"] = rec["phone"]
        if lat is not None:
            obj["lat"] = lat
            obj["lng"] = lng
        else:
            failed += 1
        stores.append(obj)
        if (i + 1) % 100 == 0:
            save_cache(cache)
            print(f"  지오코드 {i+1}/{len(merged)} 실패 누적 {failed}", file=sys.stderr)

    save_cache(cache)
    ok = sum(1 for s in stores if "lat" in s)
    print(f"좌표 {ok}/{len(stores)} 실패 {failed}", file=sys.stderr)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(stores, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"저장 {OUT_JSON}", file=sys.stderr)


if __name__ == "__main__":
    main()
