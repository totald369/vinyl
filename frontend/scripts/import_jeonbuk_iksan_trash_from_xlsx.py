#!/usr/bin/env python3
"""
전북 익산시 종량제봉투·불연성마대 판매소 xlsx → stores.jeonbuk-iksan-trash.json

  pip install openpyxl
  python3 scripts/import_jeonbuk_iksan_trash_from_xlsx.py \\
    --input ~/Downloads/지정판매소\\ 목록\\(익산시\\).xlsx
  python3 scripts/import_jeonbuk_iksan_trash_from_xlsx.py --refresh

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonbuk-iksan-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonbuk-iksan-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "지정판매소 목록(익산시).xlsx"
SHEET_TRASH = "지정판매소 목록"
SHEET_SPECIAL = "마대 판매소 목록"
DATA_START_ROW = 5
REF_DATE = "2026-06-05"
CACHE_VERSION = "v1-iksan-2026"

# 카카오에 없는 읍면동 표기 보정 (엑셀 오타·구 명칭)
DONG_ALIASES: dict[str, str] = {}

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.07

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        for sep in ("NEXT_PUBLIC_", "NEXT_", "#"):
            if sep in v:
                v = v.split(sep, 1)[0].strip()
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def flag_o(val: object) -> bool:
    t = collapse(str(val or "")).upper()
    return t in ("O", "○", "Y", "YES", "예", "여")


def in_iksan_bbox(lat: float, lng: float) -> bool:
    # 함열·웅포·용안 면 포함
    return 35.88 <= lat <= 36.13 and 126.86 <= lng <= 127.14


def iksan_in_text(blob: str) -> bool:
    return "익산" in (blob or "").replace(" ", "")


def format_display_addr(raw: str) -> str:
    """저장용 주소 — 전북(전라북도) 표기 통일."""
    a = collapse(raw)
    a = a.replace("전북특별치도", "전북특별자치도")
    a = re.sub(r"^전라북도\s+", "전북특별자치도 ", a)
    a = re.sub(r"^전북\s+", "전북특별자치도 ", a)
    if not a.startswith("전북") and "익산" in a.replace(" ", ""):
        if a.startswith("익산시"):
            return f"전북특별자치도 {a}"
        return f"전북특별자치도 익산시 {a}"
    return a


def dedupe_addr(addr: str) -> str:
    a = collapse(addr)
    a = re.sub(r"\b상가\s+(?=[가-힣]+(?:로|길|대로))", "", a)
    a = re.sub(r"\s+KR(?:\s+\d+)?(?:\s+\d+동)?\s*$", "", a, flags=re.I)
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"\s+\d+\s*층\s*", " ", a)
    a = re.sub(r"\s+\d+\s*호\s*", " ", a)
    a = re.sub(r"\s+익산시\s+전라북도\s*$", "", a)
    a = re.sub(r"\s+전라북도\s*$", "", a)
    a = re.sub(r"\s+예스마트\s+\d+\s*$", "", a, flags=re.I)
    if "," in a:
        head = a.split(",")[0].strip()
        if re.search(r"(?:로|길|대로)\s*\d", head):
            a = head
    parts = a.split()
    mid = len(parts) // 2
    if mid >= 4 and parts[:mid] == parts[mid : mid + mid]:
        a = " ".join(parts[:mid])
    half = len(a) // 2
    if half > 12:
        left, right = a[:half].strip(), a[half:].strip()
        if right.startswith(left):
            return collapse(left)
    return collapse(a)


def extract_paren_hint(raw: str) -> tuple[str, str]:
    a = collapse(raw)
    hint = ""
    m = re.search(r"\(([^)]+)\)", a)
    if m:
        parts = [collapse(p) for p in m.group(1).split(",") if collapse(p)]
        for p in parts:
            if re.search(r"(동|읍|면|리|가)$", p.replace(" ", "")):
                hint = p
                break
        if not hint and parts:
            hint = parts[0]
        a = collapse(re.sub(r"\s*\([^)]*\)\s*", " ", a))
    return a, hint


def normalize_iksan_addr(addr: str) -> str:
    core, hint = extract_paren_hint(addr)
    a = dedupe_addr(core)
    a = a.replace("전북특별치도", "전북특별자치도")
    a = a.replace("전라북도", "전북특별자치도")
    a = re.sub(r"^전북\s+", "전북특별자치도 ", a)
    a = re.sub(r"(\d+)\s+(\d+)\s*$", r"\1-\2", a)
    a = re.sub(r"([가-힣]+(?:로|길))(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+동)(\d+)", r"\1 \2", a)
    if a.startswith("익산시 "):
        a = a[len("익산시 ") :]
    if hint in ("상가", "상가동"):
        hint = "금강동" if "약촌" in a or "약촌" in core else ""
    if hint and hint not in a:
        if is_likely_jibeon(a) and not re.match(r"^[가-힣]+동", a.replace(" ", "")):
            a = f"{hint} {a}"
        elif not is_likely_jibeon(a) and not re.search(r"(읍|면)", a):
            a = f"{hint} {a}"
    compact = a.replace(" ", "")
    if "전북" in compact or "전라북" in compact:
        return collapse(a if a.startswith("전북") else f"전북특별자치도 {a}")
    if a.startswith("익산시"):
        return f"전북특별자치도 {a}"
    return f"전북특별자치도 익산시 {a}"


def iksan_tail(addr: str) -> str:
    """카카오 API 쿼리용 — `전북 익산시 {tail}` 형태가 가장 잘 맞음."""
    norm = normalize_iksan_addr(addr)
    for prefix in (
        "전북특별자치도 익산시 ",
        "전북특별자치도익산시",
        "전북특별자치도 ",
        "전북 익산시 ",
        "전라북도 익산시 ",
        "익산시 ",
    ):
        if norm.startswith(prefix):
            return collapse(norm[len(prefix) :])
    if norm.startswith("전북특별자치도"):
        return collapse(norm.replace("전북특별자치도", "", 1))
    return norm


def is_likely_jibeon(tail: str) -> bool:
    t = tail.replace(" ", "")
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", t))


def lot_query_variants(tail: str) -> list[str]:
    """지번-only 주소 — 동일 번지·인근 호수 후보."""
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = LOT_RE.match(tail)
    if m:
        prefix, main, sub = m.group("prefix"), m.group("main"), m.group("sub")
        subs: list[str | None] = []
        if sub:
            subs.extend([sub, "1", "2", "3", "4", "5"])
        else:
            subs.append(None)
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    tail = iksan_tail(addr_raw)
    out: list[str] = []
    seen: set[str] = set()

    def push_q(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    if is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push_q(f"전북 익산시 {tv}")
            push_q(f"전북특별자치도 익산시 {tv}")
    else:
        push_q(f"전북 익산시 {tail}")
        push_q(f"전북특별자치도 익산시 {tail}")
        if re.search(r"-\d+\s*$", tail):
            no_suffix = re.sub(r"-\d+\s*$", "", tail).strip()
            push_q(f"전북 익산시 {no_suffix}")

    norm = normalize_iksan_addr(addr_raw)
    push_q(norm.replace("전북특별자치도", "전북"))
    push_q(f"{name} 익산시")
    push_q(f"{name} {tail}")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_iksan_bbox(lat, lng):
        return None
    if not iksan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_iksan_bbox(lat, lng):
        return None
    if not iksan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    return data.get("documents") or []


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return None


def extract_dong_area(tail: str) -> str | None:
    m = re.match(r"^(.+?(?:동|읍|면|리))", tail)
    if m:
        return collapse(m.group(1))
    m = re.match(r"^(.+?동)\d", tail.replace(" ", ""))
    if m:
        return collapse(m.group(1))
    return None


def dong_fallback(addr_raw: str, key: str) -> GeoHit | None:
    """지번·도로명 모두 실패 시 읍면동 중심 + coord2address."""
    tail = iksan_tail(addr_raw)
    area = extract_dong_area(tail)
    if not area:
        return None
    area = DONG_ALIASES.get(area, area)
    for q in (f"전북 익산시 {area}", f"전북특별자치도 익산시 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                orig = format_display_addr(normalize_iksan_addr(addr_raw))
                hit.jibeon = orig
                if not hit.road or hit.road == hit.jibeon:
                    cj, cr = coord2address(hit.lng, hit.lat, key)
                    hit.road = cr or cj or hit.road
                return hit
    return None


def road_strip_fallback(addr_raw: str, key: str) -> GeoHit | None:
    """도로명+번지가 DB에 없을 때 도로명만으로 근사 좌표."""
    tail = iksan_tail(addr_raw)
    if is_likely_jibeon(tail):
        return None
    orig = format_display_addr(normalize_iksan_addr(addr_raw))
    candidates: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            candidates.append(t)

    push(re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail))
    push(re.sub(r"-\d+\s*$", "", tail))
    m = ROAD_RE.search(tail)
    if m:
        push(tail[: m.end()].strip())

    for road_only in candidates:
        if not road_only or road_only == tail:
            continue
        for q in (f"전북 익산시 {road_only}", f"전북특별자치도 익산시 {road_only}"):
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    hit.jibeon = orig
                    hit.road = orig if ROAD_RE.search(orig.replace(" ", "")) else hit.road
                    cj, cr = coord2address(hit.lng, hit.lat, key)
                    if not hit.road or hit.road == hit.jibeon:
                        hit.road = cr or orig
                    if not hit.jibeon:
                        hit.jibeon = cj or orig
                    return hit
    return None


@dataclass
class IksanRow:
    name: str
    addr_raw: str
    has_trash: bool
    has_special: bool


def iter_merged_rows(path: Path) -> list[IksanRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    merged: dict[str, IksanRow] = {}
    for sheet_name, set_trash, set_special in (
        (SHEET_TRASH, True, False),
        (SHEET_SPECIAL, False, True),
    ):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
            name = collapse(str(ws.cell(r, 2).value or ""))
            addr_raw = collapse(str(ws.cell(r, 3).value or ""))
            if not name or not addr_raw or name in ("판매소명", "상호"):
                continue
            norm = normalize_iksan_addr(addr_raw)
            key = f"{name}|{norm}"
            if key in merged:
                row = merged[key]
                if set_trash:
                    row.has_trash = True
                if set_special:
                    row.has_special = True
            else:
                merged[key] = IksanRow(
                    name=name,
                    addr_raw=addr_raw,
                    has_trash=set_trash,
                    has_special=set_special,
                )
    wb.close()
    return list(merged.values())


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true", help="지오코딩 캐시 무시 후 재조회")
    ap.add_argument("--skip-activity", action="store_true", help="activities.json 기록 생략")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        print(
            "오류: KAKAO_REST_API_KEY 가 없습니다. frontend/.env.local 을 확인하거나 --skip-kakao",
            file=sys.stderr,
        )
        raise SystemExit(1)

    rows = iter_merged_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    fallback_n = 0

    for row in rows:
        if not row.has_trash and not row.has_special:
            continue
        norm_addr = normalize_iksan_addr(row.addr_raw)
        ck = cache_key(row.name, norm_addr)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else ""
                jibeon = str(raw[3]) if len(raw) > 3 else ""
                if not road or not jibeon:
                    cj, cr = coord2address(lng, lat, key)
                    jibeon = jibeon or cj or norm_addr
                    road = road or cr or cj or jibeon
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jibeon)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key)
            if hit is None:
                hit = road_strip_fallback(row.addr_raw, key)
                if hit:
                    fallback_n += 1
            if hit is None:
                hit = dong_fallback(row.addr_raw, key)
                if hit:
                    fallback_n += 1
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{norm_addr}", file=sys.stderr)
                continue
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{norm_addr}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"jeonbuk-iksan-trash-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road,
                "address": hit.jibeon,
                "businessStatus": "영업",
                "hasTrashBag": row.has_trash,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": REF_DATE,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={REF_DATE}, geo={geo_n}, fallback={fallback_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from append_activity import record_region_data_added

        record_region_data_added(["익산시"])


if __name__ == "__main__":
    main()
