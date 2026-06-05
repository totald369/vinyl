#!/usr/bin/env python3
"""
서울 용산구 특수마대(불연성) 판매처 xlsx → stores.seoul-yongsan-special.json

시트 Sheet1 (2행~):
  연번 | 동명 | 상호 | 주소 | 연락처

  python3 scripts/import_seoul_yongsan_special_from_xlsx.py \\
    --input ~/Downloads/용산구_특수마대\\ 판매처(2026년).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.seoul-yongsan-special.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-seoul-yongsan-special.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "용산구_특수마대 판매처(2026년).xlsx"
SHEET_NAME = "Sheet1"
DATA_START_ROW = 2
REF_DATE = "2026-01-01"
CACHE_VERSION = "v1-seoul-yongsan-special"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def ref_date_from_path(p: Path) -> str:
    m = re.search(r"(\d{4})년", p.name)
    if m:
        return f"{m.group(1)}-01-01"
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def in_yongsan_bbox(lat: float, lng: float) -> bool:
    return 37.52 <= lat <= 37.56 and 126.94 <= lng <= 127.02


def yongsan_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    return "용산구" in t and "서울" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^서울시\s+", "서울특별시 ", a)
    if not a:
        return ""
    if a.startswith("서울특별시"):
        return a
    if a.startswith("서울 "):
        return "서울특별시 " + a[len("서울 ") :]
    if a.startswith("용산구"):
        return f"서울특별시 {a}"
    return f"서울특별시 용산구 {a}"


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def normalize_addr(addr_raw: str, dong: str = "") -> str:
    a = collapse(addr_raw)
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s+\d+\s*층\s*", " ", a)
    a = re.sub(r"\s+\d+\s*호\s*$", "", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+지하\d*\s*층.*$", "", a)
    a = re.sub(r"\s+지하\s*\d*\s*층.*$", "", a)
    a = re.sub(r"\s+[A-Za-z가-힣]+빌.*$", "", a)
    a = re.sub(r"\s+용산구종합행정타운.*$", "", a)
    a = re.sub(r"\s+주민센터.*$", "", a)
    a = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", a)
    a = re.sub(r"([가-힣]+(?:로|길))(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+동)(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+가)(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+)길\s+(\d+)", r"\1길 \2", a)
    dong_clean = collapse(dong)
    compact = a.replace(" ", "")
    if dong_clean and not re.search(r"(동|가|리)\d", compact) and not re.search(
        r"(동|가|리)$", compact
    ):
        if ROAD_RE.search(a) or re.search(r"(로|길|대로)", a):
            a = f"{dong_clean} {a}"
    return format_display_addr(a)


def geocode_query_variants(addr_raw: str, name: str, dong: str = "") -> list[str]:
    clean_name = re.sub(r"\([^)]*\)", "", name).strip()
    norm = normalize_addr(addr_raw, dong)
    tail = norm
    for prefix in ("서울특별시 용산구 ", "서울특별시 ", "용산구 "):
        if tail.startswith(prefix):
            tail = collapse(tail[len(prefix) :])
            break
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"서울 용산구 {tail}")
    push(f"서울특별시 용산구 {tail}")
    if is_likely_jibeon(tail):
        token_m = re.match(r"^([가-힣0-9]+(?:동|가))\s+(\d+(?:-\d+)?)", tail)
        if token_m:
            core_lot = f"{token_m.group(1)} {token_m.group(2)}"
            push(f"서울특별시 용산구 {core_lot}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"서울특별시 용산구 {road_only}")
    compact_road = re.sub(r"\s+", "", tail)
    if compact_road != tail.replace(" ", ""):
        push(f"서울특별시 용산구 {compact_road}")
    push(f"{name} 용산구")
    push(f"{name} 서울 용산구")
    if clean_name and clean_name != name:
        push(f"{clean_name} 용산구")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class YongsanRow:
    name: str
    addr_raw: str
    dong: str
    phone: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_yongsan_bbox(lat, lng):
        return None
    if not yongsan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_yongsan_bbox(lat, lng):
        return None
    if not yongsan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def resolve_geocode(addr_raw: str, name: str, key: str, dong: str = "") -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name, dong):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    area = collapse(dong)
    if area:
        for q in (f"서울 용산구 {area}", f"서울특별시 용산구 {area}"):
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return None


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[YongsanRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    out: list[YongsanRow] = []
    last_dong = ""
    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        dong = cell_str(ws.cell(r, 2).value)
        if dong:
            last_dong = dong
        name = cell_str(ws.cell(r, 3).value)
        addr_raw = cell_str(ws.cell(r, 4).value)
        if not name or not addr_raw or name in ("상호",):
            continue
        out.append(
            YongsanRow(
                name=name,
                addr_raw=addr_raw,
                dong=last_dong,
                phone=cell_str(ws.cell(r, 5).value),
            )
        )
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    misses = 0
    geo_n = 0
    seen: set[str] = set()

    for row in rows:
        display_road = normalize_addr(row.addr_raw, row.dong)
        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None
        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key, row.dong)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = hit.road or cr or display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 20 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        store: dict = {
            "id": f"seoul-yongsan-special-{rid}",
            "name": row.name,
            "lat": round(float(hit.lat), 7),
            "lng": round(float(hit.lng), 7),
            "roadAddress": hit.road or display_road,
            "address": hit.jibeon or display_road,
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": ref_date,
        }
        if row.phone:
            store["phone"] = row.phone
        out.append(store)

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["용산구"])


if __name__ == "__main__":
    main()
