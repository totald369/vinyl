#!/usr/bin/env python3
"""
경기도 시흥시 불연성마대 판매소 xlsx → stores.gyeonggi-siheung-special.json

시트 `요청정보` (헤더 2행, 데이터 3행~):
  판매소명칭(A) | 도로명주소(B) | 입고일(C) | 취급현황(D)

매핑:
  - 목록 등재 + D열 'Y' -> hasSpecialBag: true
  - hasTrashBag: false (불연성마대 전용 파일)
  - 기존 stores.gyeonggi-siheung-trash.json 좌표 재사용

사용:
  cd frontend
  python3 scripts/import_gyeonggi_siheung_special_from_xlsx.py \\
    --input ~/Downloads/시흥시_불연성마대\\ 판매소\\ 위치정보\\ .xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.gyeonggi-siheung-special.json"
TRASH_JSON = FRONTEND / "public" / "data" / "stores.gyeonggi-siheung-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-gyeonggi-siheung-special.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "시흥시_불연성마대 판매소 위치정보 .xlsx"
SHEET_NAME = "요청정보"
DATA_START_ROW = 3
REF_DATE = "2026-06-09"
CACHE_VERSION = "v1-siheung-special"

COL_NAME = 1
COL_ADDR = 2
COL_SPECIAL = 4

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")
SIHEUNG_DONGS = (
    "대야동",
    "신천동",
    "은행동",
    "매화동",
    "목감동",
    "정왕동",
    "거모동",
    "월곶동",
    "과림동",
    "군자동",
    "능곡동",
    "장곡동",
    "계수동",
    "연성동",
    "배곧동",
)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    t = collapse(str(val or "")).upper()
    return t in ("O", "○", "Y", "YES", "예", "여")


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def ref_date_from_path(p: Path) -> str:
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def clean_raw_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if "?" in a:
        a = a.split("?", 1)[0].strip()
    a = re.sub(r"\s*~\s*\d+.*$", "", a)
    a = re.sub(r",\s*$", "", a)
    return a


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    if not a:
        return ""
    if a.startswith("경기 시흥시"):
        return "경기도 시흥시 " + a[7:].strip()
    if a.startswith("시흥시"):
        return f"경기도 {a}"
    if a.startswith("경기도 시흥시"):
        return a
    return f"경기도 시흥시 {a}"


def dong_hint(addr_raw: str) -> str:
    m = re.search(r"\(([^)]+)\)", addr_raw)
    if m:
        return collapse(m.group(1).split(",")[0])
    for d in SIHEUNG_DONGS:
        if d in addr_raw:
            return d
    return ""


def normalize_addr(addr_raw: str) -> str:
    a = clean_raw_addr(addr_raw)
    if not a:
        return ""
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+(?:,\s*\d+)?\s*호.*$", "", a)
    a = re.sub(r"\s+상가동.*$", "", a)
    a = re.sub(r"\s+지하\s*$", "", a)
    a = re.sub(r"([가-힣]+(?:로|길))(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+번길)(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+)(번길)", r"\1\2", a)
    return format_display_addr(a)


def geocode_target(full: str) -> str:
    a = collapse(full)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+(?:,\s*\d+)?\s*호.*$", "", a)
    a = re.sub(r"\s+상가동.*$", "", a)
    return a


def lookup_key(name: str, addr_raw: str) -> str:
    road = geocode_target(normalize_addr(addr_raw))
    return f"{collapse(name)}|{road}"


def in_siheung_bbox(lat: float, lng: float) -> bool:
    return 37.30 <= lat <= 37.50 and 126.64 <= lng <= 126.87


def siheung_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if any(x in t for x in ("서울", "인천", "부천", "안산", "광명")):
        if "시흥시" not in t and "경기도시흥" not in t:
            return False
    return "시흥시" in t or ("경기" in t and "시흥" in t)


def siheung_tail(full: str) -> str:
    for prefix in (
        "경기도 시흥시 ",
        "경기 시흥시 ",
        "시흥시 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    norm = normalize_addr(addr_raw)
    hint = dong_hint(addr_raw)
    tail = siheung_tail(geocode_target(norm))
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"경기도 시흥시 {tail}")
    push(f"경기 시흥시 {tail}")
    compact = norm.replace(" ", "")
    push(compact)
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    push(re.sub(r"((?:로|길|대로))(\d+)", r"\1 \2", compact))
    if hint:
        push(f"경기도 시흥시 {hint}")
        push(f"경기 시흥시 {hint} {tail}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"경기도 시흥시 {road_only}")
    if is_likely_jibeon(tail):
        token_m = re.match(r"^([가-힣0-9]+(?:동|리))\s+(\d+(?:-\d+)?)", tail)
        if token_m:
            core = f"{token_m.group(1)} {token_m.group(2)}"
            push(f"경기도 시흥시 {core}")
    push(f"{name} 시흥시")
    push(f"{name} 시흥")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class SiheungRow:
    name: str
    addr_raw: str
    display_road: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def load_trash_seed() -> dict[str, GeoHit]:
    if not TRASH_JSON.is_file():
        return {}
    try:
        rows = json.loads(TRASH_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, GeoHit] = {}
    for row in rows:
        name = collapse(str(row.get("name") or ""))
        road = str(row.get("roadAddress") or "")
        lat = parse_float(row.get("lat"))
        lng = parse_float(row.get("lng"))
        if not name or lat is None or lng is None:
            continue
        key = lookup_key(name, road)
        out[key] = GeoHit(lat=lat, lng=lng, road=road, jibeon=str(row.get("address") or road))
    return out


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_siheung_bbox(lat, lng):
        return None
    if not siheung_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_siheung_bbox(lat, lng):
        return None
    if not siheung_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(addr_raw: str, key: str, display_road: str) -> GeoHit | None:
    hint = dong_hint(addr_raw)
    if not hint:
        tail = siheung_tail(geocode_target(display_road))
        for d in SIHEUNG_DONGS:
            if d in tail:
                hint = d
                break
    if not hint:
        return None
    for q in (f"경기도 시흥시 {hint}", f"경기 시흥시 {hint}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, _ = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, display_road: str) -> GeoHit | None:
    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", display_road))
    for q in geocode_query_variants(addr_raw, name):
        if has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
        if not has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return area_fallback(addr_raw, key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def display_road_from_raw(addr_raw: str) -> str:
    return format_display_addr(clean_raw_addr(addr_raw))


def iter_rows(path: Path) -> list[SiheungRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    merged: dict[str, SiheungRow] = {}

    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, COL_NAME).value)
        if not name or name in (" 판매소 명칭", "판매소 명칭"):
            continue
        addr_raw = cell_str(ws.cell(r, COL_ADDR).value)
        if not addr_raw:
            continue
        if "시흥" not in addr_raw.replace(" ", ""):
            continue
        if not flag_y(ws.cell(r, COL_SPECIAL).value):
            continue
        display = display_road_from_raw(addr_raw)
        k = lookup_key(name, addr_raw)
        if k not in merged:
            merged[k] = SiheungRow(name=collapse(name), addr_raw=addr_raw, display_road=display)

    wb.close()
    return list(merged.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    trash_seed = load_trash_seed()
    out: list[dict] = []
    misses = 0
    geo_n = 0
    seed_n = 0
    seen: set[str] = set()

    for row in rows:
        display_road = row.display_road
        if not display_road or "시흥" not in display_road.replace(" ", ""):
            continue

        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None:
            seed_key = lookup_key(row.name, row.addr_raw)
            if seed_key in trash_seed:
                hit = trash_seed[seed_key]
                seed_n += 1

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 30 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"gyeonggi-siheung-special-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": False,
                "hasSpecialBag": True,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref_date,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, seed={seed_n}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["시흥시"])


if __name__ == "__main__":
    main()
