#!/usr/bin/env python3
"""
충청남도 행안부 형식 종량제봉투 판매업 xlsx → stores.chungnam-trash.json

- 좌표: Korea 1985 중부원점(EPSG:2097) → WGS84
- 좌표 누락·파싱 실패 시: 카카오 주소 검색(선택) + 캐시 `geocode-cache-chungnam-trash.json`
- 주소에 `충청남도`가 없는 행은 제외(타 시도 오기입 방지)

  pip install pyproj openpyxl
  python3 scripts/import_chungnam_trash_from_xlsx.py \\
    --input ~/Downloads/충청남도_쓰레기종량제봉투판매업\\ 정보.xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

try:
    from pyproj import Transformer
except ImportError:
    raise SystemExit("pyproj 필요: pip install pyproj") from None

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.chungnam-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-chungnam-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "충청남도_쓰레기종량제봉투판매업 정보.xlsx"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
GEOCODE_DELAY = 0.07

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
_SHEET_DATE = re.compile(r"_(\d{8})\s*$")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def sheet_to_ref_date(name: str) -> str | None:
    m = _SHEET_DATE.search(name or "")
    if not m:
        return None
    d = m.group(1)
    return f"{d[:4]}-{d[4:6]}-{d[6:8]}"


def is_chungnam_addr(road: str, lot: str) -> bool:
    blob = (road or "") + (lot or "")
    return "충청남도" in blob or "충남 " in blob or blob.startswith("충남")


def addr_blob_is_chungnam_region(addr_blob: str) -> bool:
    """카카오 주소응답은 `충남`(약칭)으로 오는 경우가 많음."""
    if "충청남도" in addr_blob or "충남" in addr_blob:
        return True
    return False


def in_chungnam_bbox(lat: float, lng: float) -> bool:
    return 34.25 <= lat <= 37.55 and 125.35 <= lng <= 128.25


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())  # type: ignore[arg-type]
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def addr_variants(road: str, lot: str) -> list[str]:
    """지오코딩 쿼리 후보 — 지번(번지 포함) 우선 후 도로명(괄호 제거)."""
    out: list[str] = []
    seen: set[str] = set()

    def add(s: str) -> None:
        t = collapse(s)
        if not t or t in seen:
            return
        seen.add(t)
        out.append(t)

    def lot_short(s: str) -> None:
        add(s)
        if "번지" in s:
            add(re.sub(r"\s*번지\s*", " ", s).strip())

    # 지번 우선 — 카카오가 동·지번 검색에서 잘 맞춤
    lot_short(lot)
    add(road)
    if road:
        r2 = re.sub(r"\s*\([^)]*\)\s*", " ", road)
        r2 = collapse(r2)
        add(r2)
    return out


def _doc_addr_blob(d: dict) -> str:
    """카카오 주소검색 document에서 주소 문자열 합치기."""
    parts: list[str] = []

    def add(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for subk in ("address_name", "region_1depth_name"):
                sv = v.get(subk)
                if isinstance(sv, str) and sv.strip():
                    parts.append(sv.strip())

    add(d.get("address_name"))
    add(d.get("address"))
    add(d.get("road_address"))
    return " ".join(parts)


def kakao_geocode(
    query: str, key: str, *, require_chungnam: bool
) -> tuple[float | None, float | None]:
    if not query.strip():
        return None, None
    req = urllib.request.Request(
        f"{GEOCODE_URL}?{urllib.parse.urlencode({'query': query})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return None, None
    for d in data.get("documents") or []:
        lat = parse_float(d.get("y"))
        lng = parse_float(d.get("x"))
        if lat is None or lng is None:
            continue
        if not in_chungnam_bbox(lat, lng):
            continue
        addr_blob = _doc_addr_blob(d)
        if require_chungnam and not addr_blob_is_chungnam_region(addr_blob):
            continue
        return lat, lng
    return None, None


def resolve_lat_lng(
    *,
    ex: float | None,
    ey: float | None,
    road: str,
    lot: str,
    name: str,
    mgmt: str,
    to_wgs,
    cache: dict[str, list[float]],
    key: str | None,
    allow_kakao: bool,
) -> tuple[float | None, float | None, str]:
    """Returns (lat, lng, source) where source is 'tm'|'kakao'|'cache'."""
    sep = "\x1f"
    ck = "k:" + hashlib.sha1(f"{mgmt}{sep}{road}{sep}{lot}".encode()).hexdigest()[:32]
    hit = cache.get(ck)
    if hit and len(hit) == 2:
        return float(hit[0]), float(hit[1]), "cache"

    if ex is not None and ey is not None:
        lng, lat = to_wgs.transform(ex, ey)
        if (
            is_chungnam_addr(road, lot)
            and in_chungnam_bbox(lat, lng)
        ):
            return lat, lng, "tm"

    if not allow_kakao or not key:
        return None, None, ""

    for q in addr_variants(road, lot):
        for cand in (
            q,
            f"{name} {q}",
        ):
            glat, glng = kakao_geocode(cand, key, require_chungnam=True)
            if glat is not None and glng is not None:
                cache[ck] = [float(glat), float(glng)]
                return glat, glng, "kakao"
    return None, None, ""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument(
        "--skip-kakao",
        action="store_true",
        help="카카오 미호출(TM 좌표만 사용; 누락 좌표 행 제외)",
    )
    args = ap.parse_args()
    xlsx = args.input.expanduser()
    if not xlsx.is_file():
        raise SystemExit(f"파일 없음: {xlsx}")

    key = load_kakao_key()
    allow_kakao = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        print(
            "경고: KAKAO_REST_API_KEY 없음 — TM 좌표만 사용합니다. "
            "(전체 반영은 frontend/.env.local 설정 후 재실행)",
            file=sys.stderr,
        )

    cache = load_cache()
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_r = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col = {str(h).strip(): i for i, h in enumerate(hdr_r) if h is not None and str(h).strip()}

    def req(name: str) -> int:
        if name not in col:
            raise SystemExit(f"열 없음: {name!r}")
        return col[name]

    i_mgmt = req("관리번호")
    i_name = req("사업장명")
    i_lot = req("지번주소")
    i_road = req("도로명주소")
    i_x = req("좌표정보(X)")
    i_y = req("좌표정보(Y)")
    i_detail = req("상세영업상태명")
    i_status = req("영업상태명")
    i_phone = col.get("전화번호")

    ref_date = sheet_to_ref_date(ws.title) or "2021-10-17"
    to_wgs = Transformer.from_crs("EPSG:2097", "EPSG:4326", always_xy=True)

    out: list[dict] = []
    seen_mgmt: set[str] = set()
    stats = {"no_cn_addr": 0, "no_coord": 0, "dup": 0, "kakao_calls": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        detail = collapse(str(row[i_detail] or ""))
        if detail != "영업":
            continue

        mgmt = collapse(str(row[i_mgmt] or ""))
        if not mgmt:
            continue
        if mgmt in seen_mgmt:
            stats["dup"] += 1
            continue
        seen_mgmt.add(mgmt)

        name = collapse(str(row[i_name] or ""))
        if not name:
            continue

        road = collapse(str(row[i_road] or ""))
        lot = collapse(str(row[i_lot] or ""))
        if not road and not lot:
            continue

        if not is_chungnam_addr(road, lot):
            stats["no_cn_addr"] += 1
            continue

        ex = parse_float(row[i_x])
        ey = parse_float(row[i_y])
        ex_p = ex if ex is not None else None
        ey_p = ey if ey is not None else None

        lat, lng, src = resolve_lat_lng(
            ex=ex_p,
            ey=ey_p,
            road=road,
            lot=lot,
            name=name,
            mgmt=mgmt,
            to_wgs=to_wgs,
            cache=cache,
            key=key,
            allow_kakao=allow_kakao,
        )
        if lat is None or lng is None:
            stats["no_coord"] += 1
            continue
        if src == "kakao":
            stats["kakao_calls"] += 1
            if stats["kakao_calls"] % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {stats['kakao_calls']} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)

        st = collapse(str(row[i_status] or ""))
        business_status = "영업" if "폐업" not in st and "휴업" not in st else st or "영업"

        rid = f"chungnam-trash-{mgmt}"
        rec: dict = {
            "id": rid,
            "name": name,
            "lat": round(float(lat), 7),
            "lng": round(float(lng), 7),
            "roadAddress": road or lot,
            "address": lot or road,
            "businessStatus": business_status if business_status else "영업",
            "hasTrashBag": True,
            "hasSpecialBag": False,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": ref_date,
        }
        if i_phone is not None:
            phone = collapse(str(row[i_phone] or ""))
            if phone:
                rec["phone"] = phone
        out.append(rec)

    wb.close()
    save_cache(cache)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} (ref_date={ref_date}, "
        f"kakao_lookups≈{stats['kakao_calls']}, no_coord={stats['no_coord']}, "
        f"no_cn_addr={stats['no_cn_addr']}, dup_mgmt={stats['dup']})"
    )


if __name__ == "__main__":
    main()
