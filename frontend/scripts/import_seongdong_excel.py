#!/usr/bin/env python3
"""
서울 성동구 종량제봉투 판매소 엑셀(동별 시트) → stores.sample.json 병합.

- 시트마다 3행째 헤더: 연번 / 상호 / 주소 / … / 비고
- 비고에 '마대' 포함 시 hasSpecialBag (불연성 마대)
- 목록상 판매소는 hasTrashBag True
- 주소는 서울특별시 성동구 로 정규화 후 카카오 지오코딩

사용:
  frontend/.env.local 의 KAKAO_REST_API_KEY
  python3 scripts/import_seongdong_excel.py
  python3 scripts/import_seongdong_excel.py ~/Downloads/（판매소）....xlsx
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

DOWNLOADS = Path.home() / "Downloads"
FRONTEND = Path(__file__).resolve().parent.parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.sample.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-seongdong.json"

REF_DATE_DEFAULT = "2026-03-01"


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=8) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def norm_store_name(name: str) -> str:
    return re.sub(r"\s*\(\d+\)\s*$", "", (name or "").strip()).strip()


def norm_key(name: str, addr: str) -> str:
    a = re.sub(r"\s+", " ", addr.strip().lower())
    return f"{norm_store_name(name).lower()}|{a}"


def parse_ref_date_from_path(path: Path) -> str:
    m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", path.name)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-01"
    return REF_DATE_DEFAULT


def geocode_query_variants(full_addr: str) -> list[str]:
    seen: list[str] = []
    for q in (full_addr, re.sub(r"\s+", " ", full_addr).strip()):
        if q and q not in seen:
            seen.append(q)
    if "하왕2동" in full_addr:
        alt = full_addr.replace("하왕2동", "하왕십리2동")
        if alt not in seen:
            seen.append(alt)
    base = full_addr
    spaced_alley = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced_alley != base and spaced_alley not in seen:
        seen.append(spaced_alley)
        base = spaced_alley
    ro_gil = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", base)
    if ro_gil != base and ro_gil not in seen:
        seen.append(ro_gil)
        if "," in ro_gil:
            h0 = ro_gil.split(",")[0].strip()
            if h0 not in seen:
                seen.append(h0)
    if "," in base:
        head = base.split(",")[0].strip()
        if head and head not in seen:
            seen.append(head)
    m = re.search(r"([가-힣]+(?:로|길))\s+(\d+)-(\d+)", base)
    if m:
        stripped = base.replace(m.group(0), f"{m.group(1)} {m.group(2)}", 1)
        if stripped not in seen:
            seen.append(stripped)
        if "," in stripped:
            h = stripped.split(",")[0].strip()
            if h not in seen:
                seen.append(h)
    m = re.search(r"(면로|대로|로|길)(\d{2,4})\b", base)
    if m:
        fixed = re.sub(
            r"(면로|대로|로|길)(\d{2,4})\b",
            lambda mm: f"{mm.group(1)} {mm.group(2)}",
            base,
            count=1,
        )
        if fixed not in seen:
            seen.append(fixed)
        if "," in fixed:
            h2 = fixed.split(",")[0].strip()
            if h2 not in seen:
                seen.append(h2)
    return seen


def resolve_coords(
    road_address: str, place_name: str, cache: dict, key: str
) -> tuple[float, float] | None:
    for qv in geocode_query_variants(road_address):
        c = kakao_geocode(qv, cache, key)
        if c:
            return c
    q2 = f"{road_address} {place_name}"
    c = kakao_geocode(q2, cache, key)
    if c:
        return c
    for qv in geocode_query_variants(road_address):
        c = kakao_keyword_geocode(qv, cache, key)
        if c:
            return c
    for q in (
        f"성동구 {place_name}",
        place_name,
        f"서울 {place_name}",
    ):
        c = kakao_keyword_geocode(q, cache, key)
        if c:
            return c
    return None


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def normalize_seongdong_address(addr: str) -> str:
    a = re.sub(r"\s+", " ", addr.replace("\n", " ").strip())
    if not a:
        return a
    a = a.replace("하왕2동", "하왕십리2동")
    if a.startswith("서울특별시"):
        return a
    if a.startswith("서울시"):
        return "서울특별시" + a[3:].lstrip()
    if a.startswith("서울 "):
        rest = a[3:].lstrip()
        if rest.startswith("성동구"):
            return f"서울특별시 {rest}"
        return f"서울특별시 성동구 {rest}"
    if a.startswith("성동구"):
        return f"서울특별시 {a}"
    return f"서울특별시 성동구 {a}"


def has_madae_remark(remark: str) -> bool:
    r = _str(remark)
    return "마대" in r


def find_header_row(ws) -> int | None:
    for ri in range(1, min(20, ws.max_row + 1)):
        v0 = ws.cell(ri, 1).value
        v1 = ws.cell(ri, 2).value
        if _str(v0) == "연번" and _str(v1) == "상호":
            return ri
    return None


def parse_workbook(path: Path) -> list[dict]:
    ref_date = parse_ref_date_from_path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hri = find_header_row(ws)
        if hri is None:
            continue
        for ri in range(hri + 1, ws.max_row + 1):
            name = _str(ws.cell(ri, 2).value)
            addr_raw = _str(ws.cell(ri, 3).value)
            remark = ws.cell(ri, 6).value
            if not name or not addr_raw:
                continue
            if name == "상호":
                continue
            road = normalize_seongdong_address(addr_raw)
            out.append(
                {
                    "name": name,
                    "roadAddress": road,
                    "address": road,
                    "businessStatus": "영업",
                    "hasTrashBag": True,
                    "hasSpecialBag": has_madae_remark(remark),
                    "hasLargeWasteSticker": False,
                    "adminVerified": False,
                    "dataReferenceDate": ref_date,
                    "_source_sheet": sn,
                }
            )
    wb.close()
    return out


def discover_path(argv: list[str]) -> Path | None:
    if argv:
        p = Path(argv[0]).expanduser().resolve()
        return p if p.exists() else None
    for p in DOWNLOADS.glob("*.xlsx"):
        n = p.name.replace(" ", "")
        if "판매소" in n and "2026" in n and "현황" in n:
            return p.resolve()
    return None


def main():
    if not KAKAO_REST_KEY:
        print("KAKAO_REST_KEY 또는 KAKAO_REST_API_KEY 가 필요합니다.")
        raise SystemExit(1)

    path = discover_path(sys.argv[1:])
    if path is None:
        print("엑셀 경로를 인자로 주거나 Downloads 에 （판매소）…2026…xlsx 를 두세요.")
        raise SystemExit(1)

    raw = parse_workbook(path)
    print(f"  {path.name}: {len(raw)}행 (전 시트)")

    by_key: dict[str, dict] = {}
    for s in raw:
        k0 = norm_key(s["name"], s["roadAddress"])
        if k0 not in by_key:
            by_key[k0] = s
        else:
            prev = by_key[k0]
            prev["hasSpecialBag"] = bool(prev["hasSpecialBag"]) or bool(s["hasSpecialBag"])
            sd = s.get("dataReferenceDate")
            pd = prev.get("dataReferenceDate")
            if sd and (not pd or sd > pd):
                prev["dataReferenceDate"] = sd

    unique = list(by_key.values())
    print(f"중복 제거 후 {len(unique)}건")

    cache = load_cache()
    geocode_failed: list[str] = []
    print("지오코딩 중…")
    for i, s in enumerate(unique):
        coords = resolve_coords(s["roadAddress"], s["name"], cache, KAKAO_REST_KEY)
        if coords:
            s["lat"], s["lng"] = coords[0], coords[1]
        else:
            s["lat"] = s["lng"] = None
            geocode_failed.append(s["name"])
        s.pop("_source_sheet", None)
        if (i + 1) % 50 == 0:
            save_cache(cache)
            print(f"  …{i + 1}/{len(unique)}")
    save_cache(cache)

    if geocode_failed:
        print(f"좌표 실패 {len(geocode_failed)}건 (일부): {geocode_failed[:20]}")

    with open(OUT_JSON, "r", encoding="utf-8") as f:
        existing = json.load(f)

    max_id = 0
    for e in existing:
        try:
            max_id = max(max_id, int(str(e.get("id", "0"))))
        except ValueError:
            pass

    exist_keys = {
        norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", ""))
        for e in existing
    }

    added = 0
    updated = 0
    for s in unique:
        k0 = norm_key(s["name"], s["roadAddress"])
        if k0 in exist_keys:
            for e in existing:
                if norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", "")) != k0:
                    continue
                ch = False
                nt = bool(e.get("hasTrashBag")) or s["hasTrashBag"]
                ns = bool(e.get("hasSpecialBag")) or s["hasSpecialBag"]
                if nt != bool(e.get("hasTrashBag")):
                    ch = True
                if ns != bool(e.get("hasSpecialBag")):
                    ch = True
                e["hasTrashBag"] = nt
                e["hasSpecialBag"] = ns
                if s.get("dataReferenceDate"):
                    od = e.get("dataReferenceDate") or ""
                    if not od or (s["dataReferenceDate"] > od):
                        e["dataReferenceDate"] = s["dataReferenceDate"]
                        ch = True
                if (e.get("lat") is None or e.get("lng") is None) and s.get("lat") is not None:
                    e["lat"], e["lng"] = s["lat"], s["lng"]
                    ch = True
                if ch:
                    updated += 1
                break
            continue

        if s.get("lat") is None:
            continue

        max_id += 1
        exist_keys.add(k0)
        rec = {
            "id": str(max_id),
            "name": s["name"],
            "lat": s["lat"],
            "lng": s["lng"],
            "roadAddress": s["roadAddress"],
            "address": s["address"],
            "businessStatus": s["businessStatus"],
            "hasTrashBag": s["hasTrashBag"],
            "hasSpecialBag": s["hasSpecialBag"],
            "hasLargeWasteSticker": False,
            "adminVerified": False,
        }
        if s.get("dataReferenceDate"):
            rec["dataReferenceDate"] = s["dataReferenceDate"]
        existing.append(rec)
        added += 1

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"갱신 {updated}건, 신규 {added}건 → 총 {len(existing)} 저장: {OUT_JSON}")


if __name__ == "__main__":
    main()
