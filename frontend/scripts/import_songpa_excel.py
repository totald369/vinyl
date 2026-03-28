#!/usr/bin/env python3
"""
송파구 행정동별 종량제봉투/특수규격(불연성) 엑셀 → stores.sample.json 스키마 병합.
- 일반/가정용 종량제 열의 O → hasTrashBag
- 특수규격봉투 열의 O → hasSpecialBag
- 대표 타이틀의 (2025.07) 등 → dataReferenceDate (YYYY-MM-01)
- 주소: 서울특별시 송파구 + 도로명주소 로 카카오 지오코딩

사용:
  export KAKAO_REST_KEY=... 또는 .env.local 의 KAKAO_REST_API_KEY
  python3 scripts/import_songpa_excel.py
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

DOWNLOADS = Path.home() / "Downloads"
FRONTEND = Path(__file__).resolve().parent.parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.sample.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-songpa.json"


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def is_o(val) -> bool:
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("O", "○", "◎", "●")


def parse_ref_date(title: str) -> str | None:
    if not title:
        return None
    m = re.search(r"(\d{4})\s*[\.\-]\s*(\d{1,2})", title)
    if m:
        y, mo = m.groups()
        return f"{y}-{int(mo):02d}-01"
    return None


def is_target_workbook(f: Path) -> bool:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        t = str(next(ws.iter_rows(values_only=True))[0] or "")
        wb.close()
        return "종량제봉투" in t and ("판매" in t or "현황" in t)
    except Exception:
        return False


def find_header_indices(row1: tuple) -> tuple[int | None, int | None, int | None, int, int]:
    """대행업체, 행정동, 연번, 판매업소명, 도로명주소 열 인덱스 (0-based)."""
    h = [str(x).strip() if x is not None else "" for x in row1]

    def find(*keys: str) -> int | None:
        for i, v in enumerate(h):
            for k in keys:
                if k in v:
                    return i
        return None

    i_name = find("판매업소명", "판매소명")
    i_addr = find("도로명주소")
    i_seq = find("연번")
    i_corp = find("대행업체")
    i_dong = find("행정동")
    if i_name is None or i_addr is None:
        raise ValueError(f"필수 헤더 없음: {h[:12]!r}")
    return i_corp, i_dong, i_seq, i_name, i_addr


def find_trash_and_special_cols(rows: list[tuple]) -> tuple[list[int], int | None]:
    """row1~3에서 종량제 사이즈 열 + 특수규격 열 인덱스."""
    if len(rows) < 4:
        return [], None
    r1, r2, r3 = rows[1], rows[2], rows[3]
    special_col: int | None = None
    for i, c in enumerate(r2):
        if c is not None and "특수" in str(c):
            special_col = i
            break
    trash_cols: list[int] = []
    for i, c in enumerate(r3):
        if c is None:
            continue
        s = str(c).replace("\n", " ").strip()
        if not s:
            continue
        if special_col is not None and i >= special_col:
            break
        if re.match(r"^\d+ℓ$", s) or "재사용" in s or "ℓ" in s:
            trash_cols.append(i)
    if special_col is None:
        for i, c in enumerate(r3):
            if c is not None and str(c).strip() in ("20ℓ",) and i >= (trash_cols[-1] + 1 if trash_cols else 6):
                if i not in trash_cols:
                    special_col = i
                    break
    return trash_cols, special_col


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=8) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def pad_row(r: tuple, n: int) -> list:
    r = list(r)
    while len(r) < n:
        r.append(None)
    return r


def parse_workbook(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [pad_row(tuple(row), 40) for row in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 5:
        return []

    title = str(rows[0][0] or "")
    ref_date = parse_ref_date(title)

    try:
        i_corp, i_dong, i_seq, i_name, i_addr = find_header_indices(tuple(rows[1]))
    except ValueError:
        return []

    trash_cols, special_col = find_trash_and_special_cols(rows)
    if not trash_cols and special_col is None:
        trash_cols = list(range(6, 14))
        special_col = 14

    data_start = 4
    out: list[dict] = []
    carry_corp = ""
    carry_dong = ""

    for r in rows[data_start:]:
        name = r[i_name] if i_name < len(r) else None
        addr = r[i_addr] if i_addr < len(r) else None
        corp = r[i_corp] if i_corp < len(r) else None
        dong = r[i_dong] if i_dong < len(r) else None

        if corp:
            carry_corp = str(corp).strip()
        if dong:
            carry_dong = str(dong).strip()

        if not name or not str(name).strip():
            continue
        road = str(addr).strip() if addr else ""
        if not road:
            continue

        has_trash = any(is_o(r[i]) for i in trash_cols if i < len(r))
        has_spec = bool(special_col is not None and special_col < len(r) and is_o(r[special_col]))

        full_addr = f"서울특별시 송파구 {road}"

        out.append(
            {
                "name": str(name).strip(),
                "roadAddress": full_addr,
                "address": full_addr,
                "businessStatus": "영업",
                "hasTrashBag": has_trash,
                "hasSpecialBag": has_spec,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref_date,
                "_contractor": carry_corp,
                "_haengjeong": carry_dong,
                "_source_file": path.name,
            }
        )

    return out


def norm_key(name: str, addr: str) -> str:
    a = re.sub(r"\s+", " ", addr.strip().lower())
    return f"{name.strip().lower()}|{a}"


def main():
    if not KAKAO_REST_KEY:
        print("환경 변수 KAKAO_REST_KEY 가 필요합니다. (카카오 developers REST API 키)")
        raise SystemExit(1)

    files = sorted(f for f in DOWNLOADS.glob("*.xlsx") if is_target_workbook(f))
    if not files:
        print(f"{DOWNLOADS} 에서 대상 엑셀을 찾지 못했습니다.")
        raise SystemExit(1)

    print(f"대상 파일 {len(files)}개")
    raw: list[dict] = []
    for f in files:
        rows = parse_workbook(f)
        print(f"  {f.name}: {len(rows)}행")
        raw.extend(rows)

    by_key: dict[str, dict] = {}
    for s in raw:
        k = norm_key(s["name"], s["roadAddress"])
        if k not in by_key:
            by_key[k] = s
        else:
            prev = by_key[k]
            prev["hasTrashBag"] = prev["hasTrashBag"] or s["hasTrashBag"]
            prev["hasSpecialBag"] = prev["hasSpecialBag"] or s["hasSpecialBag"]
            rd = prev.get("dataReferenceDate")
            sd = s.get("dataReferenceDate")
            if sd and (not rd or sd > rd):
                prev["dataReferenceDate"] = sd

    unique = list(by_key.values())
    print(f"중복 제거 후 {len(unique)}건")

    cache = load_cache()
    geocode_failed: list[str] = []

    print("지오코딩 중…")
    for i, s in enumerate(unique):
        coords = kakao_geocode(s["roadAddress"], cache, KAKAO_REST_KEY)
        if not coords:
            q2 = f'{s["roadAddress"]} {s["name"]}'
            coords = kakao_geocode(q2, cache, KAKAO_REST_KEY)
        if not coords:
            coords = kakao_keyword_geocode(f'송파구 {s["name"]}', cache, KAKAO_REST_KEY)
        if not coords:
            geocode_failed.append(s["name"])
            s["lat"] = None
            s["lng"] = None
        else:
            s["lat"], s["lng"] = coords
        if (i + 1) % 50 == 0:
            save_cache(cache)
            print(f"  …{i + 1}/{len(unique)}")
    save_cache(cache)

    if geocode_failed:
        print(f"좌표 실패 {len(geocode_failed)}건 (목록 일부): {geocode_failed[:15]}")

    with open(OUT_JSON, "r", encoding="utf-8") as f:
        existing = json.load(f)

    max_id = 0
    for e in existing:
        try:
            max_id = max(max_id, int(str(e.get("id", "0"))))
        except ValueError:
            pass

    exist_keys = {norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", "")) for e in existing}

    added = 0
    for s in unique:
        if s.get("lat") is None:
            continue
        k = norm_key(s["name"], s["roadAddress"])
        if k in exist_keys:
            for e in existing:
                if norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", "")) == k:
                    e["hasTrashBag"] = bool(e.get("hasTrashBag")) or s["hasTrashBag"]
                    e["hasSpecialBag"] = bool(e.get("hasSpecialBag")) or s["hasSpecialBag"]
                    if s.get("dataReferenceDate"):
                        od = e.get("dataReferenceDate") or ""
                        if not od or (s["dataReferenceDate"] > od):
                            e["dataReferenceDate"] = s["dataReferenceDate"]
                    break
            continue

        max_id += 1
        exist_keys.add(k)
        rec = {
            "id": str(max_id),
            "name": s["name"],
            "lat": s["lat"],
            "lng": s["lng"],
            "roadAddress": s["roadAddress"],
            "address": s["address"],
            "businessStatus": s["businessStatus"],
            "hasTrashBag": s["hasTrashBag"],
            "hasSpecialBag": s["hasSpecialBag"],
            "hasLargeWasteSticker": False,
            "adminVerified": False,
        }
        if s.get("dataReferenceDate"):
            rec["dataReferenceDate"] = s["dataReferenceDate"]
        existing.append(rec)
        added += 1

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"기존 {len(existing) - added} + 신규 {added} = 총 {len(existing)} → {OUT_JSON}")


if __name__ == "__main__":
    main()
