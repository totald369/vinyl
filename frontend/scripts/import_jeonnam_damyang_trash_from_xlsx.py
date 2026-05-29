#!/usr/bin/env python3
"""
전라남도 담양군 종량제봉투 판매처 xlsx → stores.jeonnam-damyang-trash.json

시트 `판매업`: 연번 | 판매소명 | 입력 주소 — 3행부터

  python3 scripts/import_jeonnam_damyang_trash_from_xlsx.py \\
    --input ~/Downloads/종량제봉투\\ 판매처\(담양군\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonnam-damyang-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonnam-damyang-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "종량제봉투 판매처(담양군).xlsx"
REF_DATE = date.today().isoformat()
CACHE_VERSION = "v1-damyang"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")
_REF_FILENAME = re.compile(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def ref_date_from_path(p: Path) -> str:
    m = _REF_FILENAME.search(p.name)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return REF_DATE


def in_damyang_bbox(lat: float, lng: float) -> bool:
    return 35.05 <= lat <= 35.48 and 126.85 <= lng <= 127.18


def damyang_in_text(blob: str) -> bool:
    return "담양" in (blob or "").replace(" ", "")


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^전남\s+", "전라남도 ", a)
    if not a:
        return ""
    if a.startswith("전라남도"):
        return a
    if a.startswith("담양군"):
        return f"전라남도 {a}"
    return f"전라남도 담양군 {a}"


def extract_paren_hint(raw: str) -> tuple[str, str]:
    a = collapse(raw)
    hint = ""
    m = re.search(r"\(([^)]+)\)", a)
    if m:
        parts = [collapse(p) for p in m.group(1).split(",") if collapse(p)]
        for p in parts:
            if re.search(r"(동|읍|면|리)$", p.replace(" ", "")):
                hint = p
                break
        if not hint and parts:
            hint = parts[0]
        a = collapse(re.sub(r"\s*\([^)]*\)\s*", " ", a))
    return a, hint


def normalize_addr(addr_raw: str) -> str:
    core, hint = extract_paren_hint(addr_raw)
    a = collapse(core)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"\s+\d+\s*층\s*", " ", a)
    a = re.sub(r"\s+\d+\s*호\s*", " ", a)
    a = re.sub(r"\s*~\s*\d+\s*호\s*", " ", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+외\s+\d+\s*필지\s*", " ", a)
    a = re.sub(r"(\d+)\s+(\d+)\s*$", r"\1-\2", a)
    a = re.sub(r"([가-힣]+(?:리|동))(\d+)", r"\1 \2", a)
    if a.startswith("담양군 "):
        a = a[len("담양군 ") :]
    if hint and hint not in a and not re.search(r"(읍|면)", a):
        if is_likely_jibeon(a) and not re.match(r"^[가-힣]+동", a.replace(" ", "")):
            a = f"{hint} {a}"
        elif not is_likely_jibeon(a):
            a = f"{hint} {a}"
    return format_display_addr(a)


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def lot_query_variants(tail: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m2 = LOT_RE.match(tail)
    if m2:
        prefix, main, sub = m2.group("prefix"), m2.group("main"), m2.group("sub")
        subs: list[str | None] = [sub, "1", "2", "3", "4", "5"] if sub else [None]
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def damyang_tail(full: str) -> str:
    for prefix in (
        "전라남도 담양군 ",
        "전라남도담양군",
        "전라남도 ",
        "전남 담양군 ",
        "담양군 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    if full.startswith("전라남도"):
        return collapse(full.replace("전라남도", "", 1))
    return full


def geocode_query_variants(addr_full: str, name: str) -> list[str]:
    norm = normalize_addr(addr_full)
    tail = damyang_tail(norm)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    if is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push(f"전남 담양군 {tv}")
            push(f"전라남도 담양군 {tv}")
    else:
        push(f"전남 담양군 {tail}")
        push(f"전라남도 담양군 {tail}")
        push(norm)
        if re.search(r"-\d+\s*$", tail):
            rt = re.sub(r"-\d+\s*$", "", tail).strip()
            push(f"전남 담양군 {rt}")
        road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
        if road_only and road_only != tail:
            push(f"전남 담양군 {road_only}")

    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    if m:
        push(f"전남 담양군 {collapse(m.group(1))}")

    push(f"{name} 담양")
    push(f"{name} 담양군")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_damyang_bbox(lat, lng):
        return None
    if not damyang_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_damyang_bbox(lat, lng):
        return None
    if not damyang_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    return data.get("documents") or []


def area_fallback(addr_full: str, key: str, display: str) -> GeoHit | None:
    tail = damyang_tail(normalize_addr(addr_full))
    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    if not m:
        return None
    area = collapse(m.group(1))
    for q in (f"전남 담양군 {area}", f"전라남도 담양군 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display if not is_likely_jibeon(damyang_tail(display)) else hit.road
                cj, cr = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or normalize_addr(addr_full)
                if not hit.road or hit.road == hit.jibeon:
                    hit.road = cr or display
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, display: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return area_fallback(addr_raw, key, display)


def display_addresses(addr_raw: str) -> tuple[str, str]:
    norm = normalize_addr(addr_raw)
    return norm, norm


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["판매업"] if "판매업" in wb.sheetnames else wb.active
    out: list[tuple[str, str]] = []
    for r in range(3, ws.max_row + 1):
        no = ws.cell(r, 1).value
        name = collapse(str(ws.cell(r, 2).value or ""))
        addr = collapse(str(ws.cell(r, 3).value or ""))
        if not name or not addr:
            continue
        try:
            int(str(no).strip())
        except (TypeError, ValueError):
            continue
        out.append((name, addr))
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true", help="activities.json 기록 생략")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    path_ref = ref_date_from_path(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    seen: set[str] = set()

    for name, addr_raw in iter_rows(inp):
        display_road, _ = display_addresses(addr_raw)
        if not display_road or "담양" not in display_road.replace(" ", ""):
            continue

        dk = f"{name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else ""
                if not jib or not road:
                    cj, cr = coord2address(lng, lat, key)
                    jib = jib or cj or display_road
                    road = road or cr or display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(addr_raw, name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            norm = normalize_addr(addr_raw)
            tail = damyang_tail(norm)
            if is_likely_jibeon(tail):
                hit.jibeon = norm
                hit.road = cr or hit.road or norm
            else:
                hit.road = display_road
                hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"jeonnam-damyang-trash-{rid}",
                "name": name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road or display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": False,
                "dataReferenceDate": path_ref,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {len(out)} → {OUT_JSON} (ref_date={path_ref}, api≈{geo_n}, miss={misses})")

    if out and not args.skip_activity:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        from append_activity import record_region_data_added

        record_region_data_added(["담양군"])


if __name__ == "__main__":
    main()
