#!/usr/bin/env python3
"""
경상남도 하동군 종량제봉투·특수마대(불연성마대) 판매소 xlsx → stores.gyeongnam-hadong-trash.json

시트 `판매소 현황` (3행~, A열 비어 있음):
  연번(B) | 판매소명(C) | 읍면명(D) | 입력 주소(E) | 종량제봉투(F) | 특수마대(G)

  python3 scripts/import_gyeongnam_hadong_trash_from_xlsx.py \\
    --input ~/Downloads/하동군_정보공개\\ 청구에\\ 따른\\ 회신내용(종량제봉투\\ 판매소\\ 현황).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.gyeongnam-hadong-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-gyeongnam-hadong-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "하동군_정보공개 청구에 따른 회신내용(종량제봉투 판매소 현황).xlsx"
SHEET_NAME = "판매소 현황"
DATA_START_ROW = 3
REF_DATE = "2026-06-05"
CACHE_VERSION = "v1-hadong"

COL_NAME = 3
COL_AREA = 4
COL_ADDR = 5
COL_TRASH = 6
COL_SPECIAL = 7

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")
HADONG_AREAS = (
    "하동읍",
    "화개면",
    "악양면",
    "적량면",
    "횡천면",
    "고전면",
    "금남면",
    "진교면",
    "양보면",
    "북천면",
    "청암면",
    "옥종면",
    "금성면",
)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    t = collapse(str(val or "")).upper()
    return t in ("O", "○", "Y", "YES", "예", "여")


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def ref_date_from_path(p: Path) -> str:
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def in_hadong_bbox(lat: float, lng: float) -> bool:
    # 화개면·쌍계사 일대(북부) 포함
    return 34.92 <= lat <= 35.28 and 127.52 <= lng <= 127.88


def hadong_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if any(x in t for x in ("하동구", "서울", "대구", "부산", "인천")):
        return False
    if "하동군" in t:
        return True
    return ("경상남도" in t or "경남" in t) and "하동" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^경남\s+", "경상남도 ", a)
    if not a:
        return ""
    if a.startswith("경상남도"):
        return a
    if a.startswith("하동군"):
        return f"경상남도 {a}"
    return f"경상남도 하동군 {a}"


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def normalize_hadong_addr(addr_raw: str, area: str = "") -> str:
    a = collapse(addr_raw)
    if not a:
        return ""
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\.\s*\d+\s*층.*$", "", a)
    a = re.sub(r"\.\s*(?!\d)[^.]+$", "", a)
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"(\d+)\s+(\d+)\s*$", r"\1-\2", a)
    a = re.sub(r"([가-힣]+(?:리|동))(\d+)", r"\1 \2", a)
    a = re.sub(r"하동군(?=[가-힣])", "하동군 ", a)
    if not a.startswith("경상") and not a.startswith("하동"):
        prefix = collapse(area) or "하동읍"
        if prefix and not a.startswith(prefix):
            a = f"{prefix} {a}"
        a = f"하동군 {a}"
    return format_display_addr(a)


def geocode_target(full: str) -> str:
    a = collapse(full)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\.\s*\d+\s*층.*$", "", a)
    a = re.sub(r"\.\s*(?!\d)[^.]+$", "", a)
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+\s*호\s*$", "", a)
    return a


def road_format_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(addr)
    compact = addr.replace(" ", "")
    push(compact)
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    push(re.sub(r"([가-힣]+리)(\d+길)", r"\1\2", compact))
    push(re.sub(r"([가-힣]+로)(\d+)", r"\1 \2", compact))
    return out


def hadong_tail(full: str) -> str:
    for prefix in (
        "경상남도 하동군 ",
        "경상남도 ",
        "경남 하동군 ",
        "하동군 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def lot_query_variants(tail: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = LOT_RE.match(tail)
    if m:
        prefix, main, sub = m.group("prefix"), m.group("main"), m.group("sub")
        subs: list[str | None] = [sub, "1", "2", "3", "4", "5"] if sub else [None]
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def geocode_query_variants(road_full: str, name: str) -> list[str]:
    target = geocode_target(road_full)
    tail = hadong_tail(target)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    if is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push(f"경남 하동군 {tv}")
            push(f"경상남도 하동군 {tv}")
    else:
        for base in road_format_variants(target):
            t = hadong_tail(base)
            push(f"경남 하동군 {t}")
            push(f"경상남도 하동군 {t}")
            push(base)
            road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", t).strip()
            if road_only and road_only != t:
                push(f"경남 하동군 {road_only}")

    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    if m:
        push(f"경남 하동군 {collapse(m.group(1))}")

    push(f"{name} 하동")
    push(f"{name} 하동군")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class HadongRow:
    name: str
    area: str
    addr_raw: str
    has_trash: bool
    has_special: bool


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_hadong_bbox(lat, lng):
        return None
    if not hadong_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_hadong_bbox(lat, lng):
        return None
    if not hadong_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(addr_raw: str, key: str, display_road: str) -> GeoHit | None:
    tail = hadong_tail(geocode_target(addr_raw))
    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    area = collapse(m.group(1)) if m else ""
    if not area:
        for candidate in HADONG_AREAS:
            if candidate in tail:
                area = candidate
                break
    if not area:
        return None
    for q in (f"경남 하동군 {area}", f"경상남도 하동군 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, _ = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, display_road: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return area_fallback(addr_raw, key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[HadongRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    out: list[HadongRow] = []
    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, COL_NAME).value)
        if not name or name in ("판매소명",):
            continue
        area = cell_str(ws.cell(r, COL_AREA).value)
        addr_raw = cell_str(ws.cell(r, COL_ADDR).value)
        if not addr_raw:
            continue
        has_trash = flag_y(ws.cell(r, COL_TRASH).value)
        has_special = flag_y(ws.cell(r, COL_SPECIAL).value)
        if not has_trash and not has_special:
            continue
        out.append(
            HadongRow(
                name=name,
                area=area,
                addr_raw=addr_raw,
                has_trash=has_trash,
                has_special=has_special,
            )
        )
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    seen: set[str] = set()

    for row in rows:
        full_norm = normalize_hadong_addr(row.addr_raw, row.area)
        display_road = geocode_target(full_norm)
        if not display_road or "하동" not in display_road.replace(" ", ""):
            continue

        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(full_norm, row.name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 30 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"gyeongnam-hadong-trash-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": row.has_trash,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": ref_date,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["하동군"])


if __name__ == "__main__":
    main()
