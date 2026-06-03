#!/usr/bin/env python3
"""
전라남도 강진군 종량제봉투·불연성마대 판매업소 xlsx(읍면별 시트) → stores.jeonnam-gangjin-trash.json

각 시트 3행: 상호, 사업장주소(도로명), 종량제봉투 판매여부, 불연성마대 판매여부 (여/부)
4행부터 데이터

  python3 scripts/import_jeonnam_gangjin_trash_from_xlsx.py \\
    --input ~/Downloads/강진군_종량제봉투\\ 판매업소.xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonnam-gangjin-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonnam-gangjin-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "강진군_종량제봉투 판매업소.xlsx"
REF_DATE = date.today().isoformat()
CACHE_VERSION = "v1-gangjin-xlsx"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def in_gangjin_bbox(lat: float, lng: float) -> bool:
    return 34.4 <= lat <= 34.8 and 126.5 <= lng <= 127.1


def gangjin_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if "임실" in t or ("강진면" in t and "강진군" not in t):
        return False
    if "부산" in t or "용산구" in t or "동구" in t and "대구" in t:
        return False
    return "강진군" in t or ("전라남도" in t and "강진" in t)


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^전남\s+", "전라남도 ", a)
    if not a:
        return ""
    if a.startswith("전라남도"):
        return a
    if a.startswith("강진군"):
        return f"전라남도 {a}"
    if re.match(r"^강진(읍|면)", a):
        return f"전라남도 강진군 {a}"
    return f"전라남도 강진군 {a}"


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if "," in a:
        a = a.split(",")[0].strip()
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*@\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    m = re.match(r"^(.+?[로길대로])(\d[\d\-]*)$", a.replace(" ", ""))
    if m:
        a = f"{m.group(1)} {m.group(2)}"
    return format_display_addr(a)


def product_yes(v: object) -> bool:
    s = str(v or "").strip()
    return s in ("여", "Y", "O", "예", "유", "있음")


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_gangjin_bbox(lat, lng):
        return None
    if not gangjin_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_gangjin_bbox(lat, lng):
        return None
    if not gangjin_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    norm = normalize_addr(addr_raw)
    tail = norm
    for prefix in ("전라남도 강진군 ", "전라남도 ", "강진군 "):
        if tail.startswith(prefix):
            tail = collapse(tail[len(prefix) :])
            break
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"전남 강진군 {tail}")
    push(f"전라남도 강진군 {tail}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"전라남도 강진군 {road_only}")
    push(f"{name} 강진")
    push(f"{name} 강진군")
    return out


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return None


@dataclass
class GangjinRow:
    name: str
    addr_raw: str
    has_trash: bool
    has_special: bool
    sheet: str


def iter_rows(path: Path) -> tuple[list[GangjinRow], str]:
    wb = load_workbook(path, data_only=True)
    out: list[GangjinRow] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(4, ws.max_row + 1):
            name = collapse(str(ws.cell(r, 1).value or ""))
            addr = collapse(str(ws.cell(r, 2).value or ""))
            if not name or not addr or name in ("상호", "판매소명"):
                continue
            has_trash = product_yes(ws.cell(r, 3).value)
            has_special = product_yes(ws.cell(r, 4).value)
            if not has_trash and not has_special:
                continue
            out.append(
                GangjinRow(
                    name=name,
                    addr_raw=addr,
                    has_trash=has_trash,
                    has_special=has_special,
                    sheet=sheet_name,
                )
            )
    return out, REF_DATE


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow_geo = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    rows, path_ref = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    seen: set[str] = set()
    geo_n = 0
    misses = 0

    for row in rows:
        display = normalize_addr(row.addr_raw)
        dk = f"{row.name}|{display}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, row.addr_raw)
        if ck in cache and not args.refresh:
            lat, lng, road, jibeon = cache[ck]
            geo_n += 1
        elif allow_geo:
            hit = resolve_geocode(row.addr_raw, row.name, key)  # type: ignore[arg-type]
            if not hit:
                print(f"[지오코딩 실패] {row.name}\t{display}", file=sys.stderr)
                misses += 1
                continue
            lat, lng, road, jibeon = hit.lat, hit.lng, hit.road, hit.jibeon
            cache[ck] = [lat, lng, road, jibeon]
            geo_n += 1
        else:
            misses += 1
            continue

        if not in_gangjin_bbox(lat, lng):
            print(f"[좌표 제외] {row.name}\t{display}", file=sys.stderr)
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"jeonnam-gangjin-trash-{rid}",
                "name": row.name,
                "lat": round(lat, 7),
                "lng": round(lng, 7),
                "roadAddress": road,
                "address": jibeon,
                "businessStatus": "영업",
                "hasTrashBag": row.has_trash,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": path_ref,
            }
        )

    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    if allow_geo:
        save_cache(cache)

    trash_n = sum(1 for x in out if x["hasTrashBag"])
    special_n = sum(1 for x in out if x["hasSpecialBag"])
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={path_ref}, 종량제 {trash_n}, 특수마대 {special_n}, "
        f"geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["강진군"])


if __name__ == "__main__":
    main()
