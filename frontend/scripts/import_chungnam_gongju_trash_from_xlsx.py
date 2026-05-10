#!/usr/bin/env python3
"""
충청남도 공주시 종량제봉투 판매소 xlsx → stores.chungnam-gongju-trash.json

시트 「종량제봉투 판매소」: 연번, 상호, 주소, 행정동, 전화번호 (+ 주소 카카오 지오코딩)

  pip install openpyxl
  python3 scripts/import_chungnam_gongju_trash_from_xlsx.py \\
    --input ~/Downloads/종량제봉투+판매소+현황\\(홈페이지+게시\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.chungnam-gongju-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-chungnam-gongju-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "종량제봉투+판매소+현황(홈페이지+게시).xlsx"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.07

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
_REF_ROW = re.compile(r"기준\s*:\s*(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})", re.I)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())  # type: ignore[arg-type]
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def in_gongju_area_bbox(lat: float, lng: float) -> bool:
    """계룡·인근 허용을 약간 넓힘."""
    return 36.20 <= lat <= 36.90 and 126.80 <= lng <= 127.40


def is_gongju_addr_blob(s: str) -> bool:
    return "공주시" in s or ("공주" in s and "충남" in s)


def excel_is_gongju(addr: str) -> bool:
    a = (addr or "").replace(" ", "").replace("\u3000", "")
    return "충청남도공주시" in a or "충남공주시" in a


def sheet_ref_date(ws) -> str:
    for r in ws.iter_rows(min_row=1, max_row=6, max_col=8, values_only=True):
        for cell in r or ():
            if cell is None:
                continue
            t = collapse(str(cell))
            m = _REF_ROW.search(t)
            if m:
                y, mo, da = m.groups()
                return f"{int(y):04d}-{int(mo):02d}-{int(da):02d}"
    return "2025-10-20"


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    """헤더 행(연번·상호·주소) → 1-based row 번호 및 열명→인덱스."""
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1
    ):
        if not row:
            continue
        texts = [(str(x).strip() if x is not None else "") for x in row]
        if "연번" in texts and "상호" in texts and "주소" in texts:
            col = {}
            for j, h in enumerate(texts):
                if h in ("연번", "상호", "주소", "행정동", "전화번호"):
                    col[h] = j
            if "연번" in col and "상호" in col and "주소" in col:
                return i, col
    raise SystemExit("헤더 행(연번·상호·주소)을 찾지 못했습니다.")


def load_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_addr_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("address"))
    touch(d.get("road_address"))
    return " ".join(parts)


def addr_region_ok(blob: str) -> bool:
    if "충청남도" not in blob and "충남" not in blob:
        return False
    return is_gongju_addr_blob(blob)


def kakao_address(
    query: str, key: str
) -> tuple[float | None, float | None]:
    req = urllib.request.Request(
        f"{GEOCODE_URL}?{urllib.parse.urlencode({'query': query})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return None, None
    for d in data.get("documents") or []:
        lat = parse_float(d.get("y"))
        lng = parse_float(d.get("x"))
        if lat is None or lng is None:
            continue
        if not in_gongju_area_bbox(lat, lng):
            continue
        if not addr_region_ok(_doc_addr_blob(d)):
            continue
        return lat, lng
    return None, None


def kakao_keyword(query: str, key: str) -> tuple[float | None, float | None]:
    q = collapse(query)
    if not q:
        return None, None
    req = urllib.request.Request(
        f"{KEYWORD_URL}?{urllib.parse.urlencode({'query': q, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return None, None
    for d in data.get("documents") or []:
        lat = parse_float(d.get("y"))
        lng = parse_float(d.get("x"))
        if lat is None or lng is None:
            continue
        if not in_gongju_area_bbox(lat, lng):
            continue
        road = collapse(str(d.get("road_address_name") or ""))
        jibeon = collapse(str(d.get("address_name") or ""))
        blob = f"{road} {jibeon}"
        if not addr_region_ok(blob):
            continue
        return lat, lng
    return None, None


def geocode_variants(addr: str, name: str, dong: str) -> list[str]:
    a = collapse(addr)
    dong = collapse(dong)
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(a)
    if a and "번지" in a:
        push(re.sub(r"\s*번지\s*", " ", a).strip())
    if a:
        push(re.sub(r"\s*\([^)]*\)\s*", " ", a))
    push(f"{name} {dong}")
    push(f"{name} 공주시")
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        print(
            "오류: KAKAO_REST_API_KEY 가 없습니다. frontend/.env.local 을 확인하거나 --skip-kakao",
            file=sys.stderr,
        )
        raise SystemExit(1)

    wb = load_workbook(inp, read_only=True, data_only=True)
    if "종량제봉투 판매소" not in wb.sheetnames:
        raise SystemExit("시트 「종량제봉투 판매소」 없음")
    ws = wb["종량제봉투 판매소"]
    ref_date = sheet_ref_date(ws)
    hdr_row, col = find_header_row(ws)
    ii, iname, iaddr = col["연번"], col["상호"], col["주소"]
    idong = col.get("행정동", -1)
    iphone = col.get("전화번호", -1)

    cache = load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0

    data_iter = ws.iter_rows(min_row=hdr_row + 1, values_only=True)
    for row in data_iter:
        if not row:
            continue
        name = collapse(str(row[iname] or ""))
        addr = collapse(str(row[iaddr] or ""))
        if not name or not addr:
            continue
        seq = row[ii]
        if seq is None or str(seq).strip() == "":
            continue
        if not excel_is_gongju(addr):
            continue

        dong = collapse(str(row[idong] if idong >= 0 and row[idong] else ""))
        ck = hashlib.sha1(f"gongju:{seq}:{addr}:{name}".encode()).hexdigest()[:28]
        if ck in cache and len(cache[ck]) == 2:
            lat, lng = float(cache[ck][0]), float(cache[ck][1])
        elif allow:
            lat = lng = None
            used_keyword = False
            for q in geocode_variants(addr, name, dong):
                la, ln = kakao_address(q, key)
                if la is None:
                    la, ln = kakao_keyword(q, key)
                    if la is not None:
                        used_keyword = True
                if la is not None:
                    lat, lng = la, ln
                    break
            if lat is None:
                misses += 1
                print(f"[geocode 실패] {name}\t{addr}", file=sys.stderr)
                continue
            cache[ck] = [float(lat), float(lng)]
            geo_n += 1
            if geo_n % 40 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)
        else:
            misses += 1
            continue

        rid_base = hashlib.sha1(f"{seq}\n{name}\n{addr}".encode()).hexdigest()[:20]
        sid = f"chungnam-gongju-trash-{rid_base}"
        rec = {
            "id": sid,
            "name": name,
            "lat": round(float(lat), 7),
            "lng": round(float(lng), 7),
            "roadAddress": addr,
            "address": collapse(f"충청남도 공주시 {dong}") if dong else "충청남도 공주시",
            "businessStatus": "영업",
            "hasTrashBag": True,
            "hasSpecialBag": False,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": ref_date,
        }
        if iphone >= 0:
            tel = collapse(str(row[iphone] or ""))
            if tel:
                rec["phone"] = tel
        out.append(rec)

    wb.close()
    save_cache(cache)

    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref_date={ref_date}, api_rows≈{geo_n}, 미매칭={misses})"
    )


if __name__ == "__main__":
    main()
