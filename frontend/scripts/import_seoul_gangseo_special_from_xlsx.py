#!/usr/bin/env python3
"""
서울 강서구 불연성마대 판매처 xlsx → stores.seoul-gangseo-special.json

시트 `거래처목록` (2행~):
  상호명 | 연락처 | 주소 | 특수마대 판매여부(Y) | 취급규격 | 데이터기준일자

  python3 scripts/import_seoul_gangseo_special_from_xlsx.py \
    --input ~/Downloads/강서구\ 불연성마대_정보공개청구16783921.xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.seoul-gangseo-special.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-seoul-gangseo-special.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "강서구 불연성마대_정보공개청구16783921.xlsx"
SHEET_NAME = "거래처목록"
DATA_START_ROW = 2
REF_DATE = "2026-06-05"
CACHE_VERSION = "v1-seoul-gangseo-special"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def ref_date_from_row(v: object) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = collapse(str(v or ""))
    m = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return REF_DATE


def ref_date_from_path(p: Path) -> str:
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def in_gangseo_bbox(lat: float, lng: float) -> bool:
    return 37.52 <= lat <= 37.59 and 126.78 <= lng <= 126.89


def gangseo_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    return "강서구" in t and "서울" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^서울시\s+", "서울특별시 ", a)
    if not a:
        return ""
    if a.startswith("서울특별시"):
        return a
    if a.startswith("서울 "):
        return "서울특별시 " + a[len("서울 ") :]
    if a.startswith("강서구"):
        return f"서울특별시 {a}"
    return f"서울특별시 강서구 {a}"


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s+\d+\s*층\s*", " ", a)
    a = re.sub(r"\s+\d+\s*호\s*$", "", a)
    a = re.sub(r"\s+지하\s*$", "", a)
    a = re.sub(r"\s+[A-Z]?\d+\s*상가.*$", "", a, flags=re.I)
    a = re.sub(r"\s+\d+단지상가.*$", "", a)
    a = re.sub(r"\s+상가.*$", "", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"([가-힣]+(?:로|길))(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+동)(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+리)(\d+)", r"\1 \2", a)
    return format_display_addr(a)


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    clean_name = re.sub(r"\([^)]*\)", "", name).strip()
    norm = normalize_addr(addr_raw)
    tail = norm
    for prefix in ("서울특별시 강서구 ", "서울특별시 ", "강서구 "):
        if tail.startswith(prefix):
            tail = collapse(tail[len(prefix) :])
            break
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"서울 강서구 {tail}")
    push(f"서울특별시 강서구 {tail}")
    if is_likely_jibeon(tail):
        token_m = re.match(r"^([가-힣0-9]+동)\s+(\d+(?:-\d+)?)", tail)
        if token_m:
            core_lot = f"{token_m.group(1)} {token_m.group(2)}"
            push(f"서울특별시 강서구 {core_lot}")
            push(f"서울 강서구 {core_lot}")
        push(re.sub(r"\s+[가-힣A-Za-z].*$", "", tail).strip())
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"서울특별시 강서구 {road_only}")
    if ROAD_RE.search(tail):
        push(tail)
    push(f"{name} 강서구")
    push(f"{name} 서울 강서구")
    if clean_name and clean_name != name:
        push(f"{clean_name} 강서구")
        push(f"{clean_name} 서울 강서구")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class GangseoRow:
    name: str
    addr_raw: str
    has_special: bool
    ref_date: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_gangseo_bbox(lat, lng):
        return None
    if not gangseo_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_gangseo_bbox(lat, lng):
        return None
    if not gangseo_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    tail = normalize_addr(addr_raw)
    m = re.search(r"([가-힣0-9]+동)", tail)
    if m:
        area = m.group(1)
        for q in (f"서울 강서구 {area}", f"서울특별시 강서구 {area}"):
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return None


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[GangseoRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    out: list[GangseoRow] = []
    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, 1).value)
        addr_raw = cell_str(ws.cell(r, 3).value)
        if not name or not addr_raw:
            continue
        special = cell_str(ws.cell(r, 4).value).upper() == "Y"
        if not special:
            continue
        out.append(
            GangseoRow(
                name=name,
                addr_raw=addr_raw,
                has_special=True,
                ref_date=ref_date_from_row(ws.cell(r, 6).value),
            )
        )
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    default_ref = ref_date_from_path(inp)
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    misses = 0
    geo_n = 0
    seen: set[str] = set()

    for row in rows:
        display_road = normalize_addr(row.addr_raw)
        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None
        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key)  # type: ignore[arg-type]
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)  # type: ignore[arg-type]
            hit.road = hit.road or cr or display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 30 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"seoul-gangseo-special-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road or display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": False,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": row.ref_date or default_ref,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={default_ref}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["강서구"])


if __name__ == "__main__":
    main()
