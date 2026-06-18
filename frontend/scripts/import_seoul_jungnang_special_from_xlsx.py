#!/usr/bin/env python3
"""
서울특별시 중랑구 특수마대(불연성) 판매소 xlsx → stores.seoul-jungnang-special.json

시트 `Sheet1` (헤더 4행, 데이터 5행~):
  연번(B) | 상호명(C) | 전화번호(D) | 도로명주소(E) | 지번주소(F)

매핑:
  - 목록 등재 -> hasSpecialBag: true
  - hasTrashBag: false

사용:
  cd frontend
  python3 scripts/import_seoul_jungnang_special_from_xlsx.py \\
    --input ~/Downloads/중랑구_특수마대\\ 판매소\\(260616\\ 기준\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.seoul-jungnang-special.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-seoul-jungnang-special.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "중랑구_특수마대 판매소(260616 기준).xlsx"
SHEET_NAME = "Sheet1"
DATA_START_ROW = 5
REF_DATE = "2026-06-16"
CACHE_VERSION = "v1-seoul-jungnang-special"

COL_NAME = 3
COL_PHONE = 4
COL_ROAD = 5
COL_JIBEON = 6

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def ref_date_from_path(p: Path) -> str:
    m = re.search(r"\(?(20)?(\d{2})(\d{2})(\d{2})\s*기준\)?", p.name)
    if m:
        yy = m.group(2)
        mm, dd = m.group(3), m.group(4)
        return f"20{yy}-{mm}-{dd}"
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def norm_phone(raw: str) -> str | None:
    t = collapse(raw)
    if not t:
        return None
    digits = re.sub(r"\D", "", t)
    if digits.startswith("02") and len(digits) == 10:
        return f"02-{digits[2:6]}-{digits[6:]}"
    if digits.startswith("02") and len(digits) == 9:
        return f"02-{digits[2:5]}-{digits[5:]}"
    return t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^서울시\s+", "서울특별시 ", a)
    if not a:
        return ""
    if a.startswith("서울특별시"):
        return a
    if a.startswith("서울 "):
        return "서울특별시 " + a[len("서울 ") :]
    if a.startswith("중랑구"):
        return f"서울특별시 {a}"
    return f"서울특별시 중랑구 {a}"


def normalize_road(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if not a:
        return ""
    a = re.sub(r"([가-힣]+로)(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+)(번길)", r"\1\2", a)
    a = re.sub(r"(\d+번길)(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+(?:로|길))(\d+)", r"\1 \2", a)
    return format_display_addr(a)


def strip_building_tail(a: str) -> str:
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+(?:,\s*\d+)?\s*호.*$", "", a)
    m = re.match(r"^(.+?(?:로|길|대로)\s*\d+(?:-\d+)?(?:번길)?)", a)
    if m:
        return collapse(m.group(1))
    m = re.match(r"^(.+?(?:로|길|대로)\d+(?:-\d+)?(?:번길)?)", a.replace(" ", ""))
    if m:
        return normalize_road(m.group(1))
    return collapse(a)


def geocode_target(road: str) -> str:
    return strip_building_tail(normalize_road(road))


def jungnang_tail(full: str) -> str:
    for prefix in (
        "서울특별시 중랑구 ",
        "서울특별시 ",
        "서울 중랑구 ",
        "중랑구 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def in_jungnang_bbox(lat: float, lng: float) -> bool:
    return 37.57 <= lat <= 37.63 and 127.04 <= lng <= 127.12


def jungnang_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    return "중랑구" in t or (("서울" in t or "서울특별시" in t) and "중랑" in t)


def geocode_query_variants(road_raw: str, jibeon_raw: str, name: str) -> list[str]:
    road = normalize_road(road_raw)
    jibeon = format_display_addr(jibeon_raw) if jibeon_raw else ""
    target = geocode_target(road_raw)
    tail = jungnang_tail(target)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(target)
    push(road)
    push(f"서울특별시 중랑구 {tail}")
    push(f"서울 중랑구 {tail}")
    compact = target.replace(" ", "")
    push(compact)
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    push(re.sub(r"((?:로|길|대로))(\d+)", r"\1 \2", compact))
    if jibeon:
        push(jibeon)
        jt = jungnang_tail(jibeon)
        push(f"서울특별시 중랑구 {jt}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"서울특별시 중랑구 {road_only}")
    push(f"{name} 중랑구")
    push(f"{name} 서울 중랑구")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class JungnangRow:
    name: str
    road_raw: str
    jibeon_raw: str
    phone: str
    display_road: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_jungnang_bbox(lat, lng):
        return None
    if not jungnang_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_jungnang_bbox(lat, lng):
        return None
    if not jungnang_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(key: str, display_road: str) -> GeoHit | None:
    for q in ("서울 중랑구", "서울특별시 중랑구"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, _ = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(
    road_raw: str, jibeon_raw: str, name: str, key: str, display_road: str
) -> GeoHit | None:
    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", display_road))
    for q in geocode_query_variants(road_raw, jibeon_raw, name):
        if has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
        if not has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return area_fallback(key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[JungnangRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    merged: dict[str, JungnangRow] = {}

    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        num = ws.cell(r, 2).value
        if not isinstance(num, (int, float)):
            continue
        name = cell_str(ws.cell(r, COL_NAME).value)
        road_raw = cell_str(ws.cell(r, COL_ROAD).value)
        if not name or not road_raw:
            continue
        if "중랑" not in road_raw.replace(" ", ""):
            continue
        jibeon_raw = cell_str(ws.cell(r, COL_JIBEON).value)
        phone = norm_phone(cell_str(ws.cell(r, COL_PHONE).value)) or ""
        display = normalize_road(road_raw)
        k = f"{name}|{display}"
        if k not in merged:
            merged[k] = JungnangRow(
                name=name,
                road_raw=road_raw,
                jibeon_raw=jibeon_raw,
                phone=phone,
                display_road=display,
            )

    wb.close()
    return list(merged.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    misses = 0
    geo_n = 0
    seen: set[str] = set()

    for row in rows:
        display_road = row.display_road
        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(row.road_raw, row.jibeon_raw, row.name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = display_road
            jibeon = format_display_addr(row.jibeon_raw) if row.jibeon_raw else ""
            hit.jibeon = jibeon or cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 25 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        store: dict = {
            "id": f"seoul-jungnang-special-{rid}",
            "name": row.name,
            "lat": round(float(hit.lat), 7),
            "lng": round(float(hit.lng), 7),
            "roadAddress": display_road,
            "address": hit.jibeon or display_road,
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": False,
            "adminVerified": False,
            "dataReferenceDate": ref_date,
        }
        if row.phone:
            store["phone"] = row.phone
        out.append(store)

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["중랑구"])


if __name__ == "__main__":
    main()
