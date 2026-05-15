#!/usr/bin/env python3
"""
서울 양천구 특수규격봉투(불연성 PP마대) 판매처 엑셀
→ public/data/stores.seoul-yangcheon-special.json

- 시트: 첫 시트 (헤더 행에 「대행업체」)
- 열: 대행업체, 대행구역, 판매점, 주변지역, 연락처, 주소
- 카카오 주소/키워드 검색 → 위·경도 (실패 시 Nominatim)
- hasTrashBag: false, hasSpecialBag: true, storeCategory: nonBurnable

사용 (frontend 디렉터리):
  python3 scripts/import_seoul_yangcheon_special_xlsx.py \\
    ~/Downloads/특수규격봉투\\ 마대판매처\\(양천구\\).xlsx
  npx tsx scripts/assignShortCodes.ts
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

SHORT_CODE_RE = re.compile(r"^[a-zA-Z0-9]{6}$")

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as e:
    raise SystemExit("openpyxl이 필요합니다. pip install openpyxl") from e


def norm_phone(raw: str | None) -> str | None:
    if raw is None:
        return None
    t = str(raw).replace("\xa0", " ").strip()
    if not t:
        return None
    return t


def full_road_address(raw: str) -> str:
    a = str(raw).replace("\xa0", " ").strip()
    a = re.sub(r"\s+", " ", a)
    if a.startswith("서울특별시 양천구"):
        return a
    if a.startswith("서울시 양천구"):
        return "서울특별시 양천구 " + a[len("서울시 양천구") :].lstrip()
    if a.startswith("서울 양천구"):
        return "서울특별시 양천구 " + a[len("서울 양천구") :].lstrip()
    if a.startswith("양천구"):
        return "서울특별시 " + a
    return "서울특별시 양천구 " + a


def read_excel_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    start = 0
    for i, r in enumerate(rows):
        if r and r[0] == "대행업체":
            start = i + 1
            break
    out: list[dict] = []
    for r in rows[start:]:
        if not r or len(r) < 6:
            continue
        name = r[2]
        addr = r[5]
        if not name or not addr:
            continue
        name = str(name).replace("\n", " ").strip()
        addr = str(addr).replace("\n", " ").strip()
        if not name or not addr:
            continue
        phone = norm_phone(r[4] if len(r) > 4 else None)
        out.append(
            {
                "name": name,
                "roadAddress": full_road_address(addr),
                "phone": phone,
            }
        )
    return out


UA = "SseubongmapYangcheonSpecialImport/1.0 (internal data tooling)"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-yangcheon-special.json"


def load_geo_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        raw = json.loads(CACHE_PATH.read_text(encoding="utf8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, list[float]] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if isinstance(v, list) and len(v) == 2:
                out[str(k)] = [float(v[0]), float(v[1])]
    return out


def save_geo_cache(cache: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=0), encoding="utf8")


def load_kakao_rest_key() -> str | None:
    for env in ("KAKAO_REST_API_KEY", "KAKAO_REST_KEY"):
        k = os.environ.get(env)
        if k and str(k).strip():
            return str(k).strip()
    p = Path(__file__).resolve().parent.parent / ".env.local"
    if not p.exists():
        return None
    for line in p.read_text(encoding="utf8").splitlines():
        line = line.strip()
        for key in ("KAKAO_REST_API_KEY=", "KAKAO_REST_KEY="):
            if line.startswith(key):
                v = line.split("=", 1)[1].strip().strip('"').strip("'")
                return v or None
    return None


KAKAO_ADDR_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KAKAO_KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"


def kakao_geocode(q: str, api_key: str) -> tuple[float | None, float | None]:
    time.sleep(0.18)
    for base, extra in (
        (KAKAO_ADDR_URL, {"query": q}),
        (KAKAO_KEYWORD_URL, {"query": q, "size": "1"}),
    ):
        qs = urllib.parse.urlencode(extra)
        req = urllib.request.Request(
            f"{base}?{qs}", headers={"Authorization": f"KakaoAK {api_key}"}
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except (OSError, ValueError):
            continue
        docs = data.get("documents") or []
        if not docs:
            continue
        try:
            lat = float(docs[0].get("y"))
            lng = float(docs[0].get("x"))
        except (TypeError, ValueError):
            continue
        if lat and lng:
            return lat, lng
    return None, None


def nominatim(q: str) -> tuple[float | None, float | None]:
    params = urllib.parse.urlencode({"q": q, "format": "json", "limit": "1"})
    url = "https://nominatim.openstreetmap.org/search?" + params
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=45) as resp:
        data = json.loads(resp.read().decode())
    time.sleep(1.12)
    if data:
        return float(data[0]["lat"]), float(data[0]["lon"])
    return None, None


def road_variants(road: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(x: str) -> None:
        t = " ".join(x.replace("\xa0", " ").split())
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    add(road)
    head = road.split(",")[0].strip()
    add(head)
    if "," in road:
        add(re.sub(r",\s*\d+층.*$", "", road).strip())
    head = re.sub(r"\s+\d+층\s*$", "", head).strip()
    add(head)
    add(re.sub(r"\s+\d+동\s*$", "", head).strip())
    add(re.sub(r"\s+\d+호\s*$", "", head).strip())
    add(re.sub(r"\s+[가-힣0-9]+상가.*$", "", head).strip())
    add(re.sub(r"\s+B\d+.*$", "", head, flags=re.I).strip())
    add(re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", head).strip())
    add(re.sub(r"(로|길)(\d{2,})$", r"\1 \2", head).strip())
    add(re.sub(r"([가-힣]+로)(\d)", r"\1 \2", head).strip())
    add(re.sub(r"(\d+길)(\d+)$", r"\1 \2", head).strip())
    gl = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", head)
    add(gl)
    add(re.sub(r"(\d+길)(\d+)$", r"\1 \2", gl).strip())
    return out


def geocode_row(p: dict, cache: dict[str, list[float]], kakao_key: str | None) -> tuple[float, float]:
    ck = p["roadAddress"] + "\t" + p["name"]
    if ck in cache:
        lat, lng = cache[ck]
        return lat, lng

    name = p["name"]
    if kakao_key:
        for base in road_variants(p["roadAddress"]):
            for q in (base, f"{name} {base}", f"서울특별시 양천구 {name}"):
                lat, lng = kakao_geocode(q, kakao_key)
                if lat is not None and lng is not None:
                    cache[ck] = [lat, lng]
                    save_geo_cache(cache)
                    return lat, lng
    for base in road_variants(p["roadAddress"]):
        for q in (
            base,
            f"{name} {base}",
            f"{name} 서울특별시 양천구",
        ):
            lat, lng = nominatim(q)
            if lat is not None and lng is not None:
                cache[ck] = [lat, lng]
                save_geo_cache(cache)
                return lat, lng
    raise RuntimeError(f"지오코딩 실패: {json.dumps(p, ensure_ascii=False)}")


def load_existing_by_id(existing_path: Path) -> dict[str, dict]:
    if not existing_path.exists():
        return {}
    try:
        data = json.loads(existing_path.read_text(encoding="utf8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(data, list):
        return {}
    out: dict[str, dict] = {}
    for row in data:
        if not isinstance(row, dict) or not row.get("id"):
            continue
        rid = str(row["id"])
        meta: dict = {}
        sc = row.get("shortCode")
        if isinstance(sc, str):
            sc = sc.strip()
            if SHORT_CODE_RE.match(sc):
                meta["shortCode"] = sc
        dr = row.get("dataReferenceDate")
        if isinstance(dr, str) and dr.strip():
            meta["dataReferenceDate"] = dr.strip()
        if meta:
            out[rid] = meta
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help="특수규격봉투 마대판매처(양천구).xlsx")
    ap.add_argument(
        "-o",
        "--out",
        type=Path,
        default=Path("public/data/stores.seoul-yangcheon-special.json"),
    )
    args = ap.parse_args()

    parsed = read_excel_rows(args.xlsx)
    if not parsed:
        raise SystemExit("엑셀에서 유효한 데이터 행이 없습니다.")

    out_abs = (
        args.out
        if args.out.is_absolute()
        else Path(__file__).resolve().parent.parent / args.out
    )
    existing_meta = load_existing_by_id(out_abs)
    geo_cache = load_geo_cache()
    kakao_key = load_kakao_rest_key()
    ref_date = "2026-01-01"
    bundle: list[dict] = []

    for i, p in enumerate(parsed, start=1):
        sid = f"yangcheon-sp-{i:03d}"
        print(f"[{i}/{len(parsed)}] {p['name'][:40]}…", flush=True)
        lat, lng = geocode_row(p, geo_cache, kakao_key)
        row: dict = {
            "id": sid,
            "name": p["name"],
            "lat": round(lat, 7),
            "lng": round(lng, 7),
            "roadAddress": p["roadAddress"],
            "address": "서울특별시 양천구",
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": existing_meta.get(sid, {}).get(
                "dataReferenceDate", ref_date
            ),
            "storeCategory": "nonBurnable",
        }
        if p.get("phone"):
            row["phone"] = p["phone"]
        sc = existing_meta.get(sid, {}).get("shortCode")
        if sc:
            row["shortCode"] = sc
        bundle.append(row)

    out_abs.parent.mkdir(parents=True, exist_ok=True)
    out_abs.write_text(
        json.dumps(bundle, ensure_ascii=False, indent=2) + "\n", encoding="utf8"
    )
    print(f"wrote {len(bundle)} rows -> {out_abs}")


if __name__ == "__main__":
    main()
