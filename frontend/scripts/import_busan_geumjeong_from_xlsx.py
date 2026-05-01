#!/usr/bin/env python3
"""
부산 금정구 (게시) 종량제·불연성마대 판매소 xlsx → JSON 2종
  - 종량제: 연번 · 동명 · 도로명주소 · 상호 → hasTrashBag
  - 불연성 : 연번 · 동명 · 도로명주소 · 상호 · 연락처 → hasSpecialBag

  python3 scripts/import_busan_geumjeong_from_xlsx.py
  python3 scripts/import_busan_geumjeong_from_xlsx.py \\
    --trash-xlsx ~/Downloads/'(게시)종량제봉투판매소현황(20260401).xlsx' \\
    --special-xlsx ~/Downloads/'(게시)불연성마대판매소현황(20260401).xlsx'
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
OUT_TRASH = FRONTEND / "public" / "data" / "stores.busan-geumjeong-trash.json"
OUT_SPECIAL = FRONTEND / "public" / "data" / "stores.busan-geumjeong-special.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-busan-geumjeong.json"

USER_AGENT = "Mozilla/5.0 (compatible; VinylMapImport/1.0; +https://github.com/totald369/vinyl)"
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=12) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=12) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def parse_seq(v) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    try:
        return int(float(s))
    except ValueError:
        return None


def collapse(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def full_busan_geumjeong(raw: str, dong: str) -> str:
    """금정 행 대부분: 도로명 + 동명. 일부 행만 「동명」에 '동래구'가 들어가고 주소가 동래 행정구역."""
    s = collapse(raw)
    if not s:
        return s
    d = collapse(dong or "")

    if s.startswith("부산광역시"):
        return s
    if s.startswith("부산시"):
        return collapse("부산광역시 " + s[3:].lstrip())

    # 종량 제 마지막 구간 등: 「동명=동래구」 + 주소가 동래구 …
    if d == "동래구" and s.startswith("동래구"):
        return collapse(f"부산광역시 {s}")

    # 명 오타 추정 (지도 검색용)
    if "시실로" in s:
        s = collapse(s.replace("시실로", "세실로"))

    rd = s
    if d and rd.startswith(d) and len(rd) > len(d) and rd[len(d)] in ("로", "길", "대"):
        return collapse(f"부산광역시 금정구 {rd}")
    return collapse(f"부산광역시 금정구 {d} {rd}".strip())


def geocode_query_variants(full_addr: str) -> list[str]:
    def _tail_paren_strip(s: str) -> str:
        return collapse(re.sub(r"\s*[\(（][^)）]*[\)）]\s*$", "", s))

    def add(q: str, out: list[str]) -> None:
        cq = collapse(q)
        if cq and cq not in out:
            out.append(cq)

    seen: list[str] = []

    collapsed = collapse(full_addr)
    comma_head = collapsed.split(",", 1)[0].strip()
    stripped_tail = _tail_paren_strip(collapsed)

    for q in (
        collapsed,
        comma_head,
        stripped_tail,
        _tail_paren_strip(comma_head),
    ):
        add(q, seen)

    base = comma_head or stripped_tail or collapsed

    spaced_alley = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced_alley != base:
        add(spaced_alley, seen)
        base = spaced_alley
    ro_gil = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", base)
    if ro_gil != base:
        add(ro_gil, seen)
        base = ro_gil
    ro_only = re.sub(
        r"^(부산광역시\s+금정구\s+[가-힣\d\s]+?(?:로|대로))\s+\d.+",
        r"\1",
        base,
    )
    if ro_only != base:
        add(ro_only, seen)
    return seen


def resolve_coords(
    road_address: str, place_name: str, dong: str, cache: dict, key: str
) -> tuple[float, float] | None:
    for qv in geocode_query_variants(road_address):
        c = kakao_geocode(qv, cache, key)
        if c:
            return c
    c = kakao_geocode(f"{road_address} {place_name}", cache, key)
    if c:
        return c
    for qv in geocode_query_variants(road_address):
        c = kakao_keyword_geocode(qv, cache, key)
        if c:
            return c
    d = (dong or "").strip()
    name_bits = {place_name}
    if "(" in place_name:
        name_bits.add(re.sub(r"\([^)]*\)", "", place_name).strip())
    for base in name_bits:
        for q in (
            f"부산 금정구 {base}",
            f"금정구 {base}",
            f"부산광역시 금정구 {d} {base}",
            f"부산 {base}",
            base,
        ):
            if len(q.strip()) < 2:
                continue
            c = kakao_keyword_geocode(q, cache, key)
            if c:
                return c
    return None


def ref_date_default() -> str:
    return "2026-04-01"


def normalize_phone(raw: str) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    s = collapse(str(raw)).replace(" ", "")
    if re.match(r"^051-\d{3}-\d{4}$", s) or re.match(r"^051-\d{4}-\d{4}$", s):
        return s
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if digits.startswith("051") and len(digits) == 11:
        return f"051-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 10 and digits.startswith("051"):
        return f"051-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 7:
        return f"051-{digits[:3]}-{digits[3:]}"
    return s


def parse_trash_xlsx(path: Path) -> list[tuple[int, str, str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "판매소현황" not in wb.sheetnames:
        raise ValueError(f"시트 없음: 판매소현황 in {path.name}")
    sh = wb["판매소현황"]
    out: list[tuple[int, str, str, str]] = []
    dong_fill = ""
    for row in sh.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        seq = parse_seq(row[0])
        dong_cell = collapse(str(row[1] or "").replace("\u200b", ""))
        if dong_cell:
            dong_fill = dong_cell
        addr_raw = collapse(str(row[2] or ""))
        name = collapse(str(row[3] or ""))
        if seq is None or not name or not addr_raw:
            continue
        dong_use = dong_cell or dong_fill
        out.append((seq, dong_use, addr_raw, name))
    wb.close()
    return out


def parse_special_xlsx(path: Path) -> list[tuple[int, str, str, str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sh = wb[wb.sheetnames[0]]
    out: list[tuple[int, str, str, str, str]] = []
    dong_fill = ""
    for row in sh.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        seq = parse_seq(row[0])
        dong_cell = collapse(str(row[1] or "").replace("\u200b", ""))
        if dong_cell:
            dong_fill = dong_cell
        addr_raw = collapse(str(row[2] or ""))
        name = collapse(str(row[3] or ""))
        phone = collapse(str(row[4] or ""))
        if seq is None or not name or not addr_raw:
            continue
        dong_use = dong_cell or dong_fill
        out.append((seq, dong_use, addr_raw, name, phone))
    wb.close()
    return out


def run_geocode_set(
    label: str,
    rows: list[tuple[int, ...]],
    id_prefix: str,
    *,
    has_trash: bool,
    has_special: bool,
    ref_date: str,
    phone_idx: bool,
) -> tuple[list[dict], list[str]]:
    cache = load_cache()
    stores: list[dict] = []
    failed: list[str] = []

    for i, tup in enumerate(rows):
        seq = tup[0]
        dong = tup[1]
        addr_raw = tup[2]
        name = tup[3]

        oid = f"{id_prefix}-{seq}"
        road = full_busan_geumjeong(addr_raw, dong)

        lat = lng = None
        if KAKAO_REST_KEY:
            c = resolve_coords(road, name, dong, cache, KAKAO_REST_KEY)
            if c:
                lat, lng = c

        if lat is None or lng is None:
            failed.append(f"{oid} {name} | {road}")

        row: dict = {
            "id": oid,
            "name": name,
            "roadAddress": road,
            "address": road,
            "businessStatus": "영업",
            "hasTrashBag": has_trash,
            "hasSpecialBag": has_special,
            "hasLargeWasteSticker": False,
            "adminVerified": False,
            "dataReferenceDate": ref_date,
        }
        if phone_idx:
            pn = normalize_phone(str(tup[4]))
            if pn:
                row["phone"] = pn

        if lat is not None and lng is not None:
            row["lat"] = lat
            row["lng"] = lng

        stores.append(row)
        if (i + 1) % 65 == 0:
            save_cache(cache)
            print(f"    [{label}] {i+1}/{len(rows)}", file=sys.stderr)

    save_cache(cache)
    return stores, failed


def main():
    dl = Path.home() / "Downloads"
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--trash-xlsx",
        type=Path,
        default=dl / "(게시)종량제봉투판매소현황(20260401).xlsx",
    )
    ap.add_argument(
        "--special-xlsx",
        type=Path,
        default=dl / "(게시)불연성마대판매소현황(20260401).xlsx",
    )
    ap.add_argument(
        "--ref-date",
        type=str,
        default=ref_date_default(),
        help="dataReferenceDate (YYYY-MM-DD)",
    )
    args = ap.parse_args()

    trx = args.trash_xlsx.expanduser()
    spx = args.special_xlsx.expanduser()
    if not trx.exists():
        print(f"종량제 파일 없음: {trx}", file=sys.stderr)
        sys.exit(1)
    if not spx.exists():
        print(f"불연성 파일 없음: {spx}", file=sys.stderr)
        sys.exit(1)

    if not KAKAO_REST_KEY:
        print("KAKAO_REST_API_KEY 없음.", file=sys.stderr)

    print("[parse] 종량제 xlsx...", file=sys.stderr)
    trash_rows = parse_trash_xlsx(trx)
    print(f"  → {len(trash_rows)}행", file=sys.stderr)

    print("[parse] 불연성 xlsx...", file=sys.stderr)
    spec_rows = parse_special_xlsx(spx)
    print(f"  → {len(spec_rows)}행", file=sys.stderr)

    ref_date = args.ref_date.strip()

    print("[geocode] 종량제", file=sys.stderr)
    trash_stores, tfail = run_geocode_set(
        "trash",
        trash_rows,
        "busan-geumjeong-trash",
        has_trash=True,
        has_special=False,
        ref_date=ref_date,
        phone_idx=False,
    )
    ot = sum(1 for s in trash_stores if "lat" in s)
    print(f"  좌표 {ot}/{len(trash_stores)} 실패 {len(tfail)}", file=sys.stderr)
    for ln in tfail[:20]:
        print(f"    {ln}", file=sys.stderr)

    print("[geocode] 불연성", file=sys.stderr)
    spec_stores, sfail = run_geocode_set(
        "special",
        spec_rows,
        "busan-geumjeong-special",
        has_trash=False,
        has_special=True,
        ref_date=ref_date,
        phone_idx=True,
    )
    ospec = sum(1 for s in spec_stores if "lat" in s)
    print(f"  좌표 {ospec}/{len(spec_stores)} 실패 {len(sfail)}", file=sys.stderr)
    for ln in sfail[:20]:
        print(f"    {ln}", file=sys.stderr)

    OUT_TRASH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_TRASH, "w", encoding="utf-8") as f:
        json.dump(trash_stores, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(OUT_SPECIAL, "w", encoding="utf-8") as f:
        json.dump(spec_stores, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"저장: {OUT_TRASH}", file=sys.stderr)
    print(f"저장: {OUT_SPECIAL}", file=sys.stderr)


if __name__ == "__main__":
    main()
