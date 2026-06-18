#!/usr/bin/env python3
"""
서울특별시 서초구 특수규격봉투(불연성마대) 판매소 xlsx → stores.seoul-seocho-special.json

시트 `특수규격봉투취급점` (헤더 2행, 데이터 3행~):
  No(A) | 상호(B) | 소재지(C) | 규격(D)

매핑:
  - 목록 등재 -> hasSpecialBag: true
  - hasTrashBag: false

사용:
  cd frontend
  python3 scripts/import_seoul_seocho_special_from_xlsx.py \\
    --input ~/Downloads/서초구_특수규격봉투판매처\\ 명단.xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.seoul-seocho-special.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-seoul-seocho-special.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "서초구_특수규격봉투판매처 명단.xlsx"
SHEET_NAME = "특수규격봉투취급점"
DATA_START_ROW = 3
REF_DATE = "2026-06-16"
CACHE_VERSION = "v1-seoul-seocho-special"

COL_NAME = 2
COL_ADDR = 3

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
SEOCHO_DONGS = (
    "방배동",
    "양재동",
    "우면동",
    "원지동",
    "잠원동",
    "반포동",
    "서초동",
    "내곡동",
    "염곡동",
    "신원동",
)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def ref_date_from_path(p: Path) -> str:
    m = re.search(r"20(\d{2})\.(\d{1,2})\.(\d{1,2})", p.name)
    if m:
        return f"20{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return REF_DATE


def dong_hint(addr_raw: str) -> str:
    m = re.search(r"\(([^)]+)\)", addr_raw)
    if m:
        hint = collapse(m.group(1).split(",")[0])
        if hint.endswith("동"):
            return hint
    for d in SEOCHO_DONGS:
        if d in addr_raw:
            return d
    m = re.match(r"^([가-힣]+동)", collapse(addr_raw))
    if m:
        return m.group(1)
    return ""


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^서울시\s+", "서울특별시 ", a)
    if not a:
        return ""
    if a.startswith("서울특별시"):
        return a
    if a.startswith("서울 "):
        return "서울특별시 " + a[len("서울 ") :]
    if a.startswith("서초구"):
        return f"서울특별시 {a}"
    return f"서울특별시 서초구 {a}"


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if not a:
        return ""
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+(?:,\s*\d+)?\s*호.*$", "", a)
    a = re.sub(r"([가-힣]+(?:로|길))(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+번길)(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+)(번길)", r"\1\2", a)
    a = re.sub(r"([가-힣]+동)(\s*)(\d+)", r"\1 \3", a)
    hint = dong_hint(addr_raw)
    if hint and hint not in a.replace(" ", ""):
        if re.search(r"\d", a) and not a.startswith("서울"):
            a = f"{hint} {a}"
    if not a.startswith("서울") and not a.startswith("서초"):
        a = f"서초구 {a}"
    return format_display_addr(a)


def strip_building_tail(a: str) -> str:
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+(?:,\s*\d+)?\s*호.*$", "", a)
    m = re.match(r"^(.+?(?:로|길|대로)\s*\d+(?:-\d+)?)", a)
    if m:
        return collapse(m.group(1))
    m = re.match(r"^(.+?동\s+\d+(?:-\d+)?)", a)
    if m:
        return collapse(m.group(1))
    return collapse(a)


def geocode_target(addr_raw: str) -> str:
    return strip_building_tail(normalize_addr(addr_raw))


def seocho_tail(full: str) -> str:
    for prefix in (
        "서울특별시 서초구 ",
        "서울특별시 ",
        "서울 서초구 ",
        "서초구 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def in_seocho_bbox(lat: float, lng: float) -> bool:
    return 37.44 <= lat <= 37.51 and 126.97 <= lng <= 127.09


def seocho_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if "서초구" in t:
        return True
    return ("서울" in t or "서울특별시" in t) and "서초" in t


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    norm = normalize_addr(addr_raw)
    target = geocode_target(addr_raw)
    tail = seocho_tail(target)
    hint = dong_hint(addr_raw)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(target)
    push(norm)
    push(f"서울특별시 서초구 {tail}")
    push(f"서울 서초구 {tail}")
    compact = target.replace(" ", "")
    push(compact)
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    if hint:
        push(f"서울특별시 서초구 {hint}")
        push(f"서울 서초구 {hint} {tail}")
    push(f"{name} 서초구")
    push(f"{name} 서울 서초")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class SeochoRow:
    name: str
    addr_raw: str
    display_road: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_seocho_bbox(lat, lng):
        return None
    if not seocho_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_seocho_bbox(lat, lng):
        return None
    if not seocho_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(hint: str, key: str, display_road: str) -> GeoHit | None:
    area = hint or "서초동"
    for q in (f"서울 서초구 {area}", f"서울특별시 서초구 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, _ = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, display_road: str) -> GeoHit | None:
    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", display_road))
    hint = dong_hint(addr_raw)
    for q in geocode_query_variants(addr_raw, name):
        if has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
        if not has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return area_fallback(hint, key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[SeochoRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    merged: dict[str, SeochoRow] = {}

    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, COL_NAME).value)
        addr_raw = cell_str(ws.cell(r, COL_ADDR).value)
        if not name or not addr_raw or name in ("상호",):
            continue
        display = normalize_addr(addr_raw)
        k = f"{name}|{display}"
        if k not in merged:
            merged[k] = SeochoRow(name=name, addr_raw=addr_raw, display_road=display)

    wb.close()
    return list(merged.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    try:
        wb = load_workbook(inp, read_only=True, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for cell in row:
                if cell and re.search(r"20\d{2}\.\d{1,2}\.\d{1,2}", str(cell)):
                    m = re.search(r"(20\d{2})\.(\d{1,2})\.(\d{1,2})", str(cell))
                    if m:
                        ref_date = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        wb.close()
    except Exception:
        pass

    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    misses = 0
    geo_n = 0
    seen: set[str] = set()

    for row in rows:
        display_road = row.display_road
        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 20 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"seoul-seocho-special-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": False,
                "hasSpecialBag": True,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref_date,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["서초구"])


if __name__ == "__main__":
    main()
