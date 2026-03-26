#!/usr/bin/env python3
"""
Geocode Anyang Excel stores and merge them into stores.sample.json.
Deduplicates by name similarity against existing data.
"""

import json
import math
import time
import urllib.request
import urllib.parse
import urllib.error
import openpyxl
from pathlib import Path

ANYANG_XLSX = Path.home() / "Downloads" / "안양 판매소별 판매 현황(1월).xlsx"
OUT_PATH = Path(__file__).resolve().parent.parent / "public" / "data" / "stores.sample.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache.json"

KAKAO_REST_KEY = "fd1c94f46de9b58135650e9fba4b5320"
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"

def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None

def to_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0

def kakao_geocode(address):
    coords = _kakao_api(GEOCODE_URL, address)
    if coords:
        return coords
    return _kakao_api(KEYWORD_URL, address)

def _kakao_api(base_url, query):
    params = urllib.parse.urlencode({"query": query, "size": "1"})
    url = f"{base_url}?{params}"
    req = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {KAKAO_REST_KEY}"})
    try:
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    return (lat, lng)
    except Exception:
        pass
    return None

def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_cache(cache):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)

def main():
    print("Reading Anyang Excel...")
    wb = openpyxl.load_workbook(str(ANYANG_XLSX), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    data_rows = all_rows[1:]
    print(f"  → {len(data_rows)} rows")

    print("Loading existing stores.sample.json...")
    with open(OUT_PATH, "r", encoding="utf-8") as f:
        existing = json.load(f)
    print(f"  → {len(existing)} existing stores")

    existing_names = set()
    for s in existing:
        existing_names.add(s["name"].strip())

    max_id = max(int(s["id"]) for s in existing)

    cache = load_cache()
    anyang_stores = []
    geocoded = 0
    failed = 0
    skipped_dup = 0
    api_calls = 0

    for row in data_rows:
        name = str(row[3]).strip() if row[3] else ""
        addr = str(row[4]).strip() if row[4] else ""
        if not name or not addr:
            continue

        if name in existing_names:
            skipped_dup += 1
            continue

        trash_total = to_int(row[30])
        sticker_total = to_int(row[31])
        non_burn_20 = to_int(row[20])
        non_burn_50 = to_int(row[21])

        cache_key = addr
        if cache_key in cache:
            coords = cache[cache_key]
        else:
            coords = kakao_geocode(addr)
            cache[cache_key] = coords
            api_calls += 1
            if api_calls % 50 == 0:
                print(f"  API calls: {api_calls}, geocoded: {geocoded}, failed: {failed}")
                save_cache(cache)
            time.sleep(0.05)

        if not coords:
            failed += 1
            continue

        geocoded += 1
        max_id += 1

        store = {
            "id": str(max_id),
            "name": name,
            "lat": coords[0],
            "lng": coords[1],
            "address": addr,
            "businessStatus": "영업",
            "hasTrashBag": trash_total > 0,
            "hasSpecialBag": (non_burn_20 + non_burn_50) > 0,
            "hasLargeWasteSticker": sticker_total > 0,
            "adminVerified": False,
            "dataReferenceDate": "2025-01-31"
        }
        anyang_stores.append(store)

    save_cache(cache)

    print(f"\nAnyang geocoding complete:")
    print(f"  API calls: {api_calls}")
    print(f"  Geocoded: {geocoded}")
    print(f"  Failed: {failed}")
    print(f"  Skipped (dup): {skipped_dup}")

    merged = existing + anyang_stores
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    print(f"\nMerged total: {len(merged)} stores")
    print(f"  (기존 {len(existing)} + 안양 {len(anyang_stores)})")
    print(f"Written to: {OUT_PATH}")
    print(f"File size: {OUT_PATH.stat().st_size / 1024:.1f} KB")

    sb = sum(1 for s in merged if s.get("hasSpecialBag"))
    ls = sum(1 for s in merged if s.get("hasLargeWasteSticker"))
    anyang_count = sum(1 for s in merged if "안양" in (s.get("roadAddress","") + s.get("address","")))
    print(f"\n안양 매장 수: {anyang_count}")
    print(f"hasSpecialBag=true 전체: {sb}")
    print(f"hasLargeWasteSticker=true 전체: {ls}")

if __name__ == "__main__":
    main()
