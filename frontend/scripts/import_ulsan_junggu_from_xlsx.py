#!/usr/bin/env python3
"""
울산광역시 중구 종량제봉투·특수규격(불연성) 판매소 xlsx 2개 → JSON 2종

  python3 scripts/import_ulsan_junggu_from_xlsx.py \\
    --trash-xlsx ~/Downloads/종량제봉투판매소현황.xlsx \\
    --special-xlsx ~/Downloads/특수규격판매소현황.xlsx

종량제 시트: 첫 시트 (예: _20250407) — 번호, 판매소명, 사업장 주소
특수 시트: 첫 시트 (예: 기간별 판매현황_20250407) — 동, 판매소 명, 소재지

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
OUT_TRASH = FRONTEND / "public" / "data" / "stores.ulsan-junggu-trash.json"
OUT_SPECIAL = FRONTEND / "public" / "data" / "stores.ulsan-junggu-special.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-ulsan-junggu.json"
DL = Path.home() / "Downloads"
DEFAULT_TRASH = DL / "종량제봉투판매소현황.xlsx"
DEFAULT_SPECIAL = DL / "특수규격판매소현황.xlsx"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.07

# 데이터 기준일(시트명 20250407)
DATA_REF_DEFAULT = "2025-04-07"


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def collapse(s: str) -> str:
    return " ".join((s or "").replace("\xa0", " ").split())


WS_RE = re.compile(r"[ \t\r\n\v\f]+")


def normalize_addr(raw: str) -> str:
    """엑셀 주소 → 울산광역시 중구 … 도로명·지번 표기."""
    a = WS_RE.sub(" ", collapse(raw)).strip()
    if not a:
        return a
    a = re.sub(r"^\(주\)", "(주)", a)
    if a.startswith("울산광역시중구"):
        a = "울산광역시 중구 " + a[len("울산광역시중구") :].lstrip()
    if a.startswith("울산광역시 중구"):
        # 엑셀 오기: 중구 뒤에 타 구·군명
        a = re.sub(
            r"(울산광역시 중구)\s+(?:북구|동구|남구)\s+", r"\1 ", a,
        )
        a = re.sub(
            r"울산광역시 중구\s+북구\s+", "울산광역시 북구 ", a,
        )
        a = re.sub(r"울산광역시 중구\s+울주군", "울산광역시 울주군", a)
        m = re.search(r"(울산광역시 중구)([가-힣ㄱ-ㅎㅏ-ㅣ0-9])", a)
        if m and m.group(2):
            idx = m.start(2)
            a = a[:idx] + " " + a[idx:]
        if "(" in a and a.count("(") > a.count(")"):
            a = f"{a.rstrip()})"
        return _tweak_junggu_road_typos(WS_RE.sub(" ", a).strip())
    if a.startswith("중구"):
        return _tweak_junggu_road_typos(f"울산광역시 {a}".strip())
    if a.startswith("울산시 중구"):
        return _tweak_junggu_road_typos(
            ("울산광역시 " + a[len("울산시") :].lstrip()).strip()
        )
    if a.startswith("울산 중구"):
        return _tweak_junggu_road_typos(
            ("울산광역시 " + a[len("울산") :].lstrip()).strip()
        )
    if a.startswith("울산광역시") and "중구" not in a[:30]:
        if "울주군" in a[:42]:
            return WS_RE.sub(" ", a).strip()
        rest = a[len("울산광역시") :].strip()
        return f"울산광역시 중구 {rest}"
    if a.startswith("울주군"):
        return WS_RE.sub(" ", f"울산광역시 {a}".strip()).strip()
    if not a.startswith("울산"):
        return f"울산광역시 중구 {a}"
    out = WS_RE.sub(" ", a).strip()
    return _tweak_junggu_road_typos(out)


def _tweak_junggu_road_typos(a: str) -> str:
    if not a.startswith("울산광역시 중구"):
        return a
    out = a.replace("복산2동", "복산동")
    out = re.sub(r"복산동(\d)", r"복산동 \1", out)
    out = re.sub(r"당산길(\d)", r"당산길 \1", out)
    out = re.sub(r"내황(\d+길)", r"내황 \1", out)
    out = re.sub(r"내황(\d+길)(\d+)", r"내황 \1 \2", out)
    out = re.sub(r"반구정(\d+길)", r"반구정 \1", out)
    out = re.sub(r"손골(\d+길)", r"손골 \1", out)
    out = re.sub(r"종가(\d+길)", r"종가 \1", out)
    out = re.sub(r"학성(\d+길)", r"학성 \1", out)
    out = re.sub(r"종가로(\d)", r"종가로 \1", out)
    return WS_RE.sub(" ", out).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(v)  # type: ignore[arg-type]
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def in_ulsan_junggu_bbox(lat: float, lng: float) -> bool:
    """울산 중구·인접 도심(다운·성안·남외 등) + 지오코딩 오차(서측 일부 포함)."""
    return 35.505 <= lat <= 35.615 and 129.22 <= lng <= 129.385


def in_ulsan_bukgu_elampo_bbox(lat: float, lng: float) -> bool:
    """엑셀 오기 '중구 북구 진장·염포로' 등 북구 염포 인근."""
    return 35.52 <= lat <= 35.58 and 129.34 <= lng <= 129.42


def coord_acceptable_for_row(lat: float, lng: float, road: str) -> bool:
    r = road.replace(" ", "")
    if in_ulsan_junggu_bbox(lat, lng):
        return True
    if "염포로" in r and in_ulsan_bukgu_elampo_bbox(lat, lng):
        return True
    if "범서읍" in r and 35.34 <= lat <= 35.42 and 129.20 <= lng <= 129.30:
        return True
    return False


def kakao_geocode(q: str, key: str, road_hint: str = "") -> tuple[float | None, float | None]:
    hint = road_hint or q
    for base, extra in (
        (GEOCODE_URL, {"query": q}),
        (KEYWORD_URL, {"query": q, "size": "15"}),
    ):
        req = urllib.request.Request(
            f"{base}?{urllib.parse.urlencode(extra)}",
            headers={"Authorization": f"KakaoAK {key}"},
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except Exception:
            continue
        docs = data.get("documents") or []
        for d in docs:
            lat = parse_float(d.get("y"))
            lng = parse_float(d.get("x"))
            if lat is None or lng is None:
                continue
            if coord_acceptable_for_row(lat, lng, hint):
                return lat, lng
    return None, None


# 카카오 미등록 지번·블록 주소 — 동 단위 대표좌표로 보정(현장과 다를 수 있음).
_MANUAL_JUNGGU: tuple[tuple[str, str, float, float, str], ...] = (
    ("목화건재", "염포로200", 35.5545, 129.3595, "울산광역시 북구 염포로 200 (진장동 인근 · 수동보정)"),
    ("대지슈퍼", "내황5길51", 35.5518, 129.3338, "울산광역시 중구 내황5길 51 (반구동 · 수동보정)"),
    ("도연상사", "손골길26", 35.5565, 129.3285, "울산광역시 중구 손골길 26 (복산동 · 수동보정)"),
    ("송이스토아", "손골길26", 35.5565, 129.3285, "울산광역시 중구 손골길 26 (복산동 · 수동보정)"),
    ("cu번영로센트리지점", "복산1길1,5단지", 35.5596452, 129.3260182, "울산광역시 중구 복산1길 1 (번영로센트리지)"),
    ("cu울산센트리지점", "복산1길1,4단지", 35.5596452, 129.3260182, "울산광역시 중구 복산1길 1 (번영로센트리지)"),
    ("gs25리버스위트점", "화진길13-2", 35.5548446, 129.3086188, "울산광역시 중구 화진길 13-2 (리버스위트)"),
    ("lg유통", "금강타워102", 35.5606811, 129.3145204, "울산광역시 중구 우정길 14 (금강타워)"),
    ("경남식품", "푸름길35", 35.5580, 129.3270, "울산광역시 중구 푸름길 35 (복산동 · 수동보정)"),
    ("굿모닝마트", "태화동709-10", 35.5535, 129.3120, "울산광역시 중구 태화동 709-10 (수동보정)"),
    ("새복산유통", "옥교동길13", 35.5505, 129.3425, "울산광역시 중구 옥교동길 13 (북정동 · 수동보정)"),
    ("서광철물", "우정동274-4", 35.5620, 129.3185, "울산광역시 중구 우정동 274-4 (수동보정)"),
    ("서울마트", "복산동361-40", 35.5570, 129.3280, "울산광역시 중구 복산동 361-40 (수동보정)"),
    ("수전철물", "내황15길23", 35.5520, 129.3335, "울산광역시 중구 내황15길 23 (반구동 · 수동보정)"),
    ("씨유장현혁신점", "종가29길1", 35.58854, 129.34425, "울산광역시 중구 종가29길 1 (장현동)"),
    ("에이원식자재유통", "우정동28-4", 35.5615, 129.3190, "울산광역시 중구 우정동 28-4 (수동보정)"),
    ("연안슈퍼", "반구정17길7", 35.5530, 129.3360, "울산광역시 중구 반구정17길 7 (반구동 · 수동보정)"),
    ("울산원예농협하나로마트", "장검길10", 35.3845, 129.2440, "울산광역시 울주군 범서읍 장검길 10 (수동보정)"),
    ("웰빙할인마트", "손골4길15", 35.5568, 129.3290, "울산광역시 중구 손골4길 15 (복산동 · 수동보정)"),
    ("이마트24비즈파크점", "종가로406-21", 35.5895, 129.3435, "울산광역시 중구 종가로 406-21 (수동보정)"),
    ("인산매점", "남외운동장", 35.5640, 129.3480, "울산광역시 중구 남외동 남외운동장지구 (수동보정)"),
    ("일품마트", "학성로175", 35.5558, 129.3345, "울산광역시 중구 학성로 175 (학성동 · 수동보정)"),
    ("탑세일마트", "성안동45b", 35.5595, 129.3375, "울산광역시 중구 성안동 45블럭 (수동보정)"),
    ("태양할인마트", "당산길23", 35.5612, 129.3175, "울산광역시 중구 당산길 23 (우정동 · 수동보정)"),
    ("팝스토어복산점", "학성1길1", 35.5540, 129.3335, "울산광역시 중구 학성1길 1 (복산동 · 수동보정)"),
    ("한강할인마트", "운곡3길24", 35.55918, 129.27428, "울산광역시 중구 운곡3길 24 (다운동)"),
    ("한국미니스톱울산우정점", "당산길13", 35.5610, 129.3178, "울산광역시 중구 당산길 13 (우정동 · 수동보정)"),
)


def lookup_manual_junggu(name: str, road: str):
    n = collapse(name).lower().replace(" ", "").replace(",", "")
    r = collapse(road).replace(" ", "").lower()
    rows = sorted(_MANUAL_JUNGGU, key=lambda x: len(x[0]) + len(x[1]), reverse=True)
    for name_needle, road_needle, la, ln, road_out in rows:
        nd = collapse(name_needle).lower().replace(" ", "")
        if nd and nd not in n:
            continue
        rd = road_needle.replace(" ", "").lower()
        if rd and rd not in r:
            continue
        return (road_out, la, ln)
    return None


def addr_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(v: str) -> None:
        t = collapse(v)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    add(addr)
    head = re.sub(r"\(.*?\)", "", addr).strip()
    add(collapse(head))
    add(normalize_addr(head))
    return out


def load_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        raw = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, list[float]] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if isinstance(v, list) and len(v) == 2:
                out[str(k)] = [float(v[0]), float(v[1])]
    return out


def save_cache(cache: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def stable_id_digest(name: str, road: str) -> str:
    h = hashlib.sha1((name + "\x1e" + road).encode("utf-8")).hexdigest()[:12]
    return f"ulsan-junggu-{h}"


SHORT_CODE_RE = re.compile(r"^[a-zA-Z0-9]{6}$")


def load_existing_shortcodes(paths: tuple[Path, ...]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for p in paths:
        if not p.exists():
            continue
        try:
            rows = json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            rid = str(row.get("id") or "")
            sc = row.get("shortCode")
            if (
                isinstance(sc, str)
                and SHORT_CODE_RE.match(sc.strip())
                and rid
            ):
                mapping[rid] = sc.strip()
    return mapping


def read_trash_sheet(ws) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        try:
            n = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        if n < 1:
            continue
        name = collapse(str(row[1] or ""))
        addr_raw = collapse(str(row[2] or ""))
        if not name or not addr_raw:
            continue
        rows.append((name, addr_raw))
    return rows


def read_special_sheet(ws) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue
        c0 = collapse(str(row[0] or ""))
        name = collapse(str(row[1] or ""))
        addr_raw = collapse(str(row[2] or ""))
        if c0 == "동" and "판매" in name:
            continue
        if not name or not addr_raw:
            continue
        if name in ("판매소 명", "판매소명"):
            continue
        rows.append((name, addr_raw))
    return rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--trash-xlsx", type=Path, default=DEFAULT_TRASH)
    ap.add_argument("--special-xlsx", type=Path, default=DEFAULT_SPECIAL)
    args = ap.parse_args()
    trx = args.trash_xlsx.expanduser()
    spx = args.special_xlsx.expanduser()
    for p, label in ((trx, "종량제"), (spx, "특수")):
        if not p.exists():
            raise SystemExit(f"{label} xlsx 없음: {p}")
    key = load_kakao_key()
    if not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    wb_t = load_workbook(trx, read_only=True, data_only=True)
    wb_s = load_workbook(spx, read_only=True, data_only=True)
    trash_in = read_trash_sheet(wb_t[wb_t.sheetnames[0]])
    special_in = read_special_sheet(wb_s[wb_s.sheetnames[0]])
    wb_t.close()
    wb_s.close()

    combined: dict[str, dict[str, object]] = {}

    def feed(name: str, addr_raw: str, *, trash: bool, special: bool) -> None:
        road = normalize_addr(addr_raw)
        kname = collapse(name.lower().replace("\xa0", ""))
        kfix = "|".join((kname, road.replace(" ", "")))
        if kfix not in combined:
            combined[kfix] = {
                "name": name,
                "roadAddress": road,
                "address": "울산광역시 중구",
                "hasTrashBag": trash,
                "hasSpecialBag": special,
            }
        row = combined[kfix]
        row["hasTrashBag"] = bool(row["hasTrashBag"] or trash)
        row["hasSpecialBag"] = bool(row["hasSpecialBag"] or special)

    for name, ar in trash_in:
        feed(name, ar, trash=True, special=False)
    for name, ar in special_in:
        feed(name, ar, trash=False, special=True)

    cache = load_cache()
    ids_sc = load_existing_shortcodes((OUT_TRASH, OUT_SPECIAL))
    ref_date = DATA_REF_DEFAULT

    enriched: dict[str, dict[str, object]] = {}
    n_geo = 0
    for kfix, r in sorted(
        combined.items(), key=lambda x: (str(x[1]["name"]), str(x[1]["roadAddress"]))
    ):
        name = str(r["name"])
        road_orig = str(r["roadAddress"])
        ck = hashlib.sha1(
            ("ulsan:junggu:" + road_orig + "\x1f" + name).encode()
        ).hexdigest()[:24]
        latlng = cache.get(ck)
        lat = lng = None
        road_display = road_orig

        if latlng and len(latlng) == 2:
            lat, lng = latlng
        else:
            used_kakao = False
            for base in addr_variants(road_orig):
                for q in (
                    base,
                    f"{name} {base}",
                    f"{name}",
                    f"울산광역시 중구 {name}",
                    f"울산 중구 {base.replace('울산광역시 중구 ', '')}",
                    re.sub(r"^울산광역시 중구\s+", "", base),
                ):
                    glat, glng = kakao_geocode(q, key, road_orig)
                    if glat is not None and glng is not None:
                        lat, lng = glat, glng
                        used_kakao = True
                        break
                if lat is not None:
                    break
            if lat is None:
                manual = lookup_manual_junggu(name, road_orig)
                if manual:
                    road_display, lat, lng = manual
                    sys.stderr.write(f"[manual coord] {name}\t{road_display}\n")
                else:
                    sys.stderr.write(f"[skip geocode failed] {name}\t{road_orig}\n")
                    continue
            cache[ck] = [float(lat), float(lng)]
            if used_kakao:
                n_geo += 1
                if n_geo % 50 == 0:
                    save_cache(cache)
                    sys.stderr.write(f"[geocode] +{n_geo} …\n")
                time.sleep(GEOCODE_DELAY)

        if not isinstance(lat, (float, int)) or not isinstance(lng, (float, int)):
            continue
        if not coord_acceptable_for_row(float(lat), float(lng), road_display):
            sys.stderr.write(f"[bbox skip] {name}\t{road_orig}\t{lat},{lng}\n")
            continue

        sid = stable_id_digest(name, road_orig)
        addr_dist = "울산광역시 중구"
        if road_display.startswith("울산광역시 울주군"):
            addr_dist = "울산광역시 울주군"
        else:
            for pfx in ("울산광역시 북구", "울산광역시 남구", "울산광역시 동구"):
                if road_display.startswith(pfx):
                    addr_dist = pfx
                    break

        enriched[kfix] = {
            **r,
            "id": sid,
            "roadAddress": road_display,
            "address": addr_dist,
            "lat": round(float(lat), 7),
            "lng": round(float(lng), 7),
            "businessStatus": "영업",
            "hasTrashBag": r["hasTrashBag"],
            "hasSpecialBag": r["hasSpecialBag"],
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": ref_date,
            "shortCode": ids_sc.get(sid),
        }
        sc = enriched[kfix].get("shortCode")
        if isinstance(sc, str) and SHORT_CODE_RE.match(sc.strip()):
            enriched[kfix]["shortCode"] = sc.strip()

    save_cache(cache)

    trash_out: list[dict] = []
    special_out: list[dict] = []
    for rec in enriched.values():
        base = dict(rec)
        common = {
            "id": base["id"],
            "name": base["name"],
            "lat": base["lat"],
            "lng": base["lng"],
            "roadAddress": base["roadAddress"],
            "address": base["address"],
            "businessStatus": base["businessStatus"],
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": base["dataReferenceDate"],
        }
        if base.get("shortCode"):
            common["shortCode"] = base["shortCode"]

        ht = bool(base["hasTrashBag"])
        hs = bool(base["hasSpecialBag"])
        if ht:
            trash_out.append(
                {**common, "hasTrashBag": True, "hasSpecialBag": hs}
            )
        if hs and not ht:
            special_out.append(
                {**common, "hasTrashBag": False, "hasSpecialBag": True}
            )

    trash_out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    special_out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_TRASH.parent.mkdir(parents=True, exist_ok=True)
    OUT_TRASH.write_text(
        json.dumps(trash_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    OUT_SPECIAL.write_text(
        json.dumps(special_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    skip = len(combined) - len(enriched)
    print(f"wrote {len(trash_out)} → {OUT_TRASH.name}")
    print(f"wrote {len(special_out)} → {OUT_SPECIAL.name}")
    if skip:
        print(f"경고: 지오코딩·bbox 실패로 제외={skip}건 (stderr 참고)")
    print("다음: cd frontend && npm run shortcodes:assign && npm run data:build-cache")


if __name__ == "__main__":
    main()
