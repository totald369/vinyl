#!/usr/bin/env python3
"""
강남구 '태워서는 안 되는 쓰레기용 봉투'(불연성 마대) 판매소 엑셀을 지오코딩해 stores.sample.json에 병합.
→ hasSpecialBag=true. 기존 매장이 있으면 hasSpecialBag만 갱신하고 hasTrashBag 등은 유지.
"""

import importlib.util
import json
import sys
import time
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
OUT_PATH = SCRIPT_DIR.parent / "public" / "data" / "stores.sample.json"

SPECIAL_XLSX = Path.home() / "Downloads" / "태워서는 안되는 쓰레기봉투 판매소 목록(2024. 6. 15.).xlsx"
SHEET_NAME = "태워서는 안 되는 쓰레기용 봉투 판매소"
DATA_REF = "2024-06-15"


def _load_merge_gangnam():
    spec = importlib.util.spec_from_file_location("merge_gangnam", SCRIPT_DIR / "merge-gangnam.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def apply_row(mg, existing, cache, stats, name, addr):
    name = str(name).strip() if name is not None else ""
    addr = str(addr).strip() if addr is not None else ""
    if not name or not addr:
        return

    match = mg.find_existing_match(existing, name, addr)
    if match:
        match["hasSpecialBag"] = True
        match["dataReferenceDate"] = DATA_REF
        stats["updated"] += 1
        return

    if mg.excel_row_in_json(existing, name, addr):
        n = name.strip()
        an = mg.norm_addr(addr)
        for s in existing:
            if s["name"].strip() != n or "강남구" not in mg.gangnam_loc(s):
                continue
            ln = mg.norm_addr(mg.gangnam_loc(s))
            if an and ln and (an in ln or ln in an):
                s["hasSpecialBag"] = True
                s["dataReferenceDate"] = DATA_REF
                stats["updated_json"] += 1
                return

    cache_key = "specbag:" + addr
    coords = mg.coords_from_cache(cache, cache_key)
    if coords is None:
        coords = mg.kakao_geocode(addr, place_name=name)
        mg.put_coords_cache(cache, cache_key, coords)
        stats["api_calls"] += 1
        time.sleep(0.05)

    if not coords:
        stats["failed"] += 1
        print(f"  Geocode fail: {name} | {addr}")
        return

    stats["geocoded"] += 1
    stats["max_id"] += 1
    existing.append(
        {
            "id": str(stats["max_id"]),
            "name": name,
            "lat": coords[0],
            "lng": coords[1],
            "roadAddress": addr,
            "address": addr,
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": False,
            "adminVerified": False,
            "dataReferenceDate": DATA_REF,
        }
    )


def main():
    if not SPECIAL_XLSX.exists():
        print(f"파일 없음: {SPECIAL_XLSX}")
        sys.exit(1)

    import openpyxl

    mg = _load_merge_gangnam()

    print("Reading 강남구 불연성 마대 Excel...")
    wb = openpyxl.load_workbook(str(SPECIAL_XLSX), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print("시트 목록:", wb.sheetnames)
        wb.close()
        sys.exit(1)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()
    print(f"  → {len(rows)} rows (header 다음부터)")

    with open(OUT_PATH, "r", encoding="utf-8") as f:
        existing = json.load(f)
    print(f"  기존 stores: {len(existing)}")

    max_id = max(int(s["id"]) for s in existing)
    cache = mg.load_cache()
    stats = {
        "updated": 0,
        "updated_json": 0,
        "geocoded": 0,
        "failed": 0,
        "api_calls": 0,
        "max_id": max_id,
    }

    for row in rows:
        name = row[1]
        addr = row[2]
        apply_row(mg, existing, cache, stats, name, addr)
        if stats["api_calls"] and stats["api_calls"] % 80 == 0:
            mg.save_cache(cache)
            print(f"  API {stats['api_calls']}, new {stats['geocoded']}, fail {stats['failed']}")

    mg.save_cache(cache)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    sp = sum(1 for s in existing if s.get("hasSpecialBag"))
    gn_sp = sum(
        1
        for s in existing
        if s.get("hasSpecialBag") and "강남구" in mg.gangnam_loc(s)
    )
    print("\n강남 불연성 마대 병합 완료:")
    print(f"  기존 매장 hasSpecialBag 갱신: {stats['updated']}")
    print(f"  JSON상 이름·주소 매칭 갱신: {stats['updated_json']}")
    print(f"  신규 지오코딩 추가: {stats['geocoded']}")
    print(f"  지오코딩 실패: {stats['failed']}")
    print(f"  전체 매장 수: {len(existing)}")
    print(f"  hasSpecialBag 전체: {sp}, 그중 강남구 주소: {gn_sp}")
    print(f"Written: {OUT_PATH}")


if __name__ == "__main__":
    main()
