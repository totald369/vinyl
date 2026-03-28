#!/usr/bin/env python3
"""
영등포구 일반종량제봉투 / 특수종량제봉투(불연성·PP 마대) 판매소 엑셀 → stores.sample.json 병합.
- 시트 1행 제목에 '일반종량제봉투' 또는 '특수종량제봉투' 가 포함된 파일만 처리 (송파 등 다른 구 엑셀과 구분)
- 동일 상호·주소는 hasTrashBag / hasSpecialBag 을 OR 로 합침
- 주소 기준 카카오 지오코딩

사용:
  .env.local 의 KAKAO_REST_API_KEY
  python3 scripts/import_yeongdeungpo_excel.py
  python3 scripts/import_yeongdeungpo_excel.py ~/Downloads/일반....xlsx ~/Downloads/특수....xlsx
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

DOWNLOADS = Path.home() / "Downloads"
FRONTEND = Path(__file__).resolve().parent.parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.sample.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-yeongdeungpo.json"


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=8) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def norm_store_name(name: str) -> str:
    return re.sub(r"\s*\(\d+\)\s*$", "", (name or "").strip()).strip()


def norm_key(name: str, addr: str) -> str:
    a = re.sub(r"\s+", " ", addr.strip().lower())
    return f"{norm_store_name(name).lower()}|{a}"


def parse_ref_date(title: str) -> str | None:
    if not title:
        return None
    m = re.search(r"(\d{4})\s*[\.\-]\s*(\d{1,2})", title)
    if m:
        y, mo = m.groups()
        return f"{y}-{int(mo):02d}-01"
    return None


def classify_workbook(path: Path) -> tuple[bool, bool] | None:
    """(hasTrashBag, hasSpecialBag). None 이면 이 스크립트 대상 아님."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        t = str(next(ws.iter_rows(values_only=True))[0] or "").replace(" ", "")
        wb.close()
    except Exception:
        return None
    if "특수종량제봉투" in t:
        return False, True
    if "일반종량제봉투" in t:
        return True, False
    return None


def geocode_query_variants(full_addr: str) -> list[str]:
    seen: list[str] = []
    for q in (full_addr, re.sub(r"\s+", " ", full_addr).strip()):
        if q and q not in seen:
            seen.append(q)
    base = full_addr
    spaced_alley = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced_alley != base and spaced_alley not in seen:
        seen.append(spaced_alley)
        base = spaced_alley
    # 도신로 60길 → 도신로60길 (로·번지 사이 공백 오기)
    ro_gil = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", base)
    if ro_gil != base and ro_gil not in seen:
        seen.append(ro_gil)
        if "," in ro_gil:
            h0 = ro_gil.split(",")[0].strip()
            if h0 not in seen:
                seen.append(h0)
    if "," in base:
        head = base.split(",")[0].strip()
        if head and head not in seen:
            seen.append(head)
    m = re.search(r"([가-힣]+(?:로|길))\s+(\d+)-(\d+)", base)
    if m:
        stripped = base.replace(m.group(0), f"{m.group(1)} {m.group(2)}", 1)
        if stripped not in seen:
            seen.append(stripped)
        if "," in stripped:
            h = stripped.split(",")[0].strip()
            if h not in seen:
                seen.append(h)
    m = re.search(r"(면로|대로|로|길)(\d{2,4})\b", base)
    if m:
        fixed = re.sub(
            r"(면로|대로|로|길)(\d{2,4})\b",
            lambda mm: f"{mm.group(1)} {mm.group(2)}",
            base,
            count=1,
        )
        if fixed not in seen:
            seen.append(fixed)
        if "," in fixed:
            h2 = fixed.split(",")[0].strip()
            if h2 not in seen:
                seen.append(h2)
    # … 103~4호 등 물결 호수 접미 제거
    if "~" in full_addr:
        no_tilde = re.sub(r"\s+\d+~\d+호.*$", "", full_addr).strip()
        if no_tilde and no_tilde not in seen:
            seen.append(no_tilde)
            ng = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", no_tilde)
            if ng != no_tilde and ng not in seen:
                seen.append(ng)
            if "," in no_tilde:
                h3 = no_tilde.split(",")[0].strip()
                if h3 not in seen:
                    seen.append(h3)
    return seen


def resolve_coords(
    road_address: str, place_name: str, cache: dict, key: str
) -> tuple[float, float] | None:
    for qv in geocode_query_variants(road_address):
        c = kakao_geocode(qv, cache, key)
        if c:
            return c
    q2 = f"{road_address} {place_name}"
    c = kakao_geocode(q2, cache, key)
    if c:
        return c
    for qv in geocode_query_variants(road_address):
        c = kakao_keyword_geocode(qv, cache, key)
        if c:
            return c
    for q in (
        f"영등포구 {place_name}",
        place_name,
        f"서울 {place_name}",
    ):
        c = kakao_keyword_geocode(q, cache, key)
        if c:
            return c
    return None


def pad_row(r: tuple, n: int) -> list:
    r = list(r)
    while len(r) < n:
        r.append(None)
    return r


def parse_workbook(
    path: Path, has_trash: bool, has_spec: bool
) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [pad_row(tuple(row), 8) for row in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 8:
        return []

    title = str(rows[0][0] or "")
    ref_date = parse_ref_date(title)

    header_idx: int | None = None
    for i, r in enumerate(rows):
        if str(r[1] or "").strip() == "판매소명" and "주소" in str(r[2] or ""):
            header_idx = i
            break
    if header_idx is None:
        return []

    carry_dong = ""
    out: list[dict] = []
    for r in rows[header_idx + 1 :]:
        dong = str(r[0] or "").strip()
        if dong:
            carry_dong = dong
        name = str(r[1] or "").strip()
        addr = str(r[2] or "").strip()
        if not name or name == "판매소명":
            continue
        if not addr:
            continue
        addr = re.sub(r"\s+", " ", addr)
        if not addr.startswith("서울"):
            addr = f"서울특별시 영등포구 {addr}"
        out.append(
            {
                "name": name,
                "roadAddress": addr,
                "address": addr,
                "businessStatus": "영업",
                "hasTrashBag": has_trash,
                "hasSpecialBag": has_spec,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref_date,
                "_source_file": path.name,
                "_haengjeong": carry_dong,
            }
        )
    return out


def discover_jobs(argv: list[str]) -> list[tuple[Path, bool, bool]]:
    if argv:
        jobs = []
        for a in argv:
            p = Path(a).expanduser().resolve()
            c = classify_workbook(p)
            if c is None:
                print(f"건너뜀 (일반/특수종량제봉투 제목 아님): {p.name}")
                continue
            ht, hs = c
            jobs.append((p, ht, hs))
        return jobs
    jobs = []
    for p in sorted(DOWNLOADS.glob("*.xlsx")):
        c = classify_workbook(p)
        if c is None:
            continue
        ht, hs = c
        jobs.append((p.resolve(), ht, hs))
    return jobs


def main():
    if not KAKAO_REST_KEY:
        print("KAKAO_REST_KEY 또는 KAKAO_REST_API_KEY 가 필요합니다.")
        raise SystemExit(1)

    jobs = discover_jobs(sys.argv[1:])
    if not jobs:
        print("대상 엑셀이 없습니다. 인자로 경로를 주거나 Downloads 에 파일을 두세요.")
        raise SystemExit(1)

    raw: list[dict] = []
    for path, ht, hs in jobs:
        rows = parse_workbook(path, ht, hs)
        kind = "일반(종량제)" if ht else "특수(불연성)"
        print(f"  {path.name}: {kind} {len(rows)}행")
        raw.extend(rows)

    by_key: dict[str, dict] = {}
    for s in raw:
        k0 = norm_key(s["name"], s["roadAddress"])
        if k0 not in by_key:
            by_key[k0] = s
        else:
            prev = by_key[k0]
            prev["hasTrashBag"] = prev["hasTrashBag"] or s["hasTrashBag"]
            prev["hasSpecialBag"] = prev["hasSpecialBag"] or s["hasSpecialBag"]
            rd = prev.get("dataReferenceDate")
            sd = s.get("dataReferenceDate")
            if sd and (not rd or sd > rd):
                prev["dataReferenceDate"] = sd
            if "(" in prev.get("name", "") and "(" not in s.get("name", ""):
                prev["name"] = s["name"]

    unique = list(by_key.values())
    print(f"중복 제거 후 {len(unique)}건")

    cache = load_cache()
    geocode_failed: list[str] = []
    print("지오코딩 중…")
    for i, s in enumerate(unique):
        coords = resolve_coords(s["roadAddress"], s["name"], cache, KAKAO_REST_KEY)
        if not coords:
            geocode_failed.append(s["name"])
            s["lat"] = None
            s["lng"] = None
        else:
            s["lat"], s["lng"] = coords
        if (i + 1) % 50 == 0:
            save_cache(cache)
            print(f"  …{i + 1}/{len(unique)}")
    save_cache(cache)

    if geocode_failed:
        print(f"좌표 실패 {len(geocode_failed)}건 (일부): {geocode_failed[:25]}")

    with open(OUT_JSON, "r", encoding="utf-8") as f:
        existing = json.load(f)

    max_id = 0
    for e in existing:
        try:
            max_id = max(max_id, int(str(e.get("id", "0"))))
        except ValueError:
            pass

    exist_keys = {
        norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", ""))
        for e in existing
    }

    added = 0
    updated = 0
    for s in unique:
        k0 = norm_key(s["name"], s["roadAddress"])
        if k0 in exist_keys:
            for e in existing:
                if norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", "")) != k0:
                    continue
                ch = False
                nt = bool(e.get("hasTrashBag")) or s["hasTrashBag"]
                ns = bool(e.get("hasSpecialBag")) or s["hasSpecialBag"]
                if nt != bool(e.get("hasTrashBag")):
                    ch = True
                if ns != bool(e.get("hasSpecialBag")):
                    ch = True
                e["hasTrashBag"] = nt
                e["hasSpecialBag"] = ns
                if s.get("dataReferenceDate"):
                    od = e.get("dataReferenceDate") or ""
                    if not od or (s["dataReferenceDate"] > od):
                        e["dataReferenceDate"] = s["dataReferenceDate"]
                        ch = True
                if (e.get("lat") is None or e.get("lng") is None) and s.get("lat") is not None:
                    e["lat"] = s["lat"]
                    e["lng"] = s["lng"]
                    ch = True
                if ch:
                    updated += 1
                break
            continue
        if s.get("lat") is None:
            continue
        max_id += 1
        exist_keys.add(k0)
        rec = {
            "id": str(max_id),
            "name": s["name"],
            "lat": s["lat"],
            "lng": s["lng"],
            "roadAddress": s["roadAddress"],
            "address": s["address"],
            "businessStatus": s["businessStatus"],
            "hasTrashBag": s["hasTrashBag"],
            "hasSpecialBag": s["hasSpecialBag"],
            "hasLargeWasteSticker": False,
            "adminVerified": False,
        }
        if s.get("dataReferenceDate"):
            rec["dataReferenceDate"] = s["dataReferenceDate"]
        existing.append(rec)
        added += 1

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"갱신 {updated}건, 신규 {added}건 → 총 {len(existing)} 저장: {OUT_JSON}")


if __name__ == "__main__":
    main()
