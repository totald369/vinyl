#!/usr/bin/env python3
"""
서울시 강북구 종량제봉투·특수규격봉투(불연성) 판매소 엑셀 → stores.sample.json 병합.

기대 파일(또는 인자 순서):
  1) 특수규격봉투 판매소 현황.xlsx  → hasSpecialBag
  2) 2. 종량제봉투 판매소 현황.xlsx → hasTrashBag

같은 상호·주소는 한 행으로 합치고, 두 품목 모두 판매하면 플래그를 함께 켭니다.

사용:
  .env.local 의 KAKAO_REST_API_KEY
  python3 scripts/import_gangbuk_xlsx.py
  python3 scripts/import_gangbuk_xlsx.py /path/특수....xlsx /path/종량제....xlsx
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

DOWNLOADS = Path.home() / "Downloads"
FRONTEND = Path(__file__).resolve().parent.parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.sample.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-gangbuk.json"


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=8) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def norm_store_name(name: str) -> str:
    return re.sub(r"\s*\(\d+\)\s*$", "", (name or "").strip()).strip()


def norm_key(name: str, addr: str) -> str:
    a = re.sub(r"\s+", " ", addr.strip().lower())
    return f"{norm_store_name(name).lower()}|{a}"


def geocode_query_variants(full_addr: str) -> list[str]:
    seen: list[str] = []
    for q in (full_addr, re.sub(r"\s+", " ", full_addr).strip()):
        if q and q not in seen:
            seen.append(q)
    base = full_addr
    spaced_alley = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced_alley != base and spaced_alley not in seen:
        seen.append(spaced_alley)
        base = spaced_alley
    spaced2 = re.sub(r"([가-힣]+로)\s+(\d+)(길)", r"\1\2\3", base)
    if spaced2 != base and spaced2 not in seen:
        seen.append(spaced2)
    if "," in base:
        head = base.split(",")[0].strip()
        if head and head not in seen:
            seen.append(head)
    m = re.search(r"([가-힣]+(?:로|길))\s+(\d+)-(\d+)", base)
    if m:
        stripped = base.replace(m.group(0), f"{m.group(1)} {m.group(2)}", 1)
        if stripped not in seen:
            seen.append(stripped)
        if "," in stripped:
            h = stripped.split(",")[0].strip()
            if h not in seen:
                seen.append(h)
    m = re.search(r"(면로|대로|로|길)(\d{2,4})\b", base)
    if m:
        fixed = re.sub(
            r"(면로|대로|로|길)(\d{2,4})\b",
            lambda mm: f"{mm.group(1)} {mm.group(2)}",
            base,
            count=1,
        )
        if fixed not in seen:
            seen.append(fixed)
        if "," in fixed:
            h2 = fixed.split(",")[0].strip()
            if h2 not in seen:
                seen.append(h2)
    return seen


def _geocode_extra_queries(road: str) -> list[str]:
    ex: list[str] = []
    if re.search(r"[가-힣]+로\d{2,}", road):
        ex.append(re.sub(r"([가-힣]+로)(\d{2,})", r"\1 \2", road))
    if re.search(r"[가-힣]+길\d{2,}", road):
        ex.append(re.sub(r"([가-힣]+길)(\d{2,})", r"\1 \2", road))
    seen: list[str] = []
    for q in ex:
        q = re.sub(r"\s+", " ", q.strip())
        if q and q not in seen:
            seen.append(q)
    return seen


def resolve_coords_gangbuk(
    road_address: str,
    place_name: str,
    cache: dict,
    key: str,
) -> tuple[float, float] | None:
    queries = [road_address] + _geocode_extra_queries(road_address)
    for addr in queries:
        for qv in geocode_query_variants(addr):
            c = kakao_geocode(qv, cache, key)
            if c:
                return c
        c = kakao_geocode(f"{addr} {place_name}", cache, key)
        if c:
            return c
        for qv in geocode_query_variants(addr):
            c = kakao_keyword_geocode(qv, cache, key)
            if c:
                return c
    for q in (
        f"강북구 {place_name}",
        f"서울특별시 강북구 {place_name}",
        place_name,
        f"서울 {place_name}",
    ):
        c = kakao_keyword_geocode(q, cache, key)
        if c:
            return c
    return None


def normalize_gangbuk_addr(addr: str) -> str:
    a = re.sub(r"\s+", " ", (addr or "").strip())
    a = re.sub(r"\s*\.\s*", " ", a)
    a = a.replace("서울시 ", "서울특별시 ")
    if not a:
        return a
    if a.startswith("서울특별시"):
        return a
    if a.startswith("서울 "):
        return "서울특별시" + a[2:]
    if a.startswith("강북구"):
        return "서울특별시 " + a
    return "서울특별시 강북구 " + a


def find_header_row(ws) -> int:
    for r in range(1, 25):
        c4 = str(ws.cell(r, 4).value or "")
        c5 = str(ws.cell(r, 5).value or "")
        c5n = re.sub(r"\s+", "", c5)
        if "상호" in c4 and "주소" in c5n:
            return r
    raise ValueError("헤더 행(상호/주소)을 찾지 못했습니다.")


def ref_date_above_header(ws, header_row: int) -> str | None:
    blob: list[str] = []
    for r in range(1, header_row):
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v is not None:
                blob.append(str(v))
    m = re.search(r"(\d{4})\s*\.\s*(\d{1,2})", " ".join(blob))
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-01"
    return None


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v) and abs(v) < 1e9:
        return str(int(v))
    return str(v).strip()


def parse_workbook(path: Path, *, has_trash: bool, has_special: bool) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = find_header_row(ws)
    ref = ref_date_above_header(ws, hr)
    out: list[dict] = []
    for r in range(hr + 1, ws.max_row + 1):
        name = _cell_str(ws.cell(r, 4).value)
        addr = _cell_str(ws.cell(r, 5).value)
        if not name or not addr:
            continue
        if name == "상호" or "상호" == name.strip():
            continue
        road = normalize_gangbuk_addr(addr)
        out.append(
            {
                "name": norm_store_name(name),
                "roadAddress": road,
                "address": road,
                "businessStatus": "영업",
                "hasTrashBag": has_trash,
                "hasSpecialBag": has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref,
                "_source": path.name,
            }
        )
    wb.close()
    return out


def merge_by_key(rows: list[dict]) -> list[dict]:
    by_key: dict[str, dict] = {}
    for s in rows:
        k0 = norm_key(s["name"], s["roadAddress"])
        if k0 not in by_key:
            by_key[k0] = {**s}
        else:
            prev = by_key[k0]
            prev["hasTrashBag"] = bool(prev.get("hasTrashBag")) or bool(s.get("hasTrashBag"))
            prev["hasSpecialBag"] = bool(prev.get("hasSpecialBag")) or bool(
                s.get("hasSpecialBag")
            )
            rd = prev.get("dataReferenceDate")
            sd = s.get("dataReferenceDate")
            if sd and (not rd or sd > rd):
                prev["dataReferenceDate"] = sd
    return list(by_key.values())


def default_paths() -> tuple[Path | None, Path | None]:
    sp = DOWNLOADS / "특수규격봉투 판매소 현황.xlsx"
    tr = DOWNLOADS / "2. 종량제봉투 판매소 현황.xlsx"
    return (
        sp if sp.exists() else None,
        tr if tr.exists() else None,
    )


def main():
    dry = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if a not in ("--dry-run",)]

    if len(args) >= 2:
        p_special = Path(args[0]).expanduser().resolve()
        p_trash = Path(args[1]).expanduser().resolve()
    else:
        p_special, p_trash = default_paths()
        if not p_special or not p_trash:
            print(
                "인자로 [특수규격.xlsx] [종량제.xlsx] 경로를 주거나, "
                "Downloads 에 기본 파일명을 두세요."
            )
            raise SystemExit(1)

    raw: list[dict] = []
    if p_special.exists():
        part = parse_workbook(p_special, has_trash=False, has_special=True)
        print(f"  {p_special.name}: 특수규격 {len(part)}건")
        raw.extend(part)
    else:
        print(f"건너뜀 (없음): {p_special}")

    if p_trash.exists():
        part = parse_workbook(p_trash, has_trash=True, has_special=False)
        print(f"  {p_trash.name}: 종량제 {len(part)}건")
        raw.extend(part)
    else:
        print(f"건너뜀 (없음): {p_trash}")

    unique = merge_by_key(raw)
    print(f"병합 후 고유 {len(unique)}건")

    if dry:
        for s in unique[:10]:
            print(
                f"  - {s['name']} | T={s['hasTrashBag']} S={s['hasSpecialBag']} | {s['roadAddress'][:50]}"
            )
        if len(unique) > 10:
            print(f"  … 외 {len(unique) - 10}건")
        print("(dry-run)")
        return

    if not KAKAO_REST_KEY:
        print("KAKAO_REST_KEY 또는 KAKAO_REST_API_KEY 가 필요합니다.")
        raise SystemExit(1)

    cache = load_cache()
    geocode_failed: list[str] = []
    print("지오코딩 중…")
    for i, s in enumerate(unique):
        c = resolve_coords_gangbuk(s["roadAddress"], s["name"], cache, KAKAO_REST_KEY)
        if not c:
            geocode_failed.append(s["name"])
            s["lat"] = None
            s["lng"] = None
        else:
            s["lat"], s["lng"] = c
        if (i + 1) % 50 == 0:
            save_cache(cache)
            print(f"  …{i + 1}/{len(unique)}")
    save_cache(cache)

    if geocode_failed:
        print(f"좌표 실패 {len(geocode_failed)}건 (일부): {geocode_failed[:12]}")

    with open(OUT_JSON, "r", encoding="utf-8") as f:
        existing = json.load(f)

    max_id = 0
    for e in existing:
        try:
            max_id = max(max_id, int(str(e.get("id", "0"))))
        except ValueError:
            pass

    exist_keys = {
        norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", ""))
        for e in existing
    }

    added = 0
    updated = 0
    for s in unique:
        s.pop("_source", None)
        k0 = norm_key(s["name"], s["roadAddress"])
        if k0 in exist_keys:
            for e in existing:
                if norm_key(e.get("name", ""), e.get("roadAddress") or e.get("address", "")) != k0:
                    continue
                ch = False
                nt = bool(e.get("hasTrashBag")) or s["hasTrashBag"]
                ns = bool(e.get("hasSpecialBag")) or s["hasSpecialBag"]
                if nt != bool(e.get("hasTrashBag")):
                    ch = True
                if ns != bool(e.get("hasSpecialBag")):
                    ch = True
                e["hasTrashBag"] = nt
                e["hasSpecialBag"] = ns
                if s.get("dataReferenceDate"):
                    od = e.get("dataReferenceDate") or ""
                    if not od or (s["dataReferenceDate"] > od):
                        e["dataReferenceDate"] = s["dataReferenceDate"]
                        ch = True
                if (e.get("lat") is None or e.get("lng") is None) and s.get("lat") is not None:
                    e["lat"] = s["lat"]
                    e["lng"] = s["lng"]
                    ch = True
                if ch:
                    updated += 1
                break
            continue
        if s.get("lat") is None:
            continue
        max_id += 1
        exist_keys.add(k0)
        rec = {
            "id": str(max_id),
            "name": s["name"],
            "lat": s["lat"],
            "lng": s["lng"],
            "roadAddress": s["roadAddress"],
            "address": s["address"],
            "businessStatus": s["businessStatus"],
            "hasTrashBag": bool(s.get("hasTrashBag")),
            "hasSpecialBag": bool(s.get("hasSpecialBag")),
            "hasLargeWasteSticker": False,
            "adminVerified": False,
        }
        if s.get("dataReferenceDate"):
            rec["dataReferenceDate"] = s["dataReferenceDate"]
        existing.append(rec)
        added += 1

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"갱신 {updated}건, 신규 {added}건 → 총 {len(existing)} 저장: {OUT_JSON}")


if __name__ == "__main__":
    main()
