#!/usr/bin/env python3
"""
부산 해운대구 종량제봉투 판매소 xlsx (컬럼: 연번, 판매소명, 소재지 동명, 주소, 전화번호)

  python3 scripts/import_busan_haeundae_from_xlsx.py \\
    --xlsx ~/Downloads/'04_부산광역시_해운대구_종량제봉투판매소_20200210.xlsx'
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.busan-haeundae-trash.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-busan-haeundae-trash.json"

USER_AGENT = "Mozilla/5.0 (compatible; VinylMapImport/1.0; +https://github.com/totald369/vinyl)"
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.06


def _load_dotenv_local():
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()
KAKAO_REST_KEY = (
    os.environ.get("KAKAO_REST_KEY", "").strip()
    or os.environ.get("KAKAO_REST_API_KEY", "").strip()
)


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(c):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(c, f, ensure_ascii=False)


def kakao_keyword_geocode(query: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(("kw:" + query).encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)
    q2 = urllib.parse.urlencode({"query": query, "size": "1"})
    r = urllib.request.Request(
        f"{KEYWORD_URL}?{q2}", headers={"Authorization": f"KakaoAK {key}"}
    )
    try:
        with urllib.request.urlopen(r, timeout=12) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    cache[h] = [lat, lng]
                    time.sleep(GEOCODE_DELAY)
                    return lat, lng
    except (urllib.error.URLError, urllib.error.HTTPError, Exception):
        pass
    time.sleep(GEOCODE_DELAY)
    return None


def kakao_geocode(address: str, cache: dict, key: str) -> tuple[float, float] | None:
    h = hashlib.sha256(address.encode("utf-8")).hexdigest()[:16]
    if h in cache:
        lat, lng = cache[h]
        return float(lat), float(lng)

    def req(url: str) -> tuple[float, float] | None:
        r = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {key}"})
        try:
            with urllib.request.urlopen(r, timeout=12) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                docs = data.get("documents", [])
                if docs:
                    lat = to_float(docs[0].get("y"))
                    lng = to_float(docs[0].get("x"))
                    if lat and lng:
                        return lat, lng
        except (urllib.error.URLError, urllib.error.HTTPError, Exception):
            pass
        return None

    q = urllib.parse.urlencode({"query": address})
    coords = req(f"{GEOCODE_URL}?{q}")
    if not coords:
        q2 = urllib.parse.urlencode({"query": address, "size": "1"})
        coords = req(f"{KEYWORD_URL}?{q2}")
    if coords:
        cache[h] = list(coords)
        time.sleep(GEOCODE_DELAY)
        return coords
    time.sleep(GEOCODE_DELAY)
    return None


def collapse(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def full_busan_haeundae(addr_raw: str, _dong_hint: str) -> str:
    s = collapse(str(addr_raw or "")).replace("\u200b", "")
    if not s:
        return s
    if s.startswith("부산광역시"):
        return s
    if s.startswith("부산시"):
        return collapse("부산광역시 " + s[3:].lstrip())
    if s.startswith("해운대구"):
        return collapse(f"부산광역시 {s}")
    if s.startswith("부산 ") and not s.startswith("부산광역시"):
        return collapse(f"부산광역시 {s[len('부산'):].lstrip()}")
    if not s.startswith("부산"):
        return collapse(f"부산광역시 해운대구 {s}")
    return s


def massage_for_geocode(road: str) -> str:
    """지도 검색 빈약 지번·띄어쓰기 보정."""
    s = collapse(road)
    s = re.sub(r"(로)(\d+)(번길)", r"\1 \2\3", s)
    s = re.sub(r"(로)(\d번길)", r"\1 \2", s)
    # 상호·단지 문자열 과다 시 일부 줄임 (청사포로 구간)
    if "청사포로" in s and ("경남" in s or "선경" in s):
        s = re.sub(
            r",?\s*경남[^\s,]*\s*APT[^\s가-힣]*|,?\s*상가동[^\s,]*",
            "",
            s,
            flags=re.I,
        )
        s = collapse(s)
    s = re.sub(r"(\S동)\s*산(\d)", r"\1 산 \2", s)
    s = re.sub(r"(\S동)\s+산\s+(\d+(?:-\d+)?)", r"\1 산\2", s)
    return collapse(s)


def geocode_address_hint(road: str, name: str) -> str:
    """표시 주소는 유지하고, 좌표만 도로명으로 유도 (지번 미제공 케이스)."""
    if "산15-3" in road and "영외" in name:
        return "부산광역시 해운대구 우동3로 94"
    return road


def fmt_phone(raw: object) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    s = collapse(str(raw))
    if re.match(r"^051-\d{3}-\d{4}$", s) or re.match(r"^051-\d{4}-\d{4}$", s):
        return s
    if re.match(r"^052-\d{3}-\d{4}$", s):
        return s
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if digits.startswith("051") and len(digits) == 11:
        return f"051-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits.startswith("052"):
        return f"052-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 10 and digits.startswith(("051", "052")):
        prefix = digits[:3]
        return f"{prefix}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 7:
        return f"051-{digits[:3]}-{digits[3:]}"
    return s


def geocode_query_variants(full_addr: str) -> list[str]:
    def _tail_paren_strip(x: str) -> str:
        return collapse(re.sub(r"\s*[\(（][^)）]*[\)）]\s*$", "", x))

    def add(q: str, out: list[str]) -> None:
        cq = collapse(q)
        if cq and cq not in out:
            out.append(cq)

    seen: list[str] = []
    collapsed = collapse(full_addr)
    comma_head = collapsed.split(",", 1)[0].strip()
    stripped_tail = _tail_paren_strip(collapsed)
    for q in (collapsed, comma_head, stripped_tail, _tail_paren_strip(comma_head)):
        add(q, seen)

    base = comma_head or stripped_tail or collapsed
    spaced_alley = re.sub(r"([가-힣]+로)(\d+)(길)", r"\1 \2\3", base)
    if spaced_alley != base:
        add(spaced_alley, seen)
        base = spaced_alley
    ro_gil = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", base)
    if ro_gil != base:
        add(ro_gil, seen)
        base = ro_gil
    ro_only = re.sub(
        r"^(부산광역시\s+해운대구\s+[가-힣\d\s]+?(?:로|대로))\s+\d.+",
        r"\1",
        base,
    )
    if ro_only != base:
        add(ro_only, seen)
    return seen


def resolve_coords(
    road_address: str, place_name: str, cache: dict, key: str
) -> tuple[float, float] | None:
    for qv in geocode_query_variants(road_address):
        c = kakao_geocode(qv, cache, key)
        if c:
            return c
    c = kakao_geocode(f"{road_address} {place_name}", cache, key)
    if c:
        return c
    for qv in geocode_query_variants(road_address):
        c = kakao_keyword_geocode(qv, cache, key)
        if c:
            return c

    bits = {place_name}
    if "(" in place_name:
        bits.add(re.sub(r"\([^)]*\)", "", place_name).strip())
    for b in bits:
        for q in (
            f"부산 해운대구 {b}",
            f"해운대구 {b}",
            f"부산 {b}",
            b,
        ):
            if len(q.strip()) < 2:
                continue
            c = kakao_keyword_geocode(q, cache, key)
            if c:
                return c

    if " 산" in road_address:
        alt = collapse(re.sub(r"\s산(\d+(?:-\d+)?)", r" \1", road_address))
        if alt != road_address:
            c = kakao_geocode(alt, cache, key)
            if c:
                return c
            c = kakao_keyword_geocode(alt, cache, key)
            if c:
                return c

    return None


def ref_date_from_filename(path: Path) -> str:
    m = re.search(r"(\d{8})", path.name)
    if m:
        d = m.group(1)
        return f"{d[:4]}-{d[4:6]}-{d[6:]}"
    return "2020-02-10"


def parse_rows(path: Path) -> list[tuple[int, str, str, str, str]]:
    """연번, 판매소명, 소재지동명(raw), 주소, 전화"""
    wb = load_workbook(path, read_only=True, data_only=True)
    sh_name = wb.sheetnames[0]
    sh = wb[sh_name]
    rows: list[tuple[int, str, str, str, str]] = []
    for r in range(2, sh.max_row + 1):
        seq_raw = sh.cell(r, 1).value
        nm = collapse(str(sh.cell(r, 2).value or ""))
        dong = collapse(str(sh.cell(r, 3).value or ""))
        addr = collapse(str(sh.cell(r, 4).value or ""))
        tel_raw = sh.cell(r, 5).value
        if not nm or not addr:
            continue
        try:
            seq = int(float(str(seq_raw)))
        except (TypeError, ValueError):
            seq = r
        rows.append((seq, nm, dong, addr, str(tel_raw or "")))
    wb.close()
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path.home()
        / "Downloads"
        / "04_부산광역시_해운대구_종량제봉투판매소_20200210.xlsx",
    )
    args = ap.parse_args()
    xp = args.xlsx.expanduser()
    if not xp.exists():
        print(f"파일 없음: {xp}", file=sys.stderr)
        sys.exit(1)

    ref_date = ref_date_from_filename(xp)
    parsed = parse_rows(xp)
    print(f"파싱 {len(parsed)}건 · ref={ref_date}", file=sys.stderr)

    cache = load_cache()
    if not KAKAO_REST_KEY:
        print("KAKAO_REST_API_KEY 없음.", file=sys.stderr)

    stores: list[dict] = []
    failed: list[str] = []

    for i, (seq, name, dong, addr_raw, tel_raw) in enumerate(parsed):
        road = massage_for_geocode(full_busan_haeundae(addr_raw, dong))
        oid = f"busan-haeundae-trash-{seq}"
        geo_road = geocode_address_hint(road, name)

        lat = lng = None
        if KAKAO_REST_KEY:
            c = resolve_coords(geo_road, name, cache, KAKAO_REST_KEY)
            if c:
                lat, lng = c
        if lat is None:
            failed.append(f"{oid} {name} | {road}")

        row: dict = {
            "id": oid,
            "name": name,
            "roadAddress": road,
            "address": road,
            "businessStatus": "영업",
            "hasTrashBag": True,
            "hasSpecialBag": False,
            "hasLargeWasteSticker": False,
            "adminVerified": False,
            "dataReferenceDate": ref_date,
        }
        phone = fmt_phone(tel_raw)
        if phone and phone != "051-000-0000":
            row["phone"] = phone
        if lat is not None and lng is not None:
            row["lat"] = lat
            row["lng"] = lng

        stores.append(row)
        if (i + 1) % 80 == 0:
            save_cache(cache)
            print(f"  진행 {i+1}/{len(parsed)}", file=sys.stderr)

    save_cache(cache)

    ok = sum(1 for s in stores if "lat" in s)
    print(f"좌표 {ok}/{len(stores)} 실패 {len(failed)}", file=sys.stderr)
    for ln in failed[:35]:
        print(f"  {ln}", file=sys.stderr)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(stores, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"저장: {OUT_JSON}", file=sys.stderr)


if __name__ == "__main__":
    main()
