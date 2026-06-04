#!/usr/bin/env python3
"""
전라남도 장흥군 종량제봉투 판매업소 xlsx → stores.jeonnam-jangheung-trash.json

시트 Sheet1 (2행~): 연번 | 업소명 | 주소

  python3 scripts/import_jeonnam_jangheung_trash_from_xlsx.py \\
    --input ~/Downloads/장흥군\\ 종량제봉투\\ 판매업소\\ 현황.xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonnam-jangheung-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonnam-jangheung-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "장흥군 종량제봉투 판매업소 현황.xlsx"
SHEET_NAME = "Sheet1"
REF_DATE = date.today().isoformat()
CACHE_VERSION = "v1-jangheung"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def in_jangheung_bbox(lat: float, lng: float) -> bool:
    return 34.42 <= lat <= 34.86 and 126.78 <= lng <= 127.22


def jangheung_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if "남양주" in t or "부산시" in t or "영도" in t:
        return False
    if "장흥군" in t:
        return True
    return ("전라남도" in t or "전남" in t) and "장흥" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^전남\s+", "전라남도 ", a)
    if not a:
        return ""
    if a.startswith("전라남도"):
        return a
    if a.startswith("장흥군"):
        return f"전라남도 {a}"
    return f"전라남도 장흥군 {a}"


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if "," in a:
        a = a.split(",")[0].strip()
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"장흥군(?=[가-힣])", "장흥군 ", a)
    a = re.sub(r"([읍면동리])([가-힣]{2,}(?:로|길|대로))", r"\1 \2", a)
    a = re.sub(r"(대로)(\d)", r"\1 \2", a)
    a = re.sub(r"(\d)길(\s*\d)", r"\1길\2", a)
    return format_display_addr(a)


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_jangheung_bbox(lat, lng):
        return None
    if not jangheung_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_jangheung_bbox(lat, lng):
        return None
    if not jangheung_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    norm = normalize_addr(addr_raw)
    tail = norm
    for prefix in ("전라남도 장흥군 ", "전라남도 ", "장흥군 "):
        if tail.startswith(prefix):
            tail = collapse(tail[len(prefix) :])
            break
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"전남 장흥군 {tail}")
    push(f"전라남도 장흥군 {tail}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"전라남도 장흥군 {road_only}")
    push(f"{name} 장흥")
    push(f"{name} 장흥군")
    compact = re.sub(r"\s+", "", name)
    if compact != name:
        push(f"{compact} 장흥군")
    if "토요시장" in tail or "토요시장" in addr_raw:
        push("장흥 토요시장")
        push("토요시장 장흥")
    if "예양로" in tail:
        push("장흥읍 예양리")
    return out


def area_fallback(addr_raw: str, key: str, display: str) -> GeoHit | None:
    norm = normalize_addr(addr_raw)
    m = re.search(r"(장흥읍|장평면|유치면|회진면|대덕면|용산면|부산면|안양면|관산면|유등면)", norm)
    if not m:
        return None
    area = m.group(1)
    for q in (f"전라남도 장흥군 {area}", f"전남 장흥군 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display
                cj, cr = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or display
                if cr and not hit.road.startswith("전라"):
                    hit.road = display
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    display = normalize_addr(addr_raw)
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                if "토요시장" in display and "토요시장" not in _doc_blob(d):
                    continue
                return hit
    return area_fallback(addr_raw, key, display)


@dataclass
class JangheungRow:
    name: str
    addr_raw: str


def iter_rows(path: Path) -> list[JangheungRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    out: list[JangheungRow] = []
    for r in range(2, ws.max_row + 1):
        name = collapse(str(ws.cell(r, 2).value or ""))
        addr = collapse(str(ws.cell(r, 3).value or ""))
        if not name or not addr or name in ("업소명", "상호"):
            continue
        out.append(JangheungRow(name=name, addr_raw=addr))
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow_geo = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    seen: set[str] = set()
    geo_n = 0
    misses = 0

    for row in rows:
        display = normalize_addr(row.addr_raw)
        dk = f"{row.name}|{display}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, row.addr_raw)
        if ck in cache and not args.refresh:
            lat, lng, road, jibeon = cache[ck]
            geo_n += 1
        elif allow_geo:
            hit = resolve_geocode(row.addr_raw, row.name, key)  # type: ignore[arg-type]
            if not hit:
                print(f"[지오코딩 실패] {row.name}\t{display}", file=sys.stderr)
                misses += 1
                continue
            lat, lng, road, jibeon = hit.lat, hit.lng, hit.road, hit.jibeon
            cache[ck] = [lat, lng, road, jibeon]
            geo_n += 1
        else:
            misses += 1
            continue

        if not in_jangheung_bbox(lat, lng):
            print(f"[좌표 제외] {row.name}\t{display}", file=sys.stderr)
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"jeonnam-jangheung-trash-{rid}",
                "name": row.name,
                "lat": round(lat, 7),
                "lng": round(lng, 7),
                "roadAddress": road,
                "address": jibeon,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": REF_DATE,
            }
        )

    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    if allow_geo:
        save_cache(cache)

    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={REF_DATE}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["장흥군"])


if __name__ == "__main__":
    main()
