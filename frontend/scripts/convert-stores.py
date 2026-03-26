#!/usr/bin/env python3
"""
Convert the national CSV + Anyang Excel into a single stores.json
matching the StoreData type exactly.
"""

import csv
import json
import math
import openpyxl
import sys
from pathlib import Path

CSV_PATH = Path.home() / "Downloads" / "전국종량제봉투판매소표준데이터 (1).csv"
ANYANG_XLSX = Path.home() / "Downloads" / "안양 판매소별 판매 현황(1월).xlsx"
OUT_PATH = Path(__file__).resolve().parent.parent / "public" / "data" / "stores.sample.json"

def read_csv_rows():
    """Read the national CSV (EUC-KR encoded)."""
    rows = []
    with open(CSV_PATH, "r", encoding="euc-kr", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

def read_anyang_excel():
    """Read the Anyang Excel and return a dict keyed by (name, address_fragment)."""
    wb = openpyxl.load_workbook(str(ANYANG_XLSX), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = all_rows[0]
    data = all_rows[1:]
    col_map = {str(h).strip().replace("\n", ""): i for i, h in enumerate(header) if h}

    result = {}
    for row in data:
        name_val = row[col_map.get("판매소명", 3)]
        addr_val = row[col_map.get("도로명주소", 4)]
        if not name_val:
            continue
        name = str(name_val).strip()
        addr = str(addr_val).strip() if addr_val else ""

        trash_bag_total_idx = col_map.get("종량제봉투", 30)
        sticker_idx = col_map.get("스티커", 31)

        non_burn_20_idx = col_map.get("불연성20리터", None)
        non_burn_50_idx = col_map.get("불연성50리터", None)
        if non_burn_20_idx is None:
            for k, v in col_map.items():
                if "불연성" in k and "20" in k:
                    non_burn_20_idx = v
                if "불연성" in k and "50" in k:
                    non_burn_50_idx = v

        trash_total = to_int(row[trash_bag_total_idx]) if trash_bag_total_idx is not None else 0
        sticker_total = to_int(row[sticker_idx]) if sticker_idx is not None else 0
        non_burn_total = 0
        if non_burn_20_idx is not None:
            non_burn_total += to_int(row[non_burn_20_idx])
        if non_burn_50_idx is not None:
            non_burn_total += to_int(row[non_burn_50_idx])

        result[name] = {
            "hasTrashBag": trash_total > 0,
            "hasSpecialBag": non_burn_total > 0,
            "hasLargeWasteSticker": sticker_total > 0,
        }
    return result

def to_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0

def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None

def is_positive_indicator(val):
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("Y", "YES", "판매", "가능", "1", "TRUE")

def main():
    print(f"Reading CSV: {CSV_PATH}")
    csv_rows = read_csv_rows()
    print(f"  → {len(csv_rows)} rows")

    print(f"Reading Anyang Excel: {ANYANG_XLSX}")
    anyang_map = read_anyang_excel()
    print(f"  → {len(anyang_map)} stores with product data")

    stores = []
    skipped = 0

    for idx, row in enumerate(csv_rows):
        name = (row.get("판매소명") or "").strip()
        if not name:
            skipped += 1
            continue

        lat = to_float(row.get("위도"))
        lng = to_float(row.get("경도"))
        if lat is None or lng is None:
            skipped += 1
            continue

        road_addr = (row.get("소재지도로명주소") or "").strip() or None
        jibun_addr = (row.get("소재지지번주소") or "").strip() or None
        biz_status = (row.get("영업상태명") or "").strip() or None
        ref_date = (row.get("데이터기준일자") or "").strip() or None
        sticker_yn = (row.get("대형폐기물스티커판매여부") or "").strip()

        has_trash_bag = True
        has_special_bag = False
        has_large_waste_sticker = is_positive_indicator(sticker_yn)

        if name in anyang_map:
            a = anyang_map[name]
            has_trash_bag = a["hasTrashBag"]
            has_special_bag = a["hasSpecialBag"]
            has_large_waste_sticker = a["hasLargeWasteSticker"]

        store = {
            "id": str(idx + 1),
            "name": name,
            "lat": lat,
            "lng": lng,
        }

        if road_addr:
            store["roadAddress"] = road_addr
        if jibun_addr:
            store["address"] = jibun_addr
        elif road_addr:
            store["address"] = road_addr
        if biz_status:
            store["businessStatus"] = biz_status

        store["hasTrashBag"] = has_trash_bag
        store["hasSpecialBag"] = has_special_bag
        store["hasLargeWasteSticker"] = has_large_waste_sticker
        store["adminVerified"] = False

        if ref_date:
            store["dataReferenceDate"] = ref_date

        stores.append(store)

    print(f"  → {len(stores)} valid stores, {skipped} skipped")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(stores, f, ensure_ascii=False, indent=2)

    print(f"Output written to: {OUT_PATH}")
    print(f"File size: {OUT_PATH.stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    main()
