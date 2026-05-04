#!/usr/bin/env python3
"""
도봉구 특수규격 종량제봉투(20L) 판매소 엑셀 → public/data/stores.dobong-noncombust.json

- 시트「현황」: 연번 / 상호명 / 주소 / 전화번호 (3행 헤더, 4행부터 데이터)
- 주소를 도로명 기준으로 정규화 후 Nominatim 지오코딩 (구로·관악 불연성 임포트와 동일 패턴)
- 매장 분류: 불연성 마대 취급처와 동일하게 storeCategory nonBurnable

사용법 (frontend 디렉터리에서):
  python3 scripts/import_dobong_noncombust_from_xlsx.py "/path/to/도봉구....xlsx"
  npm run shortcodes:assign
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

SHORT_CODE_RE = re.compile(r"^[a-zA-Z0-9]{6}$")

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as e:
    raise SystemExit("openpyxl이 필요합니다. pip install openpyxl") from e


def full_road_address(raw: str) -> str:
    a = str(raw).replace("\xa0", " ").strip()
    a = " ".join(a.split())
    # 엑셀 오기(도붕구 등)
    a = a.replace("도붕구", "도봉구").replace("서울 도붕구", "서울 도봉구")
    # "서울특별시 도봉구 서울 도봉구 …" / "… 도봉구 …" 중복 접두 제거
    if a.startswith("서울특별시 도봉구"):
        rest = a[len("서울특별시 도봉구") :].lstrip()
        rest = re.sub(r"^서울(\s*특별시)?\s+도봉구\s+", "", rest, count=1).lstrip()
        rest = re.sub(r"^도봉구\s+", "", rest, count=1).lstrip()
        a = "서울특별시 도봉구 " + rest if rest else "서울특별시 도봉구"
    if a.startswith("서울특별시 도봉구"):
        return a
    if a.startswith("서울 도봉구"):
        return "서울특별시 도봉구 " + a[7:].strip()
    if a.startswith("도봉구"):
        return "서울특별시 " + a
    return "서울특별시 도봉구 " + a


def norm_phone(raw: str | None) -> str | None:
    if raw is None:
        return None
    t = str(raw).replace("\xa0", " ").strip()
    if not t:
        return None
    digits = re.sub(r"\D", "", t)
    if len(digits) == 8 and not t.startswith("0"):
        return "02-" + t
    return t


def read_excel_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "현황" not in wb.sheetnames:
        raise SystemExit(f"시트「현황」이 없습니다. 시트: {wb.sheetnames}")
    ws = wb["현황"]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for r in rows[3:]:
        if not r:
            continue
        name = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        addr = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        phone = norm_phone(str(r[3]).strip() if len(r) > 3 and r[3] else None)
        if not name or not addr:
            continue
        out.append(
            {
                "name": name,
                "roadAddress": full_road_address(addr),
                "phone": phone,
            }
        )
    return out


UA = "SseubongmapDobongImport/1.0 (internal data tooling)"

CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-dobong-nominatim.json"


def load_geo_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        raw = json.loads(CACHE_PATH.read_text(encoding="utf8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, list[float]] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if isinstance(v, list) and len(v) == 2:
                out[str(k)] = [float(v[0]), float(v[1])]
    return out


def save_geo_cache(cache: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=0), encoding="utf8")


def load_kakao_rest_key() -> str | None:
    k = os.environ.get("KAKAO_REST_API_KEY")
    if k and str(k).strip():
        return str(k).strip()
    p = Path(__file__).resolve().parent.parent / ".env.local"
    if not p.exists():
        return None
    for line in p.read_text(encoding="utf8").splitlines():
        line = line.strip()
        if line.startswith("KAKAO_REST_API_KEY="):
            v = line.split("=", 1)[1].strip().strip('"').strip("'")
            return v or None
    return None


KAKAO_ADDR_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KAKAO_KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"


def kakao_geocode(q: str, api_key: str) -> tuple[float | None, float | None]:
    """Nominatim 실패 시 카카오 주소·키워드 검색."""
    time.sleep(0.18)
    for base, extra in (
        (KAKAO_ADDR_URL, {"query": q}),
        (KAKAO_KEYWORD_URL, {"query": q, "size": "1"}),
    ):
        qs = urllib.parse.urlencode(extra)
        req = urllib.request.Request(
            f"{base}?{qs}", headers={"Authorization": f"KakaoAK {api_key}"}
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except (OSError, ValueError):
            continue
        docs = data.get("documents") or []
        if not docs:
            continue
        try:
            lat = float(docs[0].get("y"))
            lng = float(docs[0].get("x"))
        except (TypeError, ValueError):
            continue
        if lat and lng:
            return lat, lng
    return None, None


def nominatim(q: str) -> tuple[float | None, float | None]:
    params = urllib.parse.urlencode({"q": q, "format": "json", "limit": "1"})
    url = "https://nominatim.openstreetmap.org/search?" + params
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=45) as resp:
        data = json.loads(resp.read().decode())
    time.sleep(1.12)
    if data:
        return float(data[0]["lat"]), float(data[0]["lon"])
    return None, None


def road_variants(road: str) -> list[str]:
    """층·호·일부 등 부가 표기를 떼고 시도."""
    out: list[str] = []
    seen: set[str] = set()

    def add(x: str) -> None:
        t = " ".join(x.replace("\xa0", " ").split())
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    add(road)
    head = road.split(",")[0].strip()
    add(head)
    # "… 370, 2층" → 첫 콤마만 제거한 뒤도 시도
    if "," in road:
        add(re.sub(r",\s*\d+층.*$", "", road).strip())
    # 층·동·호·상가 접미어는 N길M 분리 전에 제거해야 `$` 패턴이 맞음
    head = re.sub(r"\s+\d+층\s*$", "", head).strip()
    add(head)
    add(re.sub(r"\s+\d+동\s*$", "", head).strip())
    add(re.sub(r"\s+\d+호\s*$", "", head).strip())
    add(re.sub(r"\s+[가-힣0-9]+상가.*$", "", head).strip())
    add(re.sub(r"\s+B\d+.*$", "", head, flags=re.I).strip())
    # "도봉로 139길" → "도봉로139길" (행안부 표기)
    add(re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", head).strip())
    # "도봉로728" → "도봉로 728"
    add(re.sub(r"(로|길)(\d{2,})$", r"\1 \2", head).strip())
    # "방학로112-9" → "방학로 112-9"
    add(re.sub(r"([가-힣]+로)(\d)", r"\1 \2", head).strip())
    # "50길3" / "150길8" → "50길 3" / "150길 8"
    add(re.sub(r"(\d+길)(\d+)$", r"\1 \2", head).strip())
    gl = re.sub(r"([가-힣]+로)\s+(\d+길)", r"\1\2", head)
    add(gl)
    add(re.sub(r"(\d+길)(\d+)$", r"\1 \2", gl).strip())
    return out


def geocode_row(p: dict, cache: dict[str, list[float]], kakao_key: str | None) -> tuple[float, float]:
    ck = p["roadAddress"] + "\t" + p["name"]
    if ck in cache:
        lat, lng = cache[ck]
        return lat, lng

    name = p["name"]
    # 한국 주소는 카카오가 빠르고 정확한 경우가 많음 — 키가 있으면 먼저 시도
    if kakao_key:
        for base in road_variants(p["roadAddress"]):
            for q in (base, f"{name} {base}", f"서울 도봉구 {name}"):
                lat, lng = kakao_geocode(q, kakao_key)
                if lat is not None and lng is not None:
                    cache[ck] = [lat, lng]
                    save_geo_cache(cache)
                    return lat, lng
    for base in road_variants(p["roadAddress"]):
        for q in (
            base,
            f"{name} {base}",
            f"{name} 서울 도봉구",
        ):
            lat, lng = nominatim(q)
            if lat is not None and lng is not None:
                cache[ck] = [lat, lng]
                save_geo_cache(cache)
                return lat, lng
    raise RuntimeError(f"지오코딩 실패: {json.dumps(p, ensure_ascii=False)}")


def load_existing_by_id(existing_path: Path) -> dict[str, dict]:
    if not existing_path.exists():
        return {}
    try:
        data = json.loads(existing_path.read_text(encoding="utf8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(data, list):
        return {}
    out: dict[str, dict] = {}
    for row in data:
        if not isinstance(row, dict) or not row.get("id"):
            continue
        rid = str(row["id"])
        meta: dict = {}
        sc = row.get("shortCode")
        if isinstance(sc, str):
            sc = sc.strip()
            if SHORT_CODE_RE.match(sc):
                meta["shortCode"] = sc
        dr = row.get("dataReferenceDate")
        if isinstance(dr, str) and dr.strip():
            meta["dataReferenceDate"] = dr.strip()
        if meta:
            out[rid] = meta
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help="도봉구 특수규격 종량제봉투(20L) 판매소 목록.xlsx")
    ap.add_argument(
        "-o",
        "--out",
        type=Path,
        default=Path("public/data/stores.dobong-noncombust.json"),
    )
    args = ap.parse_args()

    parsed = read_excel_rows(args.xlsx)
    if not parsed:
        raise SystemExit("엑셀에서 유효한 데이터 행이 없습니다.")

    out_abs = (
        args.out
        if args.out.is_absolute()
        else Path(__file__).resolve().parent.parent / args.out
    )
    existing_meta = load_existing_by_id(out_abs)
    geo_cache = load_geo_cache()
    kakao_key = load_kakao_rest_key()
    iso_today = date.today().isoformat()
    bundle: list[dict] = []

    for i, p in enumerate(parsed, start=1):
        sid = f"dobong-nc-{i:03d}"
        lat, lng = geocode_row(p, geo_cache, kakao_key)
        row: dict = {
            "id": sid,
            "name": p["name"],
            "lat": round(lat, 7),
            "lng": round(lng, 7),
            "roadAddress": p["roadAddress"],
            "address": "서울특별시 도봉구",
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": existing_meta.get(sid, {}).get(
                "dataReferenceDate", iso_today
            ),
            "storeCategory": "nonBurnable",
        }
        if p.get("phone"):
            row["phone"] = p["phone"]
        sc = existing_meta.get(sid, {}).get("shortCode")
        if sc:
            row["shortCode"] = sc
        bundle.append(row)

    out_abs.parent.mkdir(parents=True, exist_ok=True)
    out_abs.write_text(
        json.dumps(bundle, ensure_ascii=False, indent=2) + "\n", encoding="utf8"
    )
    print(f"wrote {len(bundle)} rows -> {out_abs}")


if __name__ == "__main__":
    main()
