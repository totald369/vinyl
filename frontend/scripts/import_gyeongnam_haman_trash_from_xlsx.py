#!/usr/bin/env python3
"""
경상남도 함안군 종량제봉투·불연성마대 판매소 xlsx → stores.gyeongnam-haman-trash.json

시트 `1. 종량제판매소 위치 정보` (헤더 4행, 데이터 5행~):
  명칭(A) | 도로명주소(B) | 상세주소(C) | 규격별 판매여부(D~L) | 불연성특수마대(M)

매핑:
  - D~L열 중 하나라도 'O'/'○' -> hasTrashBag
  - M열 'O'/'○' -> hasSpecialBag

사용:
  cd frontend
  python3 scripts/import_gyeongnam_haman_trash_from_xlsx.py \\
    --input ~/Downloads/\\(함안군\\)종량제봉투\\ 판매소\\ 위치정보\\ 등\\ \\(1\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.gyeongnam-haman-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-gyeongnam-haman-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "(함안군)종량제봉투 판매소 위치정보 등 (1).xlsx"
SHEET_NAME = "1. 종량제판매소 위치 정보"
DATA_START_ROW = 5
REF_DATE = "2026-06-09"
CACHE_VERSION = "v1-haman"

COL_NAME = 1
COL_ADDR = 2
COL_DETAIL = 3
COL_TRASH_START = 4
COL_TRASH_END = 12
COL_SPECIAL = 13

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")
HAMAN_AREAS = (
    "가야읍",
    "칠원읍",
    "함안면",
    "군북면",
    "법수면",
    "대산면",
    "칠서면",
    "칠북면",
    "여항면",
    "산인면",
)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    t = collapse(str(val or "")).upper()
    return t in ("O", "○", "Y", "YES", "예", "여")


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def in_haman_bbox(lat: float, lng: float) -> bool:
    return 35.18 <= lat <= 35.38 and 128.30 <= lng <= 128.55


def haman_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if any(x in t for x in ("서울", "대구", "부산", "인천", "대전", "울산")):
        return False
    if "함안군" in t:
        return True
    return ("경남" in t or "경상남도" in t) and "함안" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^경남\s+", "경상남도 ", a)
    if not a:
        return ""
    if a.startswith("경상남도"):
        return a
    if a.startswith("함안군"):
        return f"경상남도 {a}"
    return f"경상남도 함안군 {a}"


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def normalize_haman_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if not a:
        return ""
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"([가-힣]+(?:리|동))(\d+)", r"\1 \2", a)
    a = re.sub(r"((?:대로|로|길))(\d+(?:-\d+)?)", r"\1 \2", a)
    if not a.startswith("경") and not a.startswith("함안"):
        if re.match(r"^\S+?(?:읍|면)", a):
            a = f"함안군 {a}"
        else:
            a = f"함안군 가야읍 {a}"
    return format_display_addr(a)


def geocode_target(full: str) -> str:
    a = collapse(full)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+\d+\s*층.*$", "", a)
    a = re.sub(r"\s+\d+\s*호\s*$", "", a)
    return a


def road_format_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(addr)
    compact = addr.replace(" ", "")
    push(compact)
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    push(re.sub(r"((?:로|길|대로))(\d+)", r"\1 \2", compact))
    return out


def haman_tail(full: str) -> str:
    for prefix in (
        "경상남도 함안군 ",
        "경상남도 ",
        "경남 함안군 ",
        "함안군 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def lot_query_variants(tail: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = LOT_RE.match(tail)
    if m:
        prefix, main, sub = m.group("prefix"), m.group("main"), m.group("sub")
        subs: list[str | None] = [sub, "1", "2", "3", "4", "5"] if sub else [None]
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def geocode_query_variants(road_full: str, name: str) -> list[str]:
    target = geocode_target(road_full)
    tail = haman_tail(target)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", target))
    if has_road:
        for base in road_format_variants(target):
            t = haman_tail(base)
            push(f"경남 함안군 {t}")
            push(f"경상남도 함안군 {t}")
    elif is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push(f"경남 함안군 {tv}")
            push(f"경상남도 함안군 {tv}")
    else:
        for base in road_format_variants(target):
            t = haman_tail(base)
            push(f"경남 함안군 {t}")
            push(f"경상남도 함안군 {t}")

    m = re.match(r"^(.+?(?:동|읍|면|리))", tail.replace(" ", ""))
    if m:
        push(f"경남 함안군 {collapse(m.group(1))}")

    push(f"{name} 함안")
    push(f"{name} 함안군")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class HamanRow:
    name: str
    addr_raw: str
    has_trash: bool
    has_special: bool


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_haman_bbox(lat, lng):
        return None
    if not haman_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_haman_bbox(lat, lng):
        return None
    if not haman_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(addr_raw: str, key: str, display_road: str) -> GeoHit | None:
    tail = haman_tail(geocode_target(addr_raw))
    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    area = collapse(m.group(1)) if m else ""
    if not area:
        for candidate in HAMAN_AREAS:
            if candidate in tail:
                area = candidate
                break
    if not area:
        return None
    for q in (f"경남 함안군 {area}", f"경상남도 함안군 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, _ = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, display_road: str) -> GeoHit | None:
    has_road = bool(re.search(r"(?:로|길|대로)\s*\d", display_road))
    for q in geocode_query_variants(display_road, name):
        if has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
        if not has_road:
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    return hit
    return area_fallback(addr_raw, key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def row_flags(ws, r: int) -> tuple[bool, bool]:
    has_trash = any(flag_y(ws.cell(r, c).value) for c in range(COL_TRASH_START, COL_TRASH_END + 1))
    has_special = flag_y(ws.cell(r, COL_SPECIAL).value)
    return has_trash, has_special


def iter_rows(path: Path) -> list[HamanRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    merged: dict[str, HamanRow] = {}

    def key_for(name: str, addr_raw: str) -> str:
        norm = normalize_haman_addr(addr_raw)
        road = geocode_target(norm)
        return f"{name}|{road}"

    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, COL_NAME).value)
        if not name or name in ("종량제봉투 판매소 명칭",):
            continue
        addr = cell_str(ws.cell(r, COL_ADDR).value)
        if not addr:
            continue
        has_trash, has_special = row_flags(ws, r)
        if not has_trash and not has_special:
            continue
        k = key_for(name, addr)
        if k not in merged:
            merged[k] = HamanRow(name, addr, False, False)
        rec = merged[k]
        if has_trash:
            rec.has_trash = True
        if has_special:
            rec.has_special = True

    wb.close()
    return list(merged.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = REF_DATE
    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    seen: set[str] = set()

    for row in rows:
        full_norm = normalize_haman_addr(row.addr_raw)
        display_road = geocode_target(full_norm)
        if not display_road or "함안" not in display_road.replace(" ", ""):
            continue

        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(full_norm, row.name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = cr or display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 20 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"gyeongnam-haman-trash-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road or display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": row.has_trash,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": ref_date,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    trash_cnt = sum(1 for r in out if r["hasTrashBag"])
    special_cnt = sum(1 for r in out if r["hasSpecialBag"])
    both_cnt = sum(1 for r in out if r["hasTrashBag"] and r["hasSpecialBag"])
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo={geo_n}, miss={misses}, src={len(rows)}, "
        f"종량제 {trash_cnt}, 불연성마대 {special_cnt}, 겸업 {both_cnt})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["함안군"])


if __name__ == "__main__":
    main()
