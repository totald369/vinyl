#!/usr/bin/env python3
"""
전북특별자치도 정읍시 종량제봉투 판매소 xlsx → stores.jeonbuk-jeongeup-trash.json

시트 `1. 종량제판매소 위치 정보` (5행~):
  지자체명(A) | 판매소 명칭(B) | 도로명 주소(C) | 입고일(D) | 용량별 취급현황(E~) | 불연성마대 안내(Y)

  * 불연성마대 열은 전 행이 "읍면동 주민센터" 안내 텍스트라 판매소 플래그로 쓰지 않음.
  * 기존 CSV(2023) 캐시(geocode-cache-jeonbuk-jeongeup-trash.json)를 재사용한다.

  python3 scripts/import_jeonbuk_jeongeup_trash_from_xlsx.py \\
    --input ~/Downloads/정보공개\\ 청구\\ \\(붙임\\)\\ 지자체별\\ 종량제봉투\\ 판매소\\ 현황\\ 조사표\\(정읍시\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonbuk-jeongeup-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonbuk-jeongeup-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "정보공개 청구 (붙임) 지자체별 종량제봉투 판매소 현황 조사표(정읍시).xlsx"
REF_DATE = "2026-06-08"
CACHE_VERSION = "v1"
DATA_START_ROW = 5
COL_NAME = 2
COL_ADDR = 3
# 일반용·재사용·음식물용 용량 열(1-based 5~24) — Y 하나라도 있으면 종량제 취급
TRASH_BAG_COLS = tuple(range(5, 25))

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    return collapse(str(val or "")).upper() == "Y"


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def in_jeongeup_bbox(lat: float, lng: float) -> bool:
    # 감곡·산외·산내 면 포함
    return 35.50 <= lat <= 35.75 and 126.72 <= lng <= 127.08


def jeongeup_in_text(blob: str) -> bool:
    return "정읍" in (blob or "").replace(" ", "")


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = a.replace("전북특별치도", "전북특별자치도")
    a = re.sub(r"^전북\s+", "전북특별자치도 ", a)
    a = re.sub(r"^전라북도\s+", "전북특별자치도 ", a)
    if not a.startswith("전북") and "정읍" in a.replace(" ", ""):
        if a.startswith("정읍시"):
            return f"전북특별자치도 {a}"
        return f"전북특별자치도 정읍시 {a}"
    return a


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"(\d+)\s+(\d+)\s*호\s*$", r"\1-\2", a)
    a = re.sub(r"(\d+)\s+(\d+)\s*$", r"\1-\2", a)
    a = re.sub(r"([가-힣]+(?:리|동))(\d+)", r"\1 \2", a)
    return format_display_addr(a)


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def lot_query_variants(tail: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = re.search(r"([가-힣]+(?:동|읍|면|리))\s+(\d+)\s*번지\s*(\d+)\s*호", tail)
    if m:
        push(tail[: m.start()] + f"{m.group(1)} {m.group(2)}-{m.group(3)}" + tail[m.end() :])
    m2 = LOT_RE.match(tail)
    if m2:
        prefix, main, sub = m2.group("prefix"), m2.group("main"), m2.group("sub")
        subs: list[str | None] = [sub, "1", "2", "3", "4", "5"] if sub else [None]
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def jeongeup_tail(full: str) -> str:
    for prefix in (
        "전북특별자치도 정읍시 ",
        "전북특별자치도 ",
        "전북 정읍시 ",
        "정읍시 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def geocode_query_variants(addr_full: str, name: str) -> list[str]:
    norm = normalize_addr(addr_full)
    tail = jeongeup_tail(norm)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    if is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push(f"전북 정읍시 {tv}")
            push(f"전북특별자치도 정읍시 {tv}")
    else:
        push(f"전북 정읍시 {tail}")
        push(f"전북특별자치도 정읍시 {tail}")
        push(norm)
        if re.search(r"-\d+\s*$", tail):
            rt = re.sub(r"-\d+\s*$", "", tail).strip()
            push(f"전북 정읍시 {rt}")
        road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
        if road_only and road_only != tail:
            push(f"전북 정읍시 {road_only}")

    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    if m:
        push(f"전북 정읍시 {collapse(m.group(1))}")

    push(f"{name} 정읍")
    push(f"{name} 정읍시")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class JeongeupRow:
    name: str
    addr_raw: str
    has_trash: bool


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_jeongeup_bbox(lat, lng):
        return None
    if not jeongeup_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_jeongeup_bbox(lat, lng):
        return None
    if not jeongeup_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    return data.get("documents") or []


def area_fallback(addr_full: str, key: str, display: str) -> GeoHit | None:
    tail = jeongeup_tail(normalize_addr(addr_full))
    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    if not m:
        return None
    area = collapse(m.group(1))
    for q in (f"전북 정읍시 {area}", f"전북특별자치도 정읍시 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display if not is_likely_jibeon(jeongeup_tail(display)) else hit.road
                cj, cr = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or normalize_addr(addr_full)
                if not hit.road or hit.road == hit.jibeon:
                    hit.road = cr or display
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, display: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return area_fallback(addr_raw, key, display)


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def iter_rows(path: Path) -> list[JeongeupRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: list[JeongeupRow] = []
    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, COL_NAME).value)
        addr_raw = cell_str(ws.cell(r, COL_ADDR).value)
        if not name or not addr_raw or name in ("종량제봉투 판매소 명칭",):
            continue
        has_trash = any(flag_y(ws.cell(r, c).value) for c in TRASH_BAG_COLS)
        out.append(JeongeupRow(name=name, addr_raw=addr_raw, has_trash=has_trash or True))
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    seen: set[str] = set()

    for row in rows:
        display_road = normalize_addr(row.addr_raw)
        if not display_road or "정읍" not in display_road.replace(" ", ""):
            continue

        dk = f"{row.name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else ""
                if not jib or not road:
                    cj, cr = coord2address(lng, lat, key)  # type: ignore[arg-type]
                    jib = jib or cj or display_road
                    road = road or cr or display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            norm = normalize_addr(row.addr_raw)
            tail = jeongeup_tail(norm)
            if is_likely_jibeon(tail):
                hit.jibeon = norm
                hit.road = cr or hit.road or norm
            else:
                hit.road = display_road
                hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"jeonbuk-jeongeup-trash-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road or display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": REF_DATE,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {len(out)} → {OUT_JSON} (ref={REF_DATE}, geo={geo_n}, miss={misses}, src={len(rows)})")

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["정읍시"])


if __name__ == "__main__":
    main()
