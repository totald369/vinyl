#!/usr/bin/env python3
"""
울산광역시 북구 종량제봉투·특수규격(불연성) 판매소 xlsx → JSON 2종

  python3 scripts/import_ulsan_bukgu_from_xlsx.py
  python3 scripts/import_ulsan_bukgu_from_xlsx.py --xlsx ~/Downloads/jonglyangje_situation_down.xlsx

시트: 「종량제봉투 판매소」, 「특수종량제봉투 판매소」
  컬럼: 번호, 판매소명, 사업장 주소, 연락처

KAKAO_REST_API_KEY: frontend/.env.local 또는 환경변수
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
OUT_TRASH = FRONTEND / "public" / "data" / "stores.ulsan-bukgu-trash.json"
OUT_SPECIAL = FRONTEND / "public" / "data" / "stores.ulsan-bukgu-special.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-ulsan-bukgu.json"
DEFAULT_XLSX = Path.home() / "Downloads" / "jonglyangje_situation_down.xlsx"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
GEOCODE_DELAY = 0.07

SHEET_TRASH = "종량제봉투 판매소"
SHEET_SPECIAL = "특수종량제봉투 판매소"


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def collapse(s: str) -> str:
    return " ".join((s or "").replace("\xa0", " ").split())


WS_RE = re.compile(r"[ \t\r\n\v\f]+")


def normalize_addr(raw: str) -> str:
    a = WS_RE.sub(" ", collapse(raw)).strip()
    if not a:
        return a
    a = re.sub(r"^\(주\)", "(주)", a)
    if a.startswith("울산광역시북구"):
        a = "울산광역시 북구 " + a[len("울산광역시북구") :].lstrip()
    if a.startswith("울산광역시 북구"):
        # 엑셀 오기: 「북구」 뒤에 「중구」「동구」「남구」가 잘못 끼어든 경우 제거
        a = re.sub(
            r"(울산광역시 북구)\s+(?:중구|동구|남구)\s+", r"\1 ", a,
        )
        m = re.search(r"(울산광역시 북구)([가-힣ㄱ-ㅎㅏ-ㅣ0-9])", a)
        if m and m.group(2):
            idx = m.start(2)
            a = a[:idx] + " " + a[idx:]
        a = _tweak_bukgu_road_typos(a)
        return WS_RE.sub(" ", a).strip()
    if a.startswith("북구"):
        return f"울산광역시 {a}"
    if a.startswith("울산시 북구"):
        return "울산광역시 " + a[len("울산시") :].lstrip()
    if not a.startswith("울산"):
        return f"울산광역시 북구 {a}"
    if "북구" not in a[:24]:
        rest = a[len("울산광역시") :].strip()
        out = f"울산광역시 북구 {rest}"
        out = _tweak_bukgu_road_typos(out)
        return WS_RE.sub(" ", out).strip()
    out = WS_RE.sub(" ", a).strip()
    out = _tweak_bukgu_road_typos(out)
    return WS_RE.sub(" ", out).strip()


def _tweak_bukgu_road_typos(a: str) -> str:
    if not a.startswith("울산광역시 북구"):
        return a
    out = re.sub(r"호계로(\d)", r"호계로 \1", a)
    if "(" in out and out.count("(") > out.count(")"):
        out = f"{out.rstrip()})"
    return out


# 카카오가 못 찾는 지번·신축 단지 등 — 카카오/지도 검색으로 보정한 좌표 (상호+주소 패턴 매칭)
# (상호 포함 문자열, 도로 문자열 패턴 공백제거 포함, 위도, 경도, 표시용 도로주소)
# 긴 패턴부터 매칭하도록 호출 전 정렬함.
_MANUAL_COORDS_ROWS: tuple[tuple[str, str, float, float, str], ...] = (
    ("생생마트", "진장명촌토지구획정리지구115b", 35.5537984, 129.3524222, "울산광역시 북구 명촌로 94 (명촌동 · 명촌주공상가 인근, 수동보정 좌표)"),
    ("생생마트", "115b주공@", 35.5530073, 129.3532514, "울산광역시 북구 진장동 주공 상가 일대 (수동보정 좌표)"),
    ("(주)열린세상", "장춘로162", 35.5585832, 129.3258144, "울산광역시 중구 장춘로 162 (학산동)"),
    ("(주)재경", "종가3길18", 35.5606894, 129.2964272, "울산광역시 중구 종가3길 18 (유곡동)"),
    ("(주)코리아세븐울산매곡퍼스트점", "월드메르디앙", 35.6412442, 129.3599686, "울산광역시 북구 매산로 66 (매곡동, 월드메르디앙 인근 세븐)"),
    ("(주)황금철물공구안전", "삼산동)", 35.5356505, 129.3417657, "울산광역시 남구 삼산동 산업로 인근 철물상 (갈밭·삼산 지구, 수동보정 좌표)"),
    ("gs25진장으뜸점", "진장명촌지구30블럭", 35.5588544, 129.3567492, "울산광역시 북구 진장에너지로 일대 GS25 (명촌 진장 명의, 수동보정 좌표)"),
    ("h-마켓", "양정동500-29", 35.5611117, 129.3576137, "울산광역시 북구 진장로 2 단지내상가 인근 (양정 LH500 단지 일대 · 수동보정 좌표)"),
    ("강남유통", "화합로148번길20", 35.5375514, 129.3420635, "울산광역시 남구 화합로148번길 20 (삼산동)"),
    ("국일상회", "염포동419-1", 35.5330, 129.3908, "울산광역시 북구 염포동 419-1 (염포로 번지 근처 · 수동보정 좌표)"),
    ("농소농협매곡지소", "매곡동542-27", 35.6428159, 129.3537871, "울산광역시 북구 매곡1로 15 (농소농협 매곡지점)"),
    ("다모아매점", "56b-1n", 35.5659742, 129.3574077, "울산광역시 북구 진장로 56 (진장동 56블럭 근처, 수동보정 좌표)"),
    ("달감마트", "14블럭9롯트", 35.6405, 129.3485, "울산광역시 북구 신천동 신천지구 14블럭 일대 (수동보정 좌표)"),
    ("동인슈퍼", "양정동148-9", 35.5442, 129.3835, "울산광역시 북구 양정동 148-9 (율동 일대 · 수동보정 좌표)"),
    ("무룡상회", "무룡동935-18", 35.5880, 129.3665, "울산광역시 북구 무룡동 935-18 (무룡로 인근 · 수동보정 좌표)"),
    ("벽산아진관리사무소", "벽산아진아파트", 35.6199245, 129.4477614, "울산광역시 북구 정자동 벽산아진아파트"),
    ("세븐일레븐울산효문동진점", "효산1길9", 35.5710, 129.3710, "울산광역시 북구 효산1길 9 (효문동 일대 · 수동보정 좌표)"),
    ("신세기마트", "산176아이파크", 35.6349820, 129.3306438, "울산광역시 북구 달천로 50 (달천 아이파크 1차 인근)"),
    ("애뜰마트", "호계매곡1로55", 35.6349018, 129.3631838, "울산광역시 북구 호계매곡1로 55 (호계동)"),
    ("연암스토아", "연암동1276-31", 35.5810, 129.3595, "울산광역시 북구 연암동 1276-31 (상방로 인근 · 수동보정 좌표)"),
    (
        "천국할인마트",
        "블루마시티효성해링턴상가",
        35.6341373,
        129.4377241,
        "울산광역시 북구 화암7길 5 (산하동 블루마시티 효성해링턴 상가 인근)",
    ),
    ("위드미울산신천점", "호계로312", 35.6379156, 129.3472127, "울산광역시 북구 호계로 312 (신천동, 다이소 인근 · 수동보정 좌표)"),
    ("자동차마을매점", "자동차마을2층", 35.5788115, 129.3548722, "울산광역시 북구 진장유통로 95 (자동차마을 인근)"),
    ("지에스25울산송정행복주택점", "박상진4로56", 35.5993959, 129.3682464, "울산광역시 북구 박상진4로 56 (송정 LH1단지 인근)"),
    ("참진슈퍼", "126b1-1n", 35.5552709, 129.3533804, "울산광역시 북구 명촌12길 26 (진장 명촌 지구 근처, 수동보정 좌표)"),
    ("코끼리마트", "신천동309-3", 35.63788, 129.34718, "울산광역시 북구 신천동 309-3 (수동보정 좌표)"),
    ("크라운베이커리명촌점", "평창리비에르", 35.5533027, 129.3551488, "울산광역시 북구 명촌동 평창리비에르 상가 인근"),
    ("태봉종합상사", "해오름4길11", 35.5628797, 129.3443080, "울산광역시 중구 해오름4길 11 (남외동)"),
    ("팝스토아매곡점", "신천로60", 35.6438175, 129.3516514, "울산광역시 북구 신천로 60 (매곡휴먼시아 제상가 인근)"),
    ("한농스토아", "신천동309-2", 35.63795, 129.34725, "울산광역시 북구 신천동 309-2 (수동보정 좌표)"),
    ("훈이마트", "15b-11n", 35.6335, 129.3645, "울산광역시 북구 호계동 호계 매곡 15블럭 일대 (수동보정 좌표)"),
)


def lookup_manual_coordinates(name: str, road: str):
    """카카오 실패 매장 보정 — (표시 도로명, 위도, 경도) 또는 None."""
    n = collapse(name).lower().replace(" ", "").replace("|", "")
    r = collapse(road).replace(" ", "")
    sorted_rows = sorted(
        _MANUAL_COORDS_ROWS,
        key=lambda row: len(row[0]) + len(row[1]),
        reverse=True,
    )
    for name_needle, road_needle, lat, lng, road_out in sorted_rows:
        nd = collapse(name_needle).lower().replace(" ", "").replace("|", "")
        if nd and nd not in n:
            continue
        rd = road_needle.replace(" ", "").lower().replace("|", "")
        if rd and rd not in r.lower():
            continue
        return (road_out, lat, lng)
    return None


def normalize_phone(raw: object) -> str | None:
    if raw is None:
        return None
    s = collapse(str(raw))
    if not s or s in ("-", "—", "."):
        return None
    digits = re.sub(r"\D+", "", s)
    if not digits:
        return None
    # 휴대폰 / 지역번호 포함 그대로
    if digits.startswith("010") and len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if digits.startswith("052"):
        body = digits[3:]
        if len(body) == 7:
            return f"052-{body[:3]}-{body[3:]}"
        if len(body) == 8:
            return f"052-{body[:4]}-{body[4:]}"
        if len(body) == 9:
            return f"052-{body[:3]}-{body[3:]}"
        return None
    if len(digits) == 7:
        # 엑셀에 지역번호 생략(052) — 북구 일반
        return f"052-{digits[:3]}-{digits[3:]}"
    if len(digits) == 8 and digits[0] in "23456789":
        return f"052-{digits[:3]}-{digits[3:]}"
    return s.replace(" ", "")


def parse_float(v: object) -> float | None:
    try:
        f = float(v)  # type: ignore[arg-type]
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def kakao_geocode(q: str, key: str) -> tuple[float | None, float | None]:
    for base, extra in (
        (GEOCODE_URL, {"query": q}),
        (KEYWORD_URL, {"query": q, "size": "5"}),
    ):
        req = urllib.request.Request(
            f"{base}?{urllib.parse.urlencode(extra)}",
            headers={"Authorization": f"KakaoAK {key}"},
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except Exception:
            continue
        docs = data.get("documents") or []
        best: tuple[float, float] | None = None
        for d in docs:
            lat = parse_float(d.get("y"))
            lng = parse_float(d.get("x"))
            if lat is None or lng is None:
                continue
            if not in_ulsan_bukgu_bbox(lat, lng):
                continue
            best = (lat, lng)
            break
        if best:
            return best
    return None, None


def addr_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(v: str) -> None:
        t = collapse(v)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    add(addr)
    head = re.sub(r"\(.*?\)", "", addr).strip()
    add(collapse(head))
    head2 = normalize_addr(head)
    add(head2)
    return out


def in_ulsan_bukgu_bbox(lat: float, lng: float) -> bool:
    return 35.48 <= lat <= 35.72 and 129.25 <= lng <= 129.52


def load_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        raw = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, list[float]] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if isinstance(v, list) and len(v) == 2:
                out[str(k)] = [float(v[0]), float(v[1])]
    return out


def save_cache(cache: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def stable_id_digest(name: str, road: str) -> str:
    h = hashlib.sha1((name + "\x1e" + road).encode("utf-8")).hexdigest()[:12]
    return f"ulsan-bukgu-{h}"


SHORT_CODE_RE = re.compile(r"^[a-zA-Z0-9]{6}$")


def load_existing_shortcodes(paths: tuple[Path, ...]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for p in paths:
        if not p.exists():
            continue
        try:
            rows = json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            rid = str(row.get("id") or "")
            sc = row.get("shortCode")
            if (
                isinstance(sc, str)
                and SHORT_CODE_RE.match(sc.strip())
                and rid
            ):
                mapping[rid] = sc.strip()
    return mapping


def read_sheet(ws) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        try:
            n = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        if n < 1:
            continue
        name = collapse(str(row[1] or ""))
        addr_raw = collapse(str(row[2] or ""))
        ph_raw = normalize_phone(row[3]) if len(row) > 3 else None
        if not name or not addr_raw:
            continue
        phone = collapse(ph_raw) if ph_raw else ""
        rows.append((name, addr_raw, phone))
    return rows


def pick_longer_digits_phone(a: str, b: str) -> str:
    if not a.strip():
        return b
    if not b.strip():
        return a
    da = re.sub(r"\D", "", a)
    db = re.sub(r"\D", "", b)
    if len(db) > len(da):
        return b
    if len(db) < len(da):
        return a
    return a if len(a) >= len(b) else b


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = ap.parse_args()
    xp = args.xlsx.expanduser()
    if not xp.exists():
        raise SystemExit(f"xlsx 없음: {xp}")
    key = load_kakao_key()
    if not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    wb = load_workbook(xp, read_only=True, data_only=True)
    if SHEET_TRASH not in wb.sheetnames or SHEET_SPECIAL not in wb.sheetnames:
        raise SystemExit(f"필요 시트 없음. 시트 목록={wb.sheetnames}")

    trash_in = read_sheet(wb[SHEET_TRASH])
    special_in = read_sheet(wb[SHEET_SPECIAL])
    wb.close()

    # 이름+원문주소 로 통합 레코드(전화번호·종량제/특수 플래그 병합)
    combined: dict[str, dict[str, object]] = {}

    def feed(name: str, addr_raw: str, phone_csv: str, *, trash: bool, special: bool) -> None:
        road = normalize_addr(addr_raw)
        kname = collapse(name.lower().replace("\xa0", ""))
        kfix = "|".join((kname, road.replace(" ", "")))
        if kfix not in combined:
            combined[kfix] = {
                "name": name,
                "roadAddress": road,
                "address": "울산광역시 북구",
                "hasTrashBag": trash,
                "hasSpecialBag": special,
                "phone_pick": "",
            }
        row = combined[kfix]
        row["hasTrashBag"] = bool(row["hasTrashBag"] or trash)
        row["hasSpecialBag"] = bool(row["hasSpecialBag"] or special)
        if phone_csv.strip():
            row["phone_pick"] = pick_longer_digits_phone(
                str(row.get("phone_pick") or ""), phone_csv
            )

    for name, ar, ph in trash_in:
        feed(name, ar, ph, trash=True, special=False)
    for name, ar, ph in special_in:
        feed(name, ar, ph, trash=False, special=True)

    cache = load_cache()
    ids_sc = load_existing_shortcodes((OUT_TRASH, OUT_SPECIAL))
    ref_date = date.today().isoformat()

    enriched: dict[str, dict[str, object]] = {}
    n_geo = 0
    for kfix, r in sorted(
        combined.items(), key=lambda x: (str(x[1]["name"]), str(x[1]["roadAddress"]))
    ):
        name = str(r["name"])
        road_orig = str(r["roadAddress"])
        ck = hashlib.sha1(
            ("ulsan:bukgu:" + road_orig + "\x1f" + name).encode()
        ).hexdigest()[:24]
        manual = lookup_manual_coordinates(name, road_orig)
        latlng = cache.get(ck)
        lat = lng = None
        road_display = road_orig

        if manual:
            road_display, lat, lng = manual
            cache[ck] = [float(lat), float(lng)]
            sys.stderr.write(
                f"[manual coord] {name}\t{road_display}\n",
            )
        elif latlng and len(latlng) == 2:
            lat, lng = latlng
            road_display = road_orig
        else:
            for base in addr_variants(road_orig):
                qlist = (
                    base,
                    f"{name} {base}",
                    f"{name}",
                    f"울산광역시 북구 {name}",
                )
                for q in qlist:
                    glat, glng = kakao_geocode(q, key)
                    if glat is not None and glng is not None:
                        lat, lng = glat, glng
                        break
                if lat is not None:
                    break
            if lat is None:
                sys.stderr.write(
                    f"[skip geocode failed] {name}\t{road_orig}\n",
                )
                continue
            cache[ck] = [float(lat), float(lng)]  # type: ignore[list-item]
            n_geo += 1
            if n_geo % 50 == 0:
                save_cache(cache)
                sys.stderr.write(f"[geocode] +{n_geo} …\n")
            time.sleep(GEOCODE_DELAY)

        if not isinstance(lat, (float, int)) or not isinstance(lng, (float, int)):
            continue
        if not in_ulsan_bukgu_bbox(float(lat), float(lng)):
            sys.stderr.write(f"[bbox skip] {name}\t{road_orig}\t{lat},{lng}\n")
            continue

        sid = stable_id_digest(name, road_orig)
        phone_disp = normalize_phone(str(r.get("phone_pick") or "").strip())
        addr_dist = "울산광역시 북구"
        for pfx in ("울산광역시 중구", "울산광역시 남구", "울산광역시 동구"):
            if road_display.startswith(pfx):
                addr_dist = pfx
                break
        enriched[kfix] = {
            **r,
            "id": sid,
            "roadAddress": road_display,
            "address": addr_dist,
            "lat": round(float(lat), 7),
            "lng": round(float(lng), 7),
            "businessStatus": "영업",
            "hasTrashBag": r["hasTrashBag"],
            "hasSpecialBag": r["hasSpecialBag"],
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": ref_date,
            "shortCode": ids_sc.get(sid),
            "phone": phone_disp if phone_disp else None,
        }
        sc = enriched[kfix].get("shortCode")
        if isinstance(sc, str) and SHORT_CODE_RE.match(sc.strip()):
            enriched[kfix]["shortCode"] = sc.strip()

    save_cache(cache)

    trash_out: list[dict] = []
    special_out: list[dict] = []
    for rec in enriched.values():
        base = dict(rec)
        common = {
            "id": base["id"],
            "name": base["name"],
            "lat": base["lat"],
            "lng": base["lng"],
            "roadAddress": base["roadAddress"],
            "address": base["address"],
            "businessStatus": base["businessStatus"],
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": base["dataReferenceDate"],
        }
        tel = base.get("phone")
        if isinstance(tel, str) and tel.strip():
            common["phone"] = tel.strip()
        if base.get("shortCode"):
            common["shortCode"] = base["shortCode"]

        # 겹치는 판매처는 동일 id·shortCode가 두 JSON에 중복되면 shortcodes:assign 전역 유일성 검사 실패.
        # → 종량제+특수 겸업은 trash JSON 한 건만 두고 hasTrashBag·hasSpecialBag을 모두 true 로 둠.
        ht = bool(base["hasTrashBag"])
        hs = bool(base["hasSpecialBag"])
        if ht:
            trash_out.append(
                {
                    **common,
                    "hasTrashBag": True,
                    "hasSpecialBag": hs,
                }
            )
        if hs and not ht:
            special_out.append(
                {
                    **common,
                    "hasTrashBag": False,
                    "hasSpecialBag": True,
                }
            )

    trash_out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    special_out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_TRASH.parent.mkdir(parents=True, exist_ok=True)
    OUT_TRASH.write_text(json.dumps(trash_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    OUT_SPECIAL.write_text(
        json.dumps(special_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8",
    )

    skip = len(combined) - len(enriched)
    print(f"wrote {len(trash_out)} → {OUT_TRASH.name}")
    print(f"wrote {len(special_out)} → {OUT_SPECIAL.name}")
    if skip:
        print(f"경고: 지오코딩·bbox 실패로 제외={skip}건 (stderr 참고)")
    print("다음: cd frontend && npm run shortcodes:assign")


if __name__ == "__main__":
    main()
