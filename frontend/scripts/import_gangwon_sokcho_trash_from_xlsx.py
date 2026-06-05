#!/usr/bin/env python3
"""
강원특별자치도 속초시 종량제봉투·PP마대(불연성) 판매소 xlsx → stores.gangwon-sokcho-trash.json

종량제: `속초시(260601) 봉투판매소 현황(정보공개용).xlsx` 시트 `종량제 봉투판매소`
PP마대: `속초시_종량제 PP마대 판매소 현황(동별).xlsx` 시트 `마대판매처`

  python3 scripts/import_gangwon_sokcho_trash_from_xlsx.py --refresh
  python3 scripts/import_gangwon_sokcho_trash_from_xlsx.py \\
    --input-trash ~/Downloads/속초시(260601)\\ 봉투판매소\\ 현황(정보공개용).xlsx \\
    --input-special ~/Downloads/속초시_종량제\\ PP마대\\ 판매소\\ 현황(동별).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.gangwon-sokcho-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-gangwon-sokcho-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_TRASH = DL / "속초시(260601) 봉투판매소 현황(정보공개용).xlsx"
DEFAULT_SPECIAL = DL / "속초시_종량제 PP마대 판매소 현황(동별).xlsx"
SHEET_TRASH = "종량제 봉투판매소"
SHEET_SPECIAL = "마대판매처"
TRASH_START_ROW = 3
SPECIAL_START_ROW = 3
REF_DATE = "2026-06-01"
CACHE_VERSION = "v1-sokcho-2026"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def in_sokcho_bbox(lat: float, lng: float) -> bool:
    # 설악산·도문 일대 포함
    return 38.16 <= lat <= 38.25 and 128.48 <= lng <= 128.62


def sokcho_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if any(x in t for x in ("고성군", "양양", "인제", "춘천", "강릉")):
        return False
    if "속초시" in t:
        return True
    return ("강원" in t or "강원특별" in t) and "속초" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^강원\s+", "강원특별자치도 ", a)
    a = re.sub(r"^강원도\s+", "강원특별자치도 ", a)
    if not a:
        return ""
    if a.startswith("강원특별자치도"):
        return a
    if a.startswith("속초시"):
        return f"강원특별자치도 {a}"
    return f"강원특별자치도 속초시 {a}"


def extract_paren_hint(raw: str) -> tuple[str, str]:
    a = collapse(raw)
    hint = ""
    m = re.search(r"\(([^)]+)\)", a)
    if m:
        parts = [collapse(p) for p in m.group(1).split(",") if collapse(p)]
        for p in parts:
            if re.search(r"(동|읍|면|리|가)$", p.replace(" ", "")):
                hint = p
                break
        if not hint and parts:
            hint = parts[0]
        a = collapse(re.sub(r"\s*\([^)]*\)\s*", " ", a))
    return a, hint


def is_likely_jibeon(tail: str) -> bool:
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", tail))


def normalize_sokcho_addr(addr_raw: str, dong: str = "") -> str:
    core, hint = extract_paren_hint(addr_raw)
    a = collapse(core)
    a = re.sub(r"(?<![가-힣])랑동", "영랑동", a)
    a = re.sub(r"\b라동\b", "영랑동", a)
    a = re.sub(r"\)+\s*$", "", a)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+\d+동\s+\d+층.*$", "", a)
    a = re.sub(r"\s+제상가\d+.*$", "", a)
    a = re.sub(r"속초시\s+속초시", "속초시", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"\.\d+/\d+\s*$", "", a)
    a = re.sub(r"\s+\d+/\d+\s*$", "", a)
    a = re.sub(r"속초시(?=[가-힣])", "속초시 ", a)
    a = re.sub(r"속초시동", "속초시 동", a)
    a = re.sub(r"([가-힣]+동)(\d+)", r"\1 \2", a)
    a = re.sub(r"([가-힣]+(?:로|길|대로))(\d+)", r"\1 \2", a)
    a = re.sub(r"(\d+)호\s*$", r"\1", a)
    dong_clean = collapse(dong)
    if hint and hint not in a.replace(" ", ""):
        if ROAD_RE.search(a) or not re.search(r"(동|가|리)\s", a):
            a = f"{hint} {a}" if hint not in a else a
    elif dong_clean and dong_clean.replace(" ", "") not in a.replace(" ", ""):
        if ROAD_RE.search(a) and not re.search(r"(동|가|리)(\s|\d)", a):
            a_body = re.sub(r"^속초시\s+", "", a)
            a = f"{dong_clean} {a_body}"
    a = re.sub(r"^([가-힣]+동)\s+속초시\s+", r"\1 ", a)
    if not re.search(r"^(강원|속초)", a.replace(" ", "")):
        if re.match(r"^[가-힣]+(?:로|길|대로)", a):
            a = f"속초시 {a}"
        else:
            a = f"속초시 {a}"
    return format_display_addr(a)


def sokcho_tail(full: str) -> str:
    for prefix in (
        "강원특별자치도 속초시 ",
        "강원특별자치도 ",
        "강원도 속초시 ",
        "속초시 ",
    ):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def lot_query_variants(tail: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = LOT_RE.match(tail)
    if m:
        prefix, main, sub = m.group("prefix"), m.group("main"), m.group("sub")
        subs: list[str | None] = [sub, "1", "2", "3", "4", "5"] if sub else [None]
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def geocode_query_variants(addr_raw: str, name: str, dong: str = "") -> list[str]:
    norm = normalize_sokcho_addr(addr_raw, dong)
    tail = sokcho_tail(norm)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    if is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push(f"강원 속초시 {tv}")
            push(f"강원특별자치도 속초시 {tv}")
    else:
        push(f"강원 속초시 {tail}")
        push(f"강원특별자치도 속초시 {tail}")
        push(norm)
        compact = re.sub(r"\s+", "", tail)
        if compact != tail.replace(" ", ""):
            push(f"강원특별자치도 속초시 {compact}")
        road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
        if road_only and road_only != tail:
            push(f"강원특별자치도 속초시 {road_only}")

    push(f"{name} 속초시")
    push(f"{name} 속초")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


@dataclass
class SokchoRow:
    name: str
    addr_raw: str
    dong: str
    has_trash: bool
    has_special: bool


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_sokcho_bbox(lat, lng):
        return None
    if not sokcho_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_sokcho_bbox(lat, lng):
        return None
    if not sokcho_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def area_fallback(addr_raw: str, key: str, dong: str) -> GeoHit | None:
    tail = sokcho_tail(normalize_sokcho_addr(addr_raw, dong))
    area = collapse(dong)
    if not area and "장선천" in tail:
        area = "장사동"
    if not area:
        m = re.match(r"^(.+?(?:동|읍|면|리))", tail.replace(" ", ""))
        area = collapse(m.group(1)) if m else ""
    if not area:
        return None
    for q in (f"강원 속초시 {area}", f"강원특별자치도 속초시 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str, dong: str = "") -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name, dong):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return area_fallback(addr_raw, key, dong)


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def iter_special_rows(path: Path) -> list[SokchoRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_SPECIAL] if SHEET_SPECIAL in wb.sheetnames else wb.active
    out: list[SokchoRow] = []
    last_dong = ""
    for r in range(SPECIAL_START_ROW, (ws.max_row or 0) + 1):
        dong = cell_str(ws.cell(r, 1).value)
        if dong and dong not in ("거래처명",):
            last_dong = dong
        name = cell_str(ws.cell(r, 2).value)
        addr_raw = cell_str(ws.cell(r, 3).value)
        if not name or not addr_raw or name in ("거래처명",):
            continue
        out.append(
            SokchoRow(
                name=name,
                addr_raw=addr_raw,
                dong=last_dong,
                has_trash=False,
                has_special=True,
            )
        )
    wb.close()
    return out


def iter_trash_rows(path: Path) -> list[SokchoRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_TRASH] if SHEET_TRASH in wb.sheetnames else wb.active
    out: list[SokchoRow] = []
    for r in range(TRASH_START_ROW, (ws.max_row or 0) + 1):
        name = cell_str(ws.cell(r, 2).value)
        addr_raw = cell_str(ws.cell(r, 3).value)
        if not name or not addr_raw or name in ("업 체 명", "업체명"):
            continue
        out.append(
            SokchoRow(
                name=name,
                addr_raw=addr_raw,
                dong="",
                has_trash=True,
                has_special=False,
            )
        )
    wb.close()
    return out


def merge_rows(special_rows: list[SokchoRow], trash_rows: list[SokchoRow]) -> list[SokchoRow]:
    merged: dict[str, SokchoRow] = {}
    for row in special_rows:
        merged[row.name] = row
    for row in trash_rows:
        if row.name in merged:
            existing = merged[row.name]
            existing.has_trash = True
        else:
            merged[row.name] = row
    return list(merged.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-trash", type=Path, default=DEFAULT_TRASH)
    ap.add_argument("--input-special", type=Path, default=DEFAULT_SPECIAL)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    trash_path = args.input_trash.expanduser()
    special_path = args.input_special.expanduser()
    if not trash_path.is_file():
        raise SystemExit(f"종량제 파일 없음: {trash_path}")
    if not special_path.is_file():
        raise SystemExit(f"PP마대 파일 없음: {special_path}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    rows = merge_rows(iter_special_rows(special_path), iter_trash_rows(trash_path))
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0

    for row in rows:
        if not row.has_trash and not row.has_special:
            continue
        norm_addr = normalize_sokcho_addr(row.addr_raw, row.dong)
        ck = cache_key(row.name, norm_addr)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else norm_addr
                jibeon = str(raw[3]) if len(raw) > 3 else norm_addr
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jibeon)

        if hit is None and allow:
            hit = resolve_geocode(row.addr_raw, row.name, key, row.dong)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {row.name}\t{norm_addr}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = hit.road or cr or norm_addr
            hit.jibeon = cj or hit.jibeon or norm_addr
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{norm_addr}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"gangwon-sokcho-trash-{rid}",
                "name": row.name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": hit.road,
                "address": hit.jibeon,
                "businessStatus": "영업",
                "hasTrashBag": row.has_trash,
                "hasSpecialBag": row.has_special,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": REF_DATE,
            }
        )

    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={REF_DATE}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["속초시"])


if __name__ == "__main__":
    main()
