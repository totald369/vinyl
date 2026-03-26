#!/usr/bin/env python3
"""
Geocode 서울 강남구 종량제봉투 판매소 Excel and merge into stores.sample.json.
All rows are 종량제봉투 판매처 → hasTrashBag=true.
"""

import json
import math
import os
import re
import sys
import time
import urllib.request
import urllib.parse
import openpyxl
from pathlib import Path

GANGNAM_XLSX = Path.home() / "Downloads" / "종량제봉투 판매소목록(25.3.13.기준).xlsx"
OUT_PATH = Path(__file__).resolve().parent.parent / "public" / "data" / "stores.sample.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache.json"

KAKAO_REST_KEY = os.environ.get("KAKAO_REST_API_KEY") or "fd1c94f46de9b58135650e9fba4b5320"
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
DATA_REF = "2025-03-13"


def to_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


TYPO_FIXES = (
    ("서울특볈시", "서울특별시"),
    ("서울특별실", "서울특별시"),
    ("서을시", "서울특별시"),
    ("서울특별시시", "서울특별시"),
    ("강남구남부", "강남구 남부"),
    ("강남구테헤란", "강남구 테헤란"),
    ("강남구테헤란로", "강남구 테헤란로"),
)


def relax_ro_house_number(s):
    """테헤란로216 → 테헤란로 216 (건물번호가 로에 붙은 형태)."""
    return re.sub(
        r"(테헤란로|강남대로|언주로|삼성로|남부순환로)(\d{2,5})(?![길0-9])",
        r"\1 \2",
        s,
    )


def jibun_query_variants(addr):
    """'(논현동 125-8)', '(신514-21)' 형태 지번을 서울 강남 구 단위 주소로 변환."""
    if not addr or ("강남" not in addr and "서울" not in addr):
        return []
    out = []
    for m in re.finditer(r"\(([가-힣]+동)\s+(\d+[\-\d]*)\)", addr):
        out.append(f"서울특별시 강남구 {m.group(1)} {m.group(2)}")
    sm = re.search(r"\(신(\d+[\-\d]*)\)", addr)
    if sm:
        out.append(f"서울특별시 강남구 신사동 {sm.group(1)}")
    sm2 = re.search(r"\(신사(\d+[\-\d]*)\)", addr)
    if sm2:
        out.append(f"서울특별시 강남구 신사동 {sm2.group(1)}")
    return out


def comma_head_road(addr):
    """'도로명 … 22, 112호 …' 같이 쉼표 뒤가 검색을 망치는 경우 앞부분만."""
    if "," not in addr:
        return None
    head = addr.split(",")[0].strip()
    if len(head) >= 12 and re.search(r"\d", head):
        return head
    return None


def sanitize_place_name_for_keyword(name):
    """카카오 키워드 검색에 맞게 상호 정리."""
    s = (name or "").strip()
    s = re.sub(r"^[㈜\(]\s*주\s*\)?\s*|^㈜|^\(주\)\s*|주식회사\s*", "", s).strip()
    s = re.sub(r"(세븐일레븐|미니스톱)([가-힣])", r"\1 \2", s)
    s = re.sub(r"(gs25|지에스25)([가-힣])", r"\1 \2", s, flags=re.I)
    return s


def strip_trailing_floor(s):
    """'… 1F' / '… 2F' 등은 카카오 도로명 검색에서 자주 실패함."""
    t = re.sub(r"\s+[12]F\s*$", "", s, flags=re.I).strip()
    t = re.sub(r",\s*[12]F\s*$", "", t, flags=re.I).strip()
    return t


def relax_ro_gil(s):
    """테헤란로51길 → 테헤란로 51길 (한글로+숫자길 붙은 형태)."""
    return re.sub(r"([가-힣]+로)(\d+길)", r"\1 \2", s)


def relax_gil_house(s):
    """151길26 → 151길 26."""
    return re.sub(r"(\d+길)(\d)", r"\1 \2", s)


def address_variants(raw):
    a = (raw or "").strip()
    if not a:
        return []
    seen = set()
    out = []

    def add(x):
        x = x.strip()
        if not x:
            return
        for v in (
            x,
            strip_trailing_floor(x),
            relax_ro_gil(x),
            relax_ro_gil(strip_trailing_floor(x)),
            relax_gil_house(relax_ro_gil(x)),
            relax_ro_house_number(x),
            relax_ro_house_number(strip_trailing_floor(x)),
            relax_ro_gil(relax_ro_house_number(strip_trailing_floor(x))),
            relax_gil_house(relax_ro_gil(relax_ro_house_number(strip_trailing_floor(x)))),
        ):
            if v and v not in seen:
                seen.add(v)
                out.append(v)

    add(a)
    for old, new in TYPO_FIXES:
        if old in a:
            add(a.replace(old, new))
    if a.startswith("강남구"):
        add("서울특별시 " + a)
    if a.startswith("서울시 ") and "강남구" in a:
        add(a.replace("서울시 ", "서울특별시 ", 1))
    base = a.split(", 1F")[0].split(", 2F")[0].rstrip(",").strip()
    if base != a:
        add(base)
    for jb in jibun_query_variants(a):
        add(jb)
    ch = comma_head_road(a)
    if ch:
        add(ch)
    return out


def kakao_geocode(address, place_name=None):
    for q in address_variants(address):
        coords = _kakao_api(GEOCODE_URL, q)
        if coords:
            return coords
    for q in address_variants(address):
        coords = _kakao_api(KEYWORD_URL, q)
        if coords:
            return coords
    # 도로명이 DB와 달라 주소 검색이 안 되는 경우: 업체명 + 강남 (키워드)
    if place_name:
        nm = sanitize_place_name_for_keyword(place_name)
        for q in (f"{nm} 강남구", f"{nm} 서울 강남", nm):
            if len(q) < 2:
                continue
            coords = _kakao_api(KEYWORD_URL, q)
            if coords:
                return coords
        if "더프레시" in nm:
            m = re.search(r"(지에스\s*더프레시[^점]*점?)", nm)
            if m:
                for q in (f"{m.group(1).strip()} 강남구",):
                    coords = _kakao_api(KEYWORD_URL, q)
                    if coords:
                        return coords
    return None


def _kakao_api(base_url, query):
    params = urllib.parse.urlencode({"query": query, "size": "1"})
    url = f"{base_url}?{params}"
    req = urllib.request.Request(url, headers={"Authorization": f"KakaoAK {KAKAO_REST_KEY}"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            docs = data.get("documents", [])
            if docs:
                lat = to_float(docs[0].get("y"))
                lng = to_float(docs[0].get("x"))
                if lat and lng:
                    return (lat, lng)
    except Exception:
        pass
    return None


def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def coords_from_cache(cache, key):
    v = cache.get(key)
    if not v or not isinstance(v, (list, tuple)) or len(v) != 2:
        return None
    lat, lng = to_float(v[0]), to_float(v[1])
    if lat is None or lng is None:
        return None
    return (lat, lng)


def put_coords_cache(cache, key, coords):
    if coords:
        cache[key] = [coords[0], coords[1]]


def save_cache(cache):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)


def gangnam_loc(s):
    return (s.get("roadAddress") or "") + (s.get("address") or "")


def norm_addr(x):
    if not x:
        return ""
    return "".join(str(x).split()).lower()


def excel_row_in_json(stores, name, addr):
    """엑셀 한 행이 이미 JSON에 좌표 있는 매장으로 반영됐는지(이름+주소 유사)."""
    n = name.strip()
    an = norm_addr(addr)
    for s in stores:
        if s["name"].strip() != n:
            continue
        if "강남구" not in gangnam_loc(s):
            continue
        if s.get("lat") is None or s.get("lng") is None:
            continue
        ln = norm_addr(gangnam_loc(s))
        if an and ln and (an in ln or ln in an):
            return True
    return False


def find_existing_match(existing, name, addr):
    n = name.strip()
    addr_n = norm_addr(addr)
    candidates = [
        s
        for s in existing
        if s["name"].strip() == n and "강남구" in gangnam_loc(s)
    ]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    for s in candidates:
        loc_n = norm_addr(gangnam_loc(s))
        if addr_n and loc_n and (addr_n in loc_n or loc_n in addr_n):
            return s
    return None


def main():
    if not GANGNAM_XLSX.exists():
        print(f"Missing file: {GANGNAM_XLSX}")
        return

    print("Reading 강남구 Excel...")
    wb = openpyxl.load_workbook(str(GANGNAM_XLSX), read_only=True, data_only=True)
    ws = wb["최종"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()
    print(f"  → {len(rows)} rows (from row 4)")

    print("Loading stores.sample.json...")
    with open(OUT_PATH, "r", encoding="utf-8") as f:
        existing = json.load(f)
    print(f"  → {len(existing)} stores")

    max_id = max(int(s["id"]) for s in existing)
    cache = load_cache()
    new_stores = []
    updated = 0
    geocoded = 0
    failed = 0
    api_calls = 0

    for row in rows:
        name = str(row[2]).strip() if row[2] is not None else ""
        addr = str(row[3]).strip() if row[3] is not None else ""
        if not name or not addr:
            continue

        match = find_existing_match(existing, name, addr)
        if match:
            match["hasTrashBag"] = True
            match["dataReferenceDate"] = DATA_REF
            updated += 1
            continue

        cache_key = addr
        coords = coords_from_cache(cache, cache_key)
        if coords is None:
            coords = kakao_geocode(addr, place_name=name)
            put_coords_cache(cache, cache_key, coords)
            api_calls += 1
            if api_calls % 100 == 0:
                print(f"  API: {api_calls}, new: {len(new_stores)}, fail: {failed}")
                save_cache(cache)
            time.sleep(0.05)

        if not coords:
            failed += 1
            print(f"  Geocode fail: {name} | {addr}")
            continue

        geocoded += 1
        max_id += 1
        new_stores.append(
            {
                "id": str(max_id),
                "name": name,
                "lat": coords[0],
                "lng": coords[1],
                "roadAddress": addr,
                "address": addr,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": DATA_REF,
            }
        )

    save_cache(cache)
    merged = existing + new_stores
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    gn = [s for s in merged if "강남구" in gangnam_loc(s)]
    print("\n강남구 merge 완료:")
    print(f"  기존 매칭 → hasTrashBag 갱신: {updated}")
    print(f"  신규 지오코딩: {geocoded}")
    print(f"  지오코딩 실패: {failed}")
    print(f"  신규 추가: {len(new_stores)}")
    print(f"  전체 매장: {len(merged)}")
    print(f"  강남구 주소 포함 매장 수: {len(gn)}")
    print(f"Written: {OUT_PATH}")


def retry_missing_only():
    """첫 실행에서 지오코딩 실패한 엑셀 행만 보강 (주소 변형 로직 반영)."""
    if not GANGNAM_XLSX.exists():
        print(f"Missing file: {GANGNAM_XLSX}")
        return

    wb = openpyxl.load_workbook(str(GANGNAM_XLSX), read_only=True, data_only=True)
    ws = wb["최종"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()

    with open(OUT_PATH, "r", encoding="utf-8") as f:
        stores = json.load(f)

    max_id = max(int(s["id"]) for s in stores)
    cache = load_cache()
    added = 0
    still_fail = 0
    api_calls = 0

    for row in rows:
        name = str(row[2]).strip() if row[2] is not None else ""
        addr = str(row[3]).strip() if row[3] is not None else ""
        if not name or not addr:
            continue
        if excel_row_in_json(stores, name, addr):
            continue
        match = find_existing_match(stores, name, addr)
        if match:
            match["hasTrashBag"] = True
            match["dataReferenceDate"] = DATA_REF
            continue

        cache_key = "retry:" + norm_addr(addr)
        coords = coords_from_cache(cache, cache_key)
        if coords is None:
            coords = kakao_geocode(addr, place_name=name)
            put_coords_cache(cache, cache_key, coords)
            api_calls += 1
            time.sleep(0.05)

        if not coords:
            still_fail += 1
            print(f"  Still fail: {name} | {addr}")
            continue

        max_id += 1
        stores.append(
            {
                "id": str(max_id),
                "name": name,
                "lat": coords[0],
                "lng": coords[1],
                "roadAddress": addr,
                "address": addr,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": False,
                "adminVerified": False,
                "dataReferenceDate": DATA_REF,
            }
        )
        added += 1

    save_cache(cache)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(stores, f, ensure_ascii=False, indent=2)
    print(f"\nretry: added {added}, still_fail {still_fail}, api_calls {api_calls}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--retry":
        retry_missing_only()
    else:
        main()
