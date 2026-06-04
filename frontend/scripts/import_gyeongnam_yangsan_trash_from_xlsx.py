#!/usr/bin/env python3
"""
경상남도 양산시 종량제봉투 판매소 xlsx → stores.gyeongnam-yangsan-trash.json

시트 `1. 종량제판매소 위치 정보` (6행~):
  연번 | 판매소 명칭 | 도로명 주소 | 상세주소 | … | 용량별 Y(5ℓ~)

  python3 scripts/import_gyeongnam_yangsan_trash_from_xlsx.py \\
    --input ~/Downloads/양산시정보공개자료\\(종량제봉투\\ 판매처\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.gyeongnam-yangsan-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-gyeongnam-yangsan-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "양산시정보공개자료(종량제봉투 판매처).xlsx"
SHEET_NAME = "1. 종량제판매소 위치 정보"
DATA_START_ROW = 6
CACHE_VERSION = "v1-yangsan"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    return collapse(str(val or "")).upper() in ("Y", "YES", "예", "여", "O", "○")


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def ref_date_from_path(p: Path) -> str:
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).date().isoformat()
    except OSError:
        return datetime.now().date().isoformat()


def in_yangsan_bbox(lat: float, lng: float) -> bool:
    return 35.18 <= lat <= 35.52 and 128.88 <= lng <= 129.32


def yangsan_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if "양산시" in t:
        return True
    if any(x in t for x in ("광주광역시", "광주", "부산광역시", "부산", "울산", "서울", "인천")):
        return False
    return ("경상남도" in t or "경남" in t) and "양산" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^경남\s+", "경상남도 ", a)
    if a.startswith("경상남도"):
        return a
    if a.startswith("양산시"):
        return f"경상남도 {a}"
    if yangsan_in_text(a) or re.search(r"(동|읍|면)\s", a):
        return f"경상남도 양산시 {a}"
    return a


def normalize_yangsan_addr(road_raw: str, detail_raw: str) -> str:
    a = collapse(road_raw)
    if not a:
        return ""
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*@\s*", " ", a)
    if not a.startswith("양산") and not a.startswith("경상"):
        a = f"양산시 {a}"
    a = re.sub(r"^경남\s+", "경상남도 ", a)
    full = format_display_addr(a)
    detail = collapse(detail_raw)
    if detail:
        full = f"{full} {detail}"
    return collapse(full)


def geocode_target(full: str) -> str:
    a = collapse(full)
    a = re.sub(r",\s*.*$", "", a)
    a = re.sub(r"\s+\d+\s*호\s*$", "", a)
    return a


def road_format_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(addr)
    compact = addr.replace(" ", "")
    push(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", compact))
    push(re.sub(r"([가-힣]+로)(\d+)", r"\1 \2", compact))
    return out


def yangsan_tail(full: str) -> str:
    for prefix in ("경상남도 양산시 ", "경상남도 ", "경남 양산시 ", "양산시 "):
        if full.startswith(prefix):
            return collapse(full[len(prefix) :])
    return full


def geocode_query_variants(road_full: str, name: str) -> list[str]:
    target = geocode_target(road_full)
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    for base in road_format_variants(target):
        tail = yangsan_tail(base)
        push(f"경남 양산시 {tail}")
        push(f"경상남도 양산시 {tail}")
        push(base)
        road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
        if road_only and road_only != tail:
            push(f"경남 양산시 {road_only}")
        m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
        if m:
            push(f"경남 양산시 {collapse(m.group(1))}")

    push(f"{name} 양산")
    push(f"{name} 양산시")
    compact = re.sub(r"\s+", "", name)
    if compact != name:
        push(f"{compact} 양산시")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_yangsan_bbox(lat, lng):
        return None
    if not yangsan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_yangsan_bbox(lat, lng):
        return None
    if not yangsan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    return data.get("documents") or []


def area_fallback(road_full: str, key: str, display_road: str) -> GeoHit | None:
    tail = yangsan_tail(geocode_target(road_full))
    m = re.match(r"^(.+?(?:동|읍|면))", tail.replace(" ", ""))
    if not m:
        return None
    area = collapse(m.group(1))
    for q in (f"경남 양산시 {area}", f"경상남도 양산시 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display_road
                cj, cr = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or hit.jibeon
                return hit
    return None


def resolve_geocode(road_full: str, name: str, key: str, display_road: str) -> GeoHit | None:
    for q in geocode_query_variants(road_full, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
        time.sleep(GEOCODE_DELAY)
    return area_fallback(road_full, key, display_road)


def cache_key(name: str, road: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{road}".encode()).hexdigest()[:28]


def row_has_trash_bag(row: tuple) -> bool:
    for i in range(5, 13):
        if i < len(row) and flag_y(row[i]):
            return True
    return False


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    ref_date = ref_date_from_path(inp)
    wb = load_workbook(inp, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    seen: set[str] = set()

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or len(row) < 3:
            continue
        seq = row[0]
        if seq is None or cell_str(seq) == "":
            continue
        name = cell_str(row[1])
        road_raw = cell_str(row[2])
        detail_raw = cell_str(row[3]) if len(row) > 3 else ""
        if not name or not road_raw:
            continue
        if not row_has_trash_bag(row):
            continue

        full_norm = normalize_yangsan_addr(road_raw, detail_raw)
        display_road = geocode_target(full_norm)
        if not display_road or "양산" not in display_road.replace(" ", ""):
            continue

        dk = f"{name}|{display_road}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(name, display_road)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else display_road
                jib = str(raw[3]) if len(raw) > 3 else ""
                if not jib or not road:
                    cj, cr = coord2address(lng, lat, key) if key else ("", "")
                    jib = jib or cj or display_road
                    road = road or cr or display_road
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jib)

        if hit is None and allow:
            hit = resolve_geocode(full_norm, name, key, display_road)
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {name}\t{display_road}", file=sys.stderr)
                continue
            cj, cr = coord2address(hit.lng, hit.lat, key)
            hit.road = display_road
            hit.jibeon = cj or hit.jibeon or display_road
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{seq}\n{name}\n{display_road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"gyeongnam-yangsan-trash-{rid}",
                "name": name,
                "lat": round(float(hit.lat), 7),
                "lng": round(float(hit.lng), 7),
                "roadAddress": display_road,
                "address": hit.jibeon or display_road,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": False,
                "adminVerified": True,
                "dataReferenceDate": ref_date,
            }
        )

    wb.close()
    save_cache(cache)
    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={ref_date}, geo≈{geo_n}, miss={misses})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["양산시"])


if __name__ == "__main__":
    main()
