#!/usr/bin/env python3
"""
강원특별자치도 원주시 종량제봉투 판매소 xlsx -> stores.gangwon-wonju-trash.json

시트 '봉투 판매소', 헤더 행: 법정동, 판매소명, 사업장주소
기준일: 파일/제목의 2026.1. 기준 → dataReferenceDate 2026-01-01

  python3 scripts/import_gangwon_wonju_trash_xlsx.py \\
    --xlsx ~/Downloads/cts523_file01_2026.xlsx
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path.home() / "Downloads" / "cts523_file01_2026.xlsx"
OUT_JSON = FRONTEND / "public" / "data" / "stores.gangwon-wonju-trash.json"
CACHE_PATH = Path(__file__).resolve().parent / "geocode-cache-gangwon-wonju-trash.json"
SHORT_CODE_RE = re.compile(r"^[a-zA-Z0-9]{6}$")
GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"


def load_kakao_key() -> str | None:
    k = os.environ.get("KAKAO_REST_API_KEY", "").strip()
    if k:
        return k
    p = FRONTEND / ".env.local"
    if not p.exists():
        return None
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("KAKAO_REST_API_KEY="):
            return line.split("=", 1)[1].strip().strip('"').strip("'") or None
    return None


def collapse(s: str) -> str:
    return " ".join((s or "").replace("\xa0", " ").split())


def normalize_addr(addr: str) -> str:
    a = collapse(addr)
    if not a:
        return a
    if a.startswith("강원특별자치도 원주시"):
        return a
    if a.startswith("강원도 원주시"):
        return "강원특별자치도 원주시" + a[len("강원도 원주시") :].strip()
    if a.startswith("원주시"):
        return "강원특별자치도 " + a
    return f"강원특별자치도 원주시 {a}"


def addr_variants(addr: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(v: str) -> None:
        t = collapse(v)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    add(addr)
    head = re.sub(r"\(.*?\)", "", addr).strip()
    add(collapse(head))
    add(collapse(re.sub(r"(\d+길)(\d+)$", r"\1 \2", head)))
    add(collapse(re.sub(r"(\d+번길)(\d+)$", r"\1 \2", head)))
    add(collapse(re.sub(r"([가-힣]+로)(\d)", r"\1 \2", head)))
    return out


def parse_float(v: str | None) -> float | None:
    try:
        return float((v or "").strip())
    except ValueError:
        return None


def kakao_geocode(q: str, key: str) -> tuple[float | None, float | None]:
    for base, extra in (
        (GEOCODE_URL, {"query": q}),
        (KEYWORD_URL, {"query": q, "size": "1"}),
    ):
        req = urllib.request.Request(
            f"{base}?{urllib.parse.urlencode(extra)}",
            headers={"Authorization": f"KakaoAK {key}"},
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except Exception:
            continue
        docs = data.get("documents") or []
        if not docs:
            continue
        lat = parse_float(docs[0].get("y"))
        lng = parse_float(docs[0].get("x"))
        if lat and lng:
            return lat, lng
    return None, None


def load_cache() -> dict[str, list[float]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        raw = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, list[float]] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if isinstance(v, list) and len(v) == 2:
                out[str(k)] = [float(v[0]), float(v[1])]
    return out


def save_cache(cache: dict[str, list[float]]) -> None:
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=0), encoding="utf-8")


def load_existing_meta(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    try:
        rows = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    out: dict[str, dict[str, str]] = {}
    if not isinstance(rows, list):
        return out
    for row in rows:
        if not isinstance(row, dict) or not row.get("id"):
            continue
        rid = str(row["id"])
        meta: dict[str, str] = {}
        sc = row.get("shortCode")
        if isinstance(sc, str) and SHORT_CODE_RE.match(sc.strip()):
            meta["shortCode"] = sc.strip()
        dr = row.get("dataReferenceDate")
        if isinstance(dr, str) and dr.strip():
            meta["dataReferenceDate"] = dr.strip()
        if meta:
            out[rid] = meta
    return out


def in_wonju_bbox(lat: float, lng: float) -> bool:
    return 37.20 <= lat <= 37.52 and 127.75 <= lng <= 128.10


def read_rows_from_xlsx(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "봉투 판매소" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"시트 '봉투 판매소' 없음: {wb.sheetnames}")
    ws = wb["봉투 판매소"]
    out: list[dict[str, str]] = []
    for tup in ws.iter_rows(values_only=True):
        row = list(tup)
        if len(row) < 3:
            continue
        dong, name, addr = row[0], row[1], row[2]
        if str(name or "").strip() == "판매소명":
            continue
        name_s = collapse(str(name or ""))
        addr_s = normalize_addr(str(addr or ""))
        if not name_s or not addr_s:
            continue
        out.append(
            {
                "dong": collapse(str(dong or "")),
                "name": name_s,
                "roadAddress": addr_s,
            }
        )
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="원주시 판매소 xlsx 경로",
    )
    args = ap.parse_args()
    xlsx_path: Path = args.xlsx

    if not xlsx_path.exists():
        raise SystemExit(f"xlsx 없음: {xlsx_path}")
    key = load_kakao_key()
    if not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요")

    raw_rows = read_rows_from_xlsx(xlsx_path)

    dedup: dict[str, dict] = {}
    for r in raw_rows:
        k = f"{r['name'].lower()}|{r['roadAddress'].lower()}"
        if k not in dedup:
            dedup[k] = r

    rows = sorted(dedup.values(), key=lambda x: (x["name"], x["roadAddress"]))
    cache = load_cache()
    existing_meta = load_existing_meta(OUT_JSON)
    ref = "2026-01-01"
    today = date.today().isoformat()

    out: list[dict] = []
    for i, r in enumerate(rows, start=1):
        sid = f"gangwon-wonju-trash-{i:04d}"
        ck = hashlib.sha1(
            (r["name"] + "\t" + r["roadAddress"]).encode("utf-8")
        ).hexdigest()[:20]
        if ck in cache:
            lat, lng = cache[ck]
        else:
            lat = lng = None
            for base in addr_variants(r["roadAddress"]):
                for q in (
                    base,
                    f"{r['name']} {base}",
                    f"원주시 {r['name']}",
                    f"강원 원주 {r['name']}",
                ):
                    glat, glng = kakao_geocode(q, key)
                    if glat is not None and glng is not None:
                        lat, lng = glat, glng
                        break
                if lat is not None and lng is not None:
                    break
            if lat is None or lng is None:
                continue
            if not in_wonju_bbox(lat, lng):
                continue
            cache[ck] = [lat, lng]
            if i % 40 == 0:
                save_cache(cache)
            time.sleep(0.08)

        if not in_wonju_bbox(float(lat), float(lng)):  # type: ignore[arg-type]
            continue

        rec = {
            "id": sid,
            "name": r["name"],
            "lat": round(float(lat), 7),
            "lng": round(float(lng), 7),
            "roadAddress": r["roadAddress"],
            "address": "강원특별자치도 원주시",
            "businessStatus": "영업",
            "hasTrashBag": True,
            "hasSpecialBag": False,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": existing_meta.get(sid, {}).get("dataReferenceDate", ref or today),
        }
        sc = existing_meta.get(sid, {}).get("shortCode")
        if sc:
            rec["shortCode"] = sc
        out.append(rec)

    save_cache(cache)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {len(out)} rows -> {OUT_JSON}")


if __name__ == "__main__":
    main()
