#!/usr/bin/env python3
"""
강원특별자치도 홍천군 종량제봉투 판매소 xlsx → stores.gangwon-hongcheon-trash.json

시트 Sheet1 (3행~): 연번 | 판매소명 | 입력 주소 | … | 대형폐기물 취급여부 (Y/N)

  python3 scripts/import_gangwon_hongcheon_trash_from_xlsx.py \\
    --input ~/Downloads/홍천군\\ 종량제봉투판매소현황.xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.gangwon-hongcheon-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-gangwon-hongcheon-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "홍천군 종량제봉투판매소현황.xlsx"
SHEET_NAME = "Sheet1"
HEADER_ROW = 2
DATA_START_ROW = 3
REF_DATE = date.today().isoformat()
CACHE_VERSION = "v1-hongcheon"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
MYEON_RE = re.compile(
    r"(홍천읍|화촌면|두촌면|내면|서석면|동면|남면|북방면|서면|영귀면|내촌면|갑봉산동)"
)


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (str(s or "")).replace("\xa0", " ")).strip()


def cell_str(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def flag_y(val: object) -> bool:
    return collapse(str(val or "")).upper() in ("Y", "YES", "예", "여", "O", "○")


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def in_hongcheon_bbox(lat: float, lng: float) -> bool:
    return 37.52 <= lat <= 38.22 and 127.42 <= lng <= 128.45


def hongcheon_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if any(x in t for x in ("충청", "경기", "서울", "인천", "부산", "대구")):
        return False
    if "홍천군" in t:
        return True
    return ("강원" in t or "강원특별" in t) and "홍천" in t


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^강원\s+", "강원특별자치도 ", a)
    a = re.sub(r"^강원도\s+", "강원특별자치도 ", a)
    if not a:
        return ""
    if a.startswith("강원특별자치도"):
        return a
    if a.startswith("홍천군"):
        return f"강원특별자치도 {a}"
    return f"강원특별자치도 홍천군 {a}"


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if "," in a:
        a = a.split(",")[0].strip()
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"홍천군(?=[가-힣])", "홍천군 ", a)
    a = re.sub(r"([읍면동리])([가-힣]{2,}(?:로|길|대로))", r"\1 \2", a)
    a = re.sub(r"(대로)(\d)", r"\1 \2", a)
    if not a.startswith("홍천") and not a.startswith("강원"):
        a = f"홍천군 {a}"
    return format_display_addr(a)


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_hongcheon_bbox(lat, lng):
        return None
    if not hongcheon_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_hongcheon_bbox(lat, lng):
        return None
    if not hongcheon_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    norm = normalize_addr(addr_raw)
    tail = norm
    for prefix in ("강원특별자치도 홍천군 ", "강원특별자치도 ", "홍천군 "):
        if tail.startswith(prefix):
            tail = collapse(tail[len(prefix) :])
            break
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"강원 홍천군 {tail}")
    push(f"강원특별자치도 홍천군 {tail}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"강원특별자치도 홍천군 {road_only}")
    push(f"{name} 홍천")
    push(f"{name} 홍천군")
    compact = re.sub(r"\s+", "", name)
    if compact != name:
        push(f"{compact} 홍천군")
    return out


def area_fallback(addr_raw: str, key: str, display: str) -> GeoHit | None:
    norm = normalize_addr(addr_raw)
    m = MYEON_RE.search(norm)
    if not m:
        return None
    area = m.group(1)
    for q in (f"강원특별자치도 홍천군 {area}", f"강원 홍천군 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                hit.road = display
                cj, cr = coord2address(hit.lng, hit.lat, key)
                hit.jibeon = cj or display
                return hit
    return None


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return area_fallback(addr_raw, key, normalize_addr(addr_raw))


@dataclass
class HongcheonRow:
    name: str
    addr_raw: str
    has_large_waste: bool


def iter_rows(path: Path) -> list[HongcheonRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    out: list[HongcheonRow] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        name = collapse(cell_str(ws.cell(r, 3).value))
        addr = collapse(cell_str(ws.cell(r, 4).value))
        if not name or not addr or name in ("판매소명", "업소명", "상호"):
            continue
        out.append(
            HongcheonRow(
                name=name,
                addr_raw=addr,
                has_large_waste=flag_y(ws.cell(r, 8).value),
            )
        )
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow_geo = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    rows = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    seen: set[str] = set()
    geo_n = 0
    misses = 0

    for row in rows:
        display = normalize_addr(row.addr_raw)
        dk = f"{row.name}|{display}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, row.addr_raw)
        if ck in cache and not args.refresh:
            lat, lng, road, jibeon = cache[ck]
            geo_n += 1
        elif allow_geo:
            hit = resolve_geocode(row.addr_raw, row.name, key)  # type: ignore[arg-type]
            if not hit:
                print(f"[지오코딩 실패] {row.name}\t{display}", file=sys.stderr)
                misses += 1
                continue
            lat, lng, road, jibeon = hit.lat, hit.lng, hit.road, hit.jibeon
            cache[ck] = [lat, lng, road, jibeon]
            geo_n += 1
        else:
            misses += 1
            continue

        if not in_hongcheon_bbox(lat, lng):
            print(f"[좌표 제외] {row.name}\t{display}", file=sys.stderr)
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{road}".encode()).hexdigest()[:20]
        out.append(
            {
                "id": f"gangwon-hongcheon-trash-{rid}",
                "name": row.name,
                "lat": round(lat, 7),
                "lng": round(lng, 7),
                "roadAddress": road,
                "address": jibeon,
                "businessStatus": "영업",
                "hasTrashBag": True,
                "hasSpecialBag": False,
                "hasLargeWasteSticker": row.has_large_waste,
                "adminVerified": True,
                "dataReferenceDate": REF_DATE,
            }
        )

    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    if allow_geo:
        save_cache(cache)

    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={REF_DATE}, geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["홍천군"])


if __name__ == "__main__":
    main()
