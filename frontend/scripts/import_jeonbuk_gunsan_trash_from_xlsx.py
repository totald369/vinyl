#!/usr/bin/env python3
"""
전북 군산시 종량제봉투·불연성마대 판매소 xlsx → stores.jeonbuk-gunsan-trash.json

  pip install openpyxl
  python3 scripts/import_jeonbuk_gunsan_trash_from_xlsx.py \\
    --input ~/Downloads/군산시\\ 종량제봉투\\ 판매소\\ 목록.xlsx
  python3 scripts/import_jeonbuk_gunsan_trash_from_xlsx.py --refresh

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonbuk-gunsan-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonbuk-gunsan-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "군산시 종량제봉투 판매소 목록.xlsx"
REF_DATE = "2026-05-21"
CACHE_VERSION = "v2"

# 카카오에 없는 읍면동 표기 보정 (엑셀 오타·구 명칭)
DONG_ALIASES: dict[str, str] = {
    "서수동": "수송동",
    "부곡동": "미룡동",
}

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.07

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
LOT_RE = re.compile(r"^(?P<prefix>.*?)(?P<main>\d+)(?:-(?P<sub>\d+))?\s*$")
ROAD_RE = re.compile(r"(로|길|대로)\s*(\d+)")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        for sep in ("NEXT_PUBLIC_", "NEXT_", "#"):
            if sep in v:
                v = v.split(sep, 1)[0].strip()
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def flag_o(val: object) -> bool:
    t = collapse(str(val or "")).upper()
    return t in ("O", "○", "Y", "YES", "예")


def in_gunsan_bbox(lat: float, lng: float) -> bool:
    return 35.75 <= lat <= 36.15 and 126.32 <= lng <= 127.05


def gunsan_in_text(blob: str) -> bool:
    return "군산" in (blob or "").replace(" ", "")


def format_display_addr(raw: str) -> str:
    """저장용 주소 — 전북(전라북도) 표기 통일."""
    a = collapse(raw)
    a = a.replace("전북특별치도", "전북특별자치도")
    a = re.sub(r"^전라북도\s+", "전북특별자치도 ", a)
    a = re.sub(r"^전북\s+", "전북특별자치도 ", a)
    if not a.startswith("전북") and "군산" in a.replace(" ", ""):
        if a.startswith("군산시"):
            return f"전북특별자치도 {a}"
        return f"전북특별자치도 군산시 {a}"
    return a


def dedupe_addr(addr: str) -> str:
    a = collapse(addr)
    a = re.sub(r"\s+KR(?:\s+\d+)?(?:\s+\d+동)?\s*$", "", a, flags=re.I)
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"\s+\d+\s*층\s*", " ", a)
    a = re.sub(r"\s+\d+\s*호\s*", " ", a)
    a = re.sub(r"\s+군산시\s+전라북도\s*$", "", a)
    a = re.sub(r"\s+전라북도\s*$", "", a)
    a = re.sub(r"\s+예스마트\s+\d+\s*$", "", a, flags=re.I)
    parts = a.split()
    mid = len(parts) // 2
    if mid >= 4 and parts[:mid] == parts[mid : mid + mid]:
        a = " ".join(parts[:mid])
    half = len(a) // 2
    if half > 12:
        left, right = a[:half].strip(), a[half:].strip()
        if right.startswith(left):
            return collapse(left)
    return collapse(a)


def normalize_gunsan_addr(addr: str) -> str:
    a = dedupe_addr(addr)
    a = a.replace("전북특별치도", "전북특별자치도")
    a = a.replace("전라북도", "전북특별자치도")
    a = re.sub(r"^전북\s+", "전북특별자치도 ", a)
    a = re.sub(r"(\d+)\s+(\d+)\s*$", r"\1-\2", a)
    compact = a.replace(" ", "")
    if "전북" in compact or "전라북" in compact:
        return collapse(a)
    if a.startswith("군산시"):
        return f"전북특별자치도 {a}"
    if "군산시" in compact or "군산" in compact:
        return f"전북특별자치도 {a}" if not a.startswith("전북") else a
    return f"전북특별자치도 군산시 {a}"


def gunsan_tail(addr: str) -> str:
    """카카오 API 쿼리용 — `전북 군산시 {tail}` 형태가 가장 잘 맞음."""
    norm = normalize_gunsan_addr(addr)
    for prefix in (
        "전북특별자치도 군산시 ",
        "전북특별자치도군산시",
        "전북특별자치도 ",
        "전북 군산시 ",
        "전라북도 군산시 ",
        "군산시 ",
    ):
        if norm.startswith(prefix):
            return collapse(norm[len(prefix) :])
    if norm.startswith("전북특별자치도"):
        return collapse(norm.replace("전북특별자치도", "", 1))
    return norm


def is_likely_jibeon(tail: str) -> bool:
    t = tail.replace(" ", "")
    if ROAD_RE.search(tail):
        return False
    return bool(re.search(r"\d", t))


def lot_query_variants(tail: str) -> list[str]:
    """지번-only 주소 — 동일 번지·인근 호수 후보."""
    out: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(tail)
    m = LOT_RE.match(tail)
    if m:
        prefix, main, sub = m.group("prefix"), m.group("main"), m.group("sub")
        subs: list[str | None] = []
        if sub:
            subs.extend([sub, "1", "2", "3", "4", "5"])
        else:
            subs.append(None)
        for s in subs:
            if s is None:
                push(f"{prefix}{main}")
            else:
                push(f"{prefix}{main}-{s}")
    return out


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    tail = gunsan_tail(addr_raw)
    out: list[str] = []
    seen: set[str] = set()

    def push_q(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    if is_likely_jibeon(tail):
        for tv in lot_query_variants(tail):
            push_q(f"전북 군산시 {tv}")
            push_q(f"전북특별자치도 군산시 {tv}")
    else:
        push_q(f"전북 군산시 {tail}")
        push_q(f"전북특별자치도 군산시 {tail}")
        if re.search(r"-\d+\s*$", tail):
            no_suffix = re.sub(r"-\d+\s*$", "", tail).strip()
            push_q(f"전북 군산시 {no_suffix}")

    norm = normalize_gunsan_addr(addr_raw)
    push_q(norm.replace("전북특별자치도", "전북"))
    push_q(f"{name} 군산시")
    push_q(f"{name} {tail}")
    return out


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    return jibeon, road


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_gunsan_bbox(lat, lng):
        return None
    if not gunsan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_gunsan_bbox(lat, lng):
        return None
    if not gunsan_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    return data.get("documents") or []


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return None


def extract_dong_area(tail: str) -> str | None:
    m = re.match(r"^(.+?(?:동|읍|면|리))", tail)
    if m:
        return collapse(m.group(1))
    m = re.match(r"^(.+?동)\d", tail.replace(" ", ""))
    if m:
        return collapse(m.group(1))
    return None


def dong_fallback(addr_raw: str, key: str) -> GeoHit | None:
    """지번·도로명 모두 실패 시 읍면동 중심 + coord2address."""
    tail = gunsan_tail(addr_raw)
    area = extract_dong_area(tail)
    if not area:
        return None
    area = DONG_ALIASES.get(area, area)
    for q in (f"전북 군산시 {area}", f"전북특별자치도 군산시 {area}"):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                orig = format_display_addr(normalize_gunsan_addr(addr_raw))
                hit.jibeon = orig
                if not hit.road or hit.road == hit.jibeon:
                    cj, cr = coord2address(hit.lng, hit.lat, key)
                    hit.road = cr or cj or hit.road
                return hit
    return None


def road_strip_fallback(addr_raw: str, key: str) -> GeoHit | None:
    """도로명+번지가 DB에 없을 때 도로명만으로 근사 좌표."""
    tail = gunsan_tail(addr_raw)
    if is_likely_jibeon(tail):
        return None
    orig = format_display_addr(normalize_gunsan_addr(addr_raw))
    candidates: list[str] = []
    seen: set[str] = set()

    def push(s: str) -> None:
        t = collapse(s)
        if t and t not in seen:
            seen.add(t)
            candidates.append(t)

    push(re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail))
    push(re.sub(r"-\d+\s*$", "", tail))
    m = ROAD_RE.search(tail)
    if m:
        push(tail[: m.end()].strip())

    for road_only in candidates:
        if not road_only or road_only == tail:
            continue
        for q in (f"전북 군산시 {road_only}", f"전북특별자치도 군산시 {road_only}"):
            for d in kakao_get(GEOCODE_URL, q, key):
                hit = parse_address_doc(d, key)
                if hit:
                    hit.jibeon = orig
                    hit.road = orig if ROAD_RE.search(orig.replace(" ", "")) else hit.road
                    cj, cr = coord2address(hit.lng, hit.lat, key)
                    if not hit.road or hit.road == hit.jibeon:
                        hit.road = cr or orig
                    if not hit.jibeon:
                        hit.jibeon = cj or orig
                    return hit
    return None


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if not row:
            continue
        texts = [(str(x).strip() if x is not None else "") for x in row]
        if "연번" in texts and "상호" in texts and "주소" in texts:
            col: dict[str, int] = {}
            for j, h in enumerate(texts):
                if h in ("연번", "상호", "주소", "종량제봉투", "불연성 마대", "비고"):
                    col[h] = j
            if {"연번", "상호", "주소"}.issubset(col):
                return i, col
    raise SystemExit("헤더 행(연번·상호·주소)을 찾지 못했습니다.")


def cache_key(seq: object, name: str, addr: str) -> str:
    h = hashlib.sha1(f"{CACHE_VERSION}:{seq}:{name}:{addr}".encode()).hexdigest()[:28]
    return h


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true", help="지오코딩 캐시 무시 후 재조회")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        print(
            "오류: KAKAO_REST_API_KEY 가 없습니다. frontend/.env.local 을 확인하거나 --skip-kakao",
            file=sys.stderr,
        )
        raise SystemExit(1)

    wb = load_workbook(inp, read_only=True, data_only=True)
    ws = wb.active
    hdr_row, col = find_header_row(ws)
    ii, iname, iaddr = col["연번"], col["상호"], col["주소"]
    itrash = col.get("종량제봉투", -1)
    ispecial = col.get("불연성 마대", -1)

    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    geo_n = 0
    misses = 0
    fallback_n = 0
    seen_keys: set[str] = set()

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row:
            continue
        name = collapse(str(row[iname] or ""))
        addr_raw = collapse(str(row[iaddr] or ""))
        if not name or not addr_raw:
            continue
        seq = row[ii]
        if seq is None or str(seq).strip() == "":
            continue

        dedupe_key = f"{name}|{addr_raw}"
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)

        norm_addr = normalize_gunsan_addr(addr_raw)
        has_trash = flag_o(row[itrash]) if itrash >= 0 else True
        has_special = flag_o(row[ispecial]) if ispecial >= 0 else False
        if not has_trash and not has_special:
            continue

        ck = cache_key(seq, name, norm_addr)
        hit: GeoHit | None = None

        if not args.refresh and ck in cache:
            raw = cache[ck]
            if isinstance(raw, list) and len(raw) >= 2:
                lat, lng = float(raw[0]), float(raw[1])
                road = str(raw[2]) if len(raw) > 2 else ""
                jibeon = str(raw[3]) if len(raw) > 3 else ""
                if not road or not jibeon:
                    cj, cr = coord2address(lng, lat, key)
                    jibeon = jibeon or cj or norm_addr
                    road = road or cr or cj or jibeon
                hit = GeoHit(lat=lat, lng=lng, road=road, jibeon=jibeon)

        if hit is None and allow:
            hit = resolve_geocode(addr_raw, name, key)
            if hit is None:
                hit = road_strip_fallback(addr_raw, key)
                if hit:
                    fallback_n += 1
            if hit is None:
                hit = dong_fallback(addr_raw, key)
                if hit:
                    fallback_n += 1
            if hit is None:
                misses += 1
                print(f"[geocode 실패] {name}\t{norm_addr}", file=sys.stderr)
                continue
            cache[ck] = [hit.lat, hit.lng, hit.road, hit.jibeon]
            geo_n += 1
            if geo_n % 50 == 0:
                save_cache(cache)
                print(f"[geocode] {geo_n} …", file=sys.stderr)
            time.sleep(GEOCODE_DELAY)
        elif hit is None:
            misses += 1
            continue

        rid = hashlib.sha1(f"{seq}\n{name}\n{norm_addr}".encode()).hexdigest()[:20]
        rec = {
            "id": f"jeonbuk-gunsan-trash-{rid}",
            "name": name,
            "lat": round(float(hit.lat), 7),
            "lng": round(float(hit.lng), 7),
            "roadAddress": hit.road,
            "address": hit.jibeon,
            "businessStatus": "영업",
            "hasTrashBag": has_trash,
            "hasSpecialBag": has_special,
            "hasLargeWasteSticker": False,
            "dataReferenceDate": REF_DATE,
        }
        out.append(rec)

    wb.close()
    save_cache(cache)

    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref_date={REF_DATE}, api_rows≈{geo_n}, dong_fallback≈{fallback_n}, 미매칭={misses})"
    )


if __name__ == "__main__":
    main()
