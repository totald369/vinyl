#!/usr/bin/env python3
"""
구로구 불연성마대 판매처 엑셀 → public/data/stores.guro-noncombust.json

- 엑셀: 행정동 / 주민센터 / 주소 / 대표연락처 형식(Sheet1, 4번째 줄부터 헤더)
- 주소로 Nominatim 지오코딩 → 위·경도
- 기존 JSON이 있으면 같은 id(guro-nc-001 …)의 shortCode 유지

사용법:
  python3 scripts/importGuroNoncombustFromXlsx.py /path/to/file.xlsx
  npm run shortcodes:assign   # 새 행만 shortCode 보정 시
"""

from __future__ import annotations

import argparse
import json
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

SHORT_CODE_RE = re.compile(r"^[a-zA-Z0-9]{6}$")

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as e:
    raise SystemExit("openpyxl이 필요합니다. pip install openpyxl") from e


def full_road_address(raw: str) -> str:
    a = str(raw).replace("\xa0", " ").strip()
    a = " ".join(a.split())
    if a.startswith("서울 구로구"):
        return "서울특별시 구로구 " + a[6:].strip()
    if a.startswith("서울특별시 구로구"):
        return a
    if a.startswith("구로구"):
        return "서울특별시 " + a
    return "서울특별시 구로구 " + a


def read_excel_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    out: list[dict] = []
    # 헤더: row index 3 (행정동, 주민센터, 주소…)
    for r in rows[4:]:
        if not r:
            continue
        dong = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        center = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        addr = str(r[3]).strip() if len(r) > 3 and r[3] else ""
        phone = str(r[4]).strip() if len(r) > 4 and r[4] else ""
        if not addr or not phone:
            continue
        fa = full_road_address(addr)
        name = center or f"{dong} 주민센터"
        out.append(
            {"dong": dong, "name": name, "roadAddress": fa, "phone": phone}
        )
    return out


UA = "SseubongmapGuroImport/2.0 (internal data tooling)"


def nominatim(q: str) -> tuple[float | None, float | None]:
    params = urllib.parse.urlencode({"q": q, "format": "json", "limit": "1"})
    url = "https://nominatim.openstreetmap.org/search?" + params
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=45) as resp:
        data = json.loads(resp.read().decode())
    time.sleep(1.12)
    if data:
        return float(data[0]["lat"]), float(data[0]["lon"])
    return None, None


def geocode_row(p: dict) -> tuple[float, float]:
    for q in (
        p["roadAddress"],
        f"{p['name']} {p['roadAddress']}",
        p["name"] + " 서울특별시 구로구",
    ):
        lat, lng = nominatim(q)
        if lat is not None and lng is not None:
            return lat, lng
    raise RuntimeError(f"지오코딩 실패: {json.dumps(p, ensure_ascii=False)}")


def load_existing_by_id(existing_path: Path) -> dict[str, dict]:
    """id → 기존 shortCode, dataReferenceDate (재실행 시 보존)"""
    if not existing_path.exists():
        return {}
    try:
        data = json.loads(existing_path.read_text(encoding="utf8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(data, list):
        return {}
    out: dict[str, dict] = {}
    for row in data:
        if not isinstance(row, dict) or not row.get("id"):
            continue
        rid = str(row["id"])
        meta: dict = {}
        sc = row.get("shortCode")
        if isinstance(sc, str):
            sc = sc.strip()
            if SHORT_CODE_RE.match(sc):
                meta["shortCode"] = sc
        dr = row.get("dataReferenceDate")
        if isinstance(dr, str) and dr.strip():
            meta["dataReferenceDate"] = dr.strip()
        if meta:
            out[rid] = meta
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "xlsx",
        type=Path,
        help="불연성포대 판매처 정보.xlsx 경로",
    )
    ap.add_argument(
        "-o",
        "--out",
        type=Path,
        default=Path("public/data/stores.guro-noncombust.json"),
        help="출력 JSON (프로젝트 frontend 디렉터리 기준)",
    )
    args = ap.parse_args()

    parsed = read_excel_rows(args.xlsx)
    if not parsed:
        raise SystemExit("엑셀에서 유효한 데이터 행이 없습니다.")

    out_abs = (
        args.out
        if args.out.is_absolute()
        else Path(__file__).resolve().parent.parent / args.out
    )
    existing_meta = load_existing_by_id(out_abs)

    iso_today = date.today().isoformat()
    bundle: list[dict] = []

    for i, p in enumerate(parsed, start=1):
        sid = f"guro-nc-{i:03d}"
        lat, lng = geocode_row(p)
        row: dict = {
            "id": sid,
            "name": p["name"],
            "lat": round(lat, 7),
            "lng": round(lng, 7),
            "roadAddress": p["roadAddress"],
            "address": f"서울특별시 구로구 {p['dong']}",
            "phone": p["phone"],
            "businessStatus": "영업",
            "hasTrashBag": False,
            "hasSpecialBag": True,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": existing_meta.get(sid, {}).get(
                "dataReferenceDate", iso_today
            ),
            "storeCategory": "nonBurnable",
        }
        sc = existing_meta.get(sid, {}).get("shortCode")
        if sc:
            row["shortCode"] = sc
        bundle.append(row)

    out_abs.parent.mkdir(parents=True, exist_ok=True)
    out_abs.write_text(
        json.dumps(bundle, ensure_ascii=False, indent=2) + "\n", encoding="utf8"
    )
    print(f"wrote {len(bundle)} rows -> {out_abs}")


if __name__ == "__main__":
    main()
