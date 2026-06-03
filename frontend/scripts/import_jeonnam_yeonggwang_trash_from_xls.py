#!/usr/bin/env python3
"""
전라남도 영광군 종량제봉투·특수마대(불연성 마대) 판매소 xls/xlsx → stores.jeonnam-yeonggwang-trash.json

※ Fasoo DRM 암호화 파일은 읽을 수 없습니다. Excel에서 열어
  「다른 이름으로 저장」(.xlsx / .csv) 후 --input 으로 지정하세요.

  python3 scripts/import_jeonnam_yeonggwang_trash_from_xls.py \\
    --input ~/Downloads/영광군_정보공개자료\\(판매소현황\\).xlsx

KAKAO_REST_API_KEY: frontend/.env.local
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date
from pathlib import Path

FRONTEND = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
OUT_JSON = FRONTEND / "public" / "data" / "stores.jeonnam-yeonggwang-trash.json"
CACHE_PATH = SCRIPT_DIR / "geocode-cache-jeonnam-yeonggwang-trash.json"
DL = Path.home() / "Downloads"
DEFAULT_INPUT = DL / "영광군_정보공개자료(판매소현황).xls"
REF_DATE = date.today().isoformat()
CACHE_VERSION = "v1-yeonggwang"

GEOCODE_URL = "https://dapi.kakao.com/v2/local/search/address.json"
KEYWORD_URL = "https://dapi.kakao.com/v2/local/search/keyword.json"
COORD2_URL = "https://dapi.kakao.com/v2/local/geo/coord2address.json"
GEOCODE_DELAY = 0.06

WS_RE = re.compile(r"[ \t\r\n\v\f]+")
_REF_DATE_CELL = re.compile(r"(\d{4})\s*[\.\-]\s*(\d{1,2})\s*[\.\-]?\s*(\d{1,2})?")


def _load_dotenv_local() -> None:
    p = FRONTEND / ".env.local"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv_local()


def collapse(s: str) -> str:
    return WS_RE.sub(" ", (s or "").replace("\xa0", " ")).strip()


def parse_float(v: object) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


def load_kakao_key() -> str | None:
    return (
        os.environ.get("KAKAO_REST_API_KEY", "").strip()
        or os.environ.get("KAKAO_REST_KEY", "").strip()
    )


def assert_not_drm(path: Path) -> None:
    head = path.read_bytes()[:128]
    if b"DRMONE" in head or b"Fasoo DRM" in head:
        raise SystemExit(
            "입력 파일이 Fasoo DRM으로 암호화되어 있습니다.\n"
            "Excel(또는 DRM 뷰어)에서 문서를 연 뒤\n"
            "  · 다른 이름으로 저장 → .xlsx 또는 .csv\n"
            "후 --input 경로를 지정해 다시 실행해 주세요."
        )


def ref_date_from_cells(cells: list[list[str]]) -> str:
    for row in cells[:6]:
        for cell in row:
            m = _REF_DATE_CELL.search(cell)
            if m:
                y, mo, d = m.group(1), m.group(2), m.group(3) or "1"
                return f"{y}-{int(mo):02d}-{int(d):02d}"
    return REF_DATE


def in_yeonggwang_bbox(lat: float, lng: float) -> bool:
    return 35.1 <= lat <= 35.5 and 126.3 <= lng <= 126.8


def yeonggwang_in_text(blob: str) -> bool:
    t = (blob or "").replace(" ", "")
    if "남양주" in t or "부산" in t or "충주" in t:
        return False
    return "영광군" in t or ("전라남도" in t and "영광" in t)


def format_display_addr(raw: str) -> str:
    a = collapse(raw)
    a = re.sub(r"^전남\s+", "전라남도 ", a)
    if not a:
        return ""
    if a.startswith("전라남도"):
        return a
    if a.startswith("영광군"):
        return f"전라남도 {a}"
    if re.match(r"^영광(읍|면)", a):
        return f"전라남도 영광군 {a}"
    return f"전라남도 영광군 {a}"


def normalize_addr(addr_raw: str) -> str:
    a = collapse(addr_raw)
    if "," in a:
        a = a.split(",")[0].strip()
    a = re.sub(r"\s*\([^)]*\)\s*", " ", a)
    a = re.sub(r"\s*번지\s*", " ", a)
    a = re.sub(r"\s+\d+\s*층\s*", " ", a)
    a = re.sub(r"\s+\d+\s*호.*$", "", a)
    m = re.match(r"^(.+?[로길대로])(\d[\d\-]*)$", a.replace(" ", ""))
    if m:
        a = f"{m.group(1)} {m.group(2)}"
    return format_display_addr(a)


def trash_ox(v: object) -> bool:
    s = str(v or "").strip().upper()
    return s in ("O", "Y", "예", "유")


def special_ox(v: object) -> bool:
    s = str(v or "").strip().upper()
    if s in ("X", "N", "해당없음", "없음", "-", ""):
        return False
    return s in ("O", "Y", "예", "유")


def normalize_phone(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v == int(v):
        s = str(int(v))
    else:
        s = collapse(str(v))
    return re.sub(r"[^\d\-]", "", s)


@dataclass
class ColMap:
    name: int
    addr: int
    phone: int | None
    trash: int | None
    special: int | None
    header_row: int


def detect_columns(header_cells: list[str]) -> ColMap | None:
    joined = " ".join(header_cells)

    def find_idx(patterns: list[str]) -> int | None:
        for i, h in enumerate(header_cells):
            h2 = collapse(h).replace("\n", "")
            for p in patterns:
                if re.search(p, h2, re.I):
                    return i
        return None

    name_i = find_idx([r"상호", r"판매소.*명", r"업체.*명", r"명칭"])
    addr_i = find_idx([r"주소", r"소재지"])
    if name_i is None or addr_i is None:
        return None
    phone_i = find_idx([r"전화"])
    trash_i = find_idx([r"종량제"])
    special_i = find_idx([r"특수", r"불연성"])
    return ColMap(
        name=name_i,
        addr=addr_i,
        phone=phone_i,
        trash=trash_i,
        special=special_i,
        header_row=0,
    )


def load_cells(path: Path) -> list[list[str]]:
    assert_not_drm(path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                text = path.read_text(encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = path.read_text(encoding="cp949", errors="replace")
        reader = csv.reader(text.splitlines())
        return [[collapse(c) for c in row] for row in reader]

    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        ws = load_workbook(path, data_only=True).worksheets[0]
        return [
            [collapse(str(ws.cell(r, c).value or "")) for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)
        ]

    try:
        import xlrd
    except ImportError as e:
        raise SystemExit("xlrd 필요: pip install 'xlrd==1.2.0'") from e

    sh = xlrd.open_workbook(str(path)).sheet_by_index(0)
    return [
        [collapse(str(sh.cell_value(r, c))) for c in range(sh.ncols)]
        for r in range(sh.nrows)
    ]


@dataclass
class YeonggwangRow:
    name: str
    addr_raw: str
    phone: str
    has_trash: bool
    has_special: bool


def iter_rows(path: Path) -> tuple[list[YeonggwangRow], str]:
    cells = load_cells(path)
    ref = ref_date_from_cells(cells)

    colmap: ColMap | None = None
    data_start = 0
    for r, row in enumerate(cells[:20]):
        cm = detect_columns(row)
        if cm:
            colmap = cm
            colmap.header_row = r
            data_start = r + 1
            break

    if colmap is None:
        # 여수형 고정 열 (연번|이름|주소|전화|종량제|특수|비고)
        colmap = ColMap(name=2, addr=3, phone=4, trash=5, special=6, header_row=2)
        data_start = 3

    out: list[YeonggwangRow] = []
    for row in cells[data_start:]:
        if len(row) <= max(colmap.name, colmap.addr):
            continue
        name = collapse(row[colmap.name])
        addr = collapse(row[colmap.addr])
        if not name or not addr or name in ("판매소명", "상호명"):
            continue
        phone = normalize_phone(row[colmap.phone]) if colmap.phone is not None and colmap.phone < len(row) else ""
        has_trash = (
            trash_ox(row[colmap.trash])
            if colmap.trash is not None and colmap.trash < len(row)
            else True
        )
        has_special = (
            special_ox(row[colmap.special])
            if colmap.special is not None and colmap.special < len(row)
            else False
        )
        if not has_trash and not has_special:
            continue
        out.append(
            YeonggwangRow(
                name=name,
                addr_raw=addr,
                phone=phone,
                has_trash=has_trash,
                has_special=has_special,
            )
        )
    return out, ref


@dataclass
class GeoHit:
    lat: float
    lng: float
    road: str
    jibeon: str


def load_cache() -> dict[str, list]:
    if not CACHE_PATH.is_file():
        return {}
    try:
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(c: dict[str, list]) -> None:
    CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")


def cache_key(name: str, addr: str) -> str:
    return hashlib.sha1(f"{CACHE_VERSION}:{name}:{addr}".encode()).hexdigest()[:28]


def kakao_get(url: str, query: str, key: str) -> list[dict]:
    req = urllib.request.Request(
        f"{url}?{urllib.parse.urlencode({'query': query, 'size': '15'})}",
        headers={"Authorization": f"KakaoAK {key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return []
    time.sleep(GEOCODE_DELAY)
    return data.get("documents") or []


def coord2address(lng: float, lat: float, kakao_key: str) -> tuple[str, str]:
    req = urllib.request.Request(
        f"{COORD2_URL}?{urllib.parse.urlencode({'x': lng, 'y': lat, 'input_coord': 'WGS84'})}",
        headers={"Authorization": f"KakaoAK {kakao_key}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return "", ""
    docs = data.get("documents") or []
    if not docs:
        return "", ""
    d = docs[0]
    jibeon = format_display_addr(str((d.get("address") or {}).get("address_name") or ""))
    road = format_display_addr(str((d.get("road_address") or {}).get("address_name") or ""))
    time.sleep(GEOCODE_DELAY)
    return jibeon, road


def _doc_blob(d: dict) -> str:
    parts: list[str] = []

    def touch(v: object) -> None:
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
        elif isinstance(v, dict):
            for sk in ("address_name", "region_1depth_name", "region_2depth_name"):
                s = v.get(sk)
                if isinstance(s, str) and s.strip():
                    parts.append(s.strip())

    touch(d.get("address_name"))
    touch(d.get("road_address"))
    touch(d.get("road_address_name"))
    touch(d.get("place_name"))
    return " ".join(parts)


def parse_address_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_yeonggwang_bbox(lat, lng):
        return None
    if not yeonggwang_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    ra = d.get("road_address")
    road = ""
    if isinstance(ra, dict):
        road = format_display_addr(str(ra.get("address_name") or ""))
    if not road:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = cr or cj
        if not jibeon:
            jibeon = cj
    if not jibeon:
        jibeon = format_display_addr(str(d.get("address_name") or ""))
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def parse_keyword_doc(d: dict, kakao_key: str) -> GeoHit | None:
    lat = parse_float(d.get("y"))
    lng = parse_float(d.get("x"))
    if lat is None or lng is None or not in_yeonggwang_bbox(lat, lng):
        return None
    if not yeonggwang_in_text(_doc_blob(d)):
        return None
    jibeon = format_display_addr(str(d.get("address_name") or ""))
    road = format_display_addr(str(d.get("road_address_name") or ""))
    if not road or not jibeon:
        cj, cr = coord2address(lng, lat, kakao_key)
        road = road or cr
        jibeon = jibeon or cj
    if not road and not jibeon:
        return None
    return GeoHit(lat=lat, lng=lng, road=road or jibeon, jibeon=jibeon or road)


def geocode_query_variants(addr_raw: str, name: str) -> list[str]:
    norm = normalize_addr(addr_raw)
    tail = norm
    for prefix in ("전라남도 영광군 ", "전라남도 ", "영광군 "):
        if tail.startswith(prefix):
            tail = collapse(tail[len(prefix) :])
            break
    out: list[str] = []
    seen: set[str] = set()

    def push(q: str) -> None:
        t = collapse(q)
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    push(norm)
    push(f"전남 영광군 {tail}")
    push(f"전라남도 영광군 {tail}")
    road_only = re.sub(r"\s+\d+(?:-\d+)?\s*$", "", tail).strip()
    if road_only and road_only != tail:
        push(f"전라남도 영광군 {road_only}")
    push(f"{name} 영광")
    push(f"{name} 영광군")
    return out


def resolve_geocode(addr_raw: str, name: str, key: str) -> GeoHit | None:
    for q in geocode_query_variants(addr_raw, name):
        for d in kakao_get(GEOCODE_URL, q, key):
            hit = parse_address_doc(d, key)
            if hit:
                return hit
        for d in kakao_get(KEYWORD_URL, q, key):
            hit = parse_keyword_doc(d, key)
            if hit:
                return hit
    return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--skip-kakao", action="store_true")
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--skip-activity", action="store_true")
    args = ap.parse_args()
    inp = args.input.expanduser()
    if not inp.is_file():
        raise SystemExit(f"파일 없음: {inp}")

    key = load_kakao_key()
    allow_geo = not args.skip_kakao and bool(key)
    if not args.skip_kakao and not key:
        raise SystemExit("KAKAO_REST_API_KEY 필요 (frontend/.env.local)")

    rows, path_ref = iter_rows(inp)
    cache = {} if args.refresh else load_cache()
    out: list[dict] = []
    seen: set[str] = set()
    geo_n = 0
    misses = 0

    for row in rows:
        display = normalize_addr(row.addr_raw)
        dk = f"{row.name}|{display}"
        if dk in seen:
            continue
        seen.add(dk)

        ck = cache_key(row.name, row.addr_raw)
        if ck in cache and not args.refresh:
            lat, lng, road, jibeon = cache[ck]
            geo_n += 1
        elif allow_geo:
            hit = resolve_geocode(row.addr_raw, row.name, key)  # type: ignore[arg-type]
            if not hit:
                print(f"[지오코딩 실패] {row.name}\t{display}", file=sys.stderr)
                misses += 1
                continue
            lat, lng, road, jibeon = hit.lat, hit.lng, hit.road, hit.jibeon
            cache[ck] = [lat, lng, road, jibeon]
            geo_n += 1
        else:
            misses += 1
            continue

        if not in_yeonggwang_bbox(lat, lng):
            print(f"[좌표 제외] {row.name}\t{display}", file=sys.stderr)
            misses += 1
            continue

        rid = hashlib.sha1(f"{row.name}\n{road}".encode()).hexdigest()[:20]
        item: dict = {
            "id": f"jeonnam-yeonggwang-trash-{rid}",
            "name": row.name,
            "lat": round(lat, 7),
            "lng": round(lng, 7),
            "roadAddress": road,
            "address": jibeon,
            "businessStatus": "영업",
            "hasTrashBag": row.has_trash,
            "hasSpecialBag": row.has_special,
            "hasLargeWasteSticker": False,
            "adminVerified": True,
            "dataReferenceDate": path_ref,
        }
        if row.phone:
            item["phone"] = row.phone
        out.append(item)

    out.sort(key=lambda x: (x["name"], x["roadAddress"]))
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    if allow_geo:
        save_cache(cache)

    trash_n = sum(1 for x in out if x["hasTrashBag"])
    special_n = sum(1 for x in out if x["hasSpecialBag"])
    print(
        f"wrote {len(out)} → {OUT_JSON} "
        f"(ref={path_ref}, 종량제 {trash_n}, 특수마대 {special_n}, "
        f"geo={geo_n}, miss={misses}, src={len(rows)})"
    )

    if out and not args.skip_activity:
        sys.path.insert(0, str(SCRIPT_DIR))
        from append_activity import record_region_data_added

        record_region_data_added(["영광군"])


if __name__ == "__main__":
    main()
